"""
Debug script to test Excel import functionality
Run this to diagnose import issues
"""

import os
import sys
from datetime import datetime, timedelta

# Add app directory to path
sys.path.insert(0, os.path.dirname(__file__))

try:
    from openpyxl import load_workbook
    print("✓ openpyxl imported successfully")
except ImportError:
    print("✗ openpyxl NOT installed")
    sys.exit(1)

from database.connection import get_session
from database.models import Member, Loan, MemberStatus, LoanStatus

print("\n" + "="*60)
print("DEBUG IMPORT SCRIPT")
print("="*60 + "\n")

# Ask user for file path
file_path = input("Enter the path to your Excel file: ").strip()

if not os.path.exists(file_path):
    print(f"✗ File not found: {file_path}")
    sys.exit(1)

print(f"✓ File found: {file_path}")

# Try to open and read the file
try:
    print("\n[1] Loading workbook...")
    workbook = load_workbook(file_path)
    sheet = workbook.active
    print(f"✓ Workbook loaded. Sheet name: {sheet.title}")
    print(f"✓ Max row: {sheet.max_row}, Max column: {sheet.max_column}")
    
    # Print first few rows to see structure
    print("\n[2] Reading file structure (first 6 rows):")
    for idx, row in enumerate(sheet.iter_rows(min_row=1, max_row=6, values_only=True), 1):
        print(f"  Row {idx}: {row}")
    
    # Now try the actual import
    print("\n[3] Starting import process...")
    session = get_session()
    
    imported_count = 0
    error_count = 0
    
    print("\n[4] Processing data rows (starting from row 5):")
    for row_idx, row in enumerate(sheet.iter_rows(min_row=5, values_only=True), start=5):
        try:
            if row is None or len(row) < 10:
                print(f"  Row {row_idx}: Skipped (empty or incomplete)")
                continue
            
            # Extract values
            s_n = row[0]
            ippis = row[1]
            loan_type = row[2]
            full_name = row[3]
            loan_amount = row[4]
            batch_number = row[5]
            cheque_no = row[6]
            loan_duration = row[7]
            loan_issue_date = row[8]
            interest_str = row[9]
            
            # Skip empty rows
            if not ippis or not full_name or loan_amount is None:
                print(f"  Row {row_idx}: Skipped (missing critical data)")
                continue
            
            print(f"\n  Row {row_idx}: Processing")
            print(f"    IPPIS: {ippis}")
            print(f"    Name: {full_name}")
            print(f"    Amount: {loan_amount}")
            print(f"    Duration: {loan_duration} months")
            print(f"    Interest: {interest_str}%")
            print(f"    Date: {loan_issue_date}")
            
            # Convert IPPIS to string
            ippis = str(int(ippis)) if isinstance(ippis, (int, float)) else str(ippis).strip()
            print(f"    → IPPIS converted to: {ippis}")
            
            # Get or create member
            member = session.query(Member).filter(Member.ippis_number == ippis).first()
            if not member:
                print(f"    → Creating new member...")
                member = Member(
                    ippis_number=ippis,
                    name=str(full_name).strip(),
                    status=MemberStatus.ACTIVE
                )
                session.add(member)
                session.flush()
                print(f"    → Member created with ID: {member.id}")
            else:
                print(f"    → Member exists with ID: {member.id}")
            
            # Parse loan amount
            loan_amt = float(loan_amount) if isinstance(loan_amount, (int, float)) else float(str(loan_amount).replace(',', ''))
            print(f"    → Loan amount: {loan_amt}")
            
            # Parse duration
            duration = int(loan_duration) if loan_duration else 0
            print(f"    → Duration: {duration}")
            
            # Parse interest rate
            if isinstance(interest_str, str):
                interest_rate = float(interest_str.strip().rstrip('%'))
            else:
                interest_rate = float(interest_str) if interest_str else 0
            
            # If interest rate is less than 1, assume it's a decimal and convert to percentage
            if interest_rate < 1:
                interest_rate = interest_rate * 100
                print(f"    → Interest converted from decimal {interest_str} to percentage: {interest_rate}%")
            else:
                print(f"    → Interest rate: {interest_rate}%")
            
            # Parse date
            start_date = datetime.now()
            if loan_issue_date:
                if isinstance(loan_issue_date, datetime):
                    start_date = loan_issue_date
                    print(f"    → Date (from datetime): {start_date}")
                elif isinstance(loan_issue_date, str):
                    for fmt in ("%d-%b", "%d-%B", "%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y"):
                        try:
                            if fmt == "%d-%b" or fmt == "%d-%B":
                                start_date = datetime.strptime(f"{loan_issue_date}-{datetime.now().year}", fmt + "-%Y")
                            else:
                                start_date = datetime.strptime(loan_issue_date, fmt)
                            print(f"    → Date parsed as: {start_date} (format: {fmt})")
                            break
                        except ValueError:
                            continue
            
            # Calculate end date and interest
            end_date = start_date + timedelta(days=30 * duration) if duration > 0 else start_date
            total_interest = loan_amt * (interest_rate / 100)
            print(f"    → End date: {end_date}")
            print(f"    → Total interest: {total_interest}")
            
            # Create loan
            new_loan = Loan(
                member_id=member.id,
                amount=loan_amt,
                interest_rate=interest_rate,
                total_interest=total_interest,
                amount_repaid=0.0,
                start_date=start_date,
                end_date=end_date,
                status=LoanStatus.ACTIVE,
                batch_number=str(batch_number).strip() if batch_number else None,
                cheque_number=str(cheque_no).strip() if cheque_no else None,
                is_member=True,
            )
            session.add(new_loan)
            imported_count += 1
            print(f"    ✓ Loan added to session")
            
        except Exception as e:
            error_count += 1
            print(f"  Row {row_idx}: ✗ ERROR: {str(e)}")
            import traceback
            traceback.print_exc()
            continue
    
    # Commit
    print(f"\n[5] Committing to database...")
    session.commit()
    session.close()
    workbook.close()
    
    print(f"\n" + "="*60)
    print(f"IMPORT COMPLETE")
    print(f"="*60)
    print(f"✓ Successfully imported: {imported_count}")
    print(f"✗ Errors encountered: {error_count}")
    print(f"="*60 + "\n")
    
except Exception as e:
    print(f"✗ FATAL ERROR: {str(e)}")
    import traceback
    traceback.print_exc()
    sys.exit(1)
