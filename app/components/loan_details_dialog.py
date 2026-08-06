import flet as ft
from database.connection import (
    get_repayments_by_loan,
    get_refunds_by_loan,
    get_topups_by_loan,
    get_penalties_by_loan,
    get_member_by_id,
    get_non_member_by_id,
    process_refund,
    update_loan,
    record_loan_penalty_change,
    calculate_interest,
    update_single_loan_overdue_interest,
    get_loan_by_id,
    get_loan_total_due_amount,
    get_loan_balance_amount,
)
from database.models import LoanStatus
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from components.ui_components import ToastNotification, NotificationType
from datetime import datetime
import threading
import os
import flet as ft
import os
import threading


def export_loan_to_excel(loan, borrower_name, repayments, refunds, topups, penalties, page, status_container):
    """
    Export loan details to Excel file
    """
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Loan Details"
        
        # Define styles
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=14)
        title_font = Font(bold=True, size=16, color="1F4E78")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Title
        ws['A1'] = f"LOAN DETAILS - {borrower_name}"
        ws['A1'].font = title_font
        ws.merge_cells('A1:D1')
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # Borrower Info Section
        row = 3
        ws[f'A{row}'] = "BORROWER INFORMATION"
        ws[f'A{row}'].font = Font(bold=True, size=13)
        row += 1
        
        ws[f'A{row}'] = "Name:"
        ws[f'B{row}'] = borrower_name
        row += 1
        
        ws[f'A{row}'] = "Loan ID:"
        ws[f'B{row}'] = f"#{loan.id}"
        row += 1
        
        ws[f'A{row}'] = "Principal:"
        ws[f'B{row}'] = f"₦{loan.amount:.2f}"
        row += 1
        
        ws[f'A{row}'] = "Interest Rate:"
        ws[f'B{row}'] = f"{loan.interest_rate}%"
        row += 1
        
        ws[f'A{row}'] = "Total Interest:"
        ws[f'B{row}'] = f"₦{loan.total_interest:.2f}"
        row += 1
        
        ws[f'A{row}'] = "Overdue Penalty:"
        ws[f'B{row}'] = f"₦{(getattr(loan, 'overdue_penalty', 0.0) or 0.0):.2f}"
        row += 1

        ws[f'A{row}'] = "Total Amount:"
        ws[f'B{row}'] = f"₦{get_loan_total_due_amount(loan):.2f}"
        row += 1
        
        ws[f'A{row}'] = "Amount Repaid:"
        ws[f'B{row}'] = f"₦{loan.amount_repaid:.2f}"
        row += 1
        
        ws[f'A{row}'] = "Balance:"
        ws[f'B{row}'] = f"₦{get_loan_balance_amount(loan):.2f}"
        row += 1
        
        ws[f'A{row}'] = "Status:"
        ws[f'B{row}'] = loan.status.value
        row += 2
        
        # Payment History Section
        ws[f'A{row}'] = "PAYMENT & TOP-UP HISTORY"
        ws[f'A{row}'].font = Font(bold=True, size=13)
        row += 1
        
        # Headers for payment table
        headers = ['S/N', 'Amount/Type', 'Balance', 'Date', 'Notes']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
        row += 1
        
        # Chronological event sorting to calculate running balance
        events = []
        for r in repayments:
            events.append({'type': 'repayment', 'amount': r.amount_paid, 'date': r.payment_date})
        for t in topups:
            events.append({'type': 'topup', 'amount': t.topup_amount, 'interest': t.interest_on_topup, 'rate': t.interest_rate, 'date': t.topup_date})
        for p in penalties:
            events.append(
                {
                    'type': 'penalty',
                    'amount': p.penalty_change,
                    'new_penalty': p.new_penalty,
                    'date': p.penalty_date,
                    'notes': p.notes,
                }
            )
            
        sorted_events = sorted(events, key=lambda x: x['date'])
        
        initial_principal = loan.amount - sum(t.topup_amount for t in topups)
        initial_interest = loan.total_interest - sum(t.interest_on_topup for t in topups)
        initial_penalty = (getattr(loan, "overdue_penalty", 0.0) or 0.0) - sum(
            p.penalty_change for p in penalties
        )
        running_balance = initial_principal + initial_interest + initial_penalty
        
        total_paid = 0
        seq_num = 1
        for event in sorted_events:
            if event['type'] == 'repayment':
                total_paid += event['amount']
                running_balance -= event['amount']
                running_balance = max(0, running_balance)
                
                ws.cell(row=row, column=1).value = seq_num
                ws.cell(row=row, column=2).value = event['amount']
                ws.cell(row=row, column=3).value = f"₦{running_balance:.2f}"
                ws.cell(row=row, column=4).value = event['date'].strftime("%Y-%m-%d")
                ws.cell(row=row, column=5).value = "Repayment"
            else:
                if event['type'] == 'topup':
                    running_balance += event['amount'] + event['interest']
                    
                    ws.cell(row=row, column=1).value = seq_num
                    ws.cell(row=row, column=2).value = f"₦{event['amount']:.2f} (+₦{event['interest']:.2f})"
                    ws.cell(row=row, column=3).value = f"₦{running_balance:.2f}"
                    ws.cell(row=row, column=4).value = event['date'].strftime("%Y-%m-%d")
                    ws.cell(row=row, column=5).value = f"Top-up ({event['rate']}%)"
                else:
                    running_balance += event['amount']

                    ws.cell(row=row, column=1).value = seq_num
                    ws.cell(row=row, column=2).value = f"₦{event['amount']:.2f}"
                    ws.cell(row=row, column=3).value = f"₦{running_balance:.2f}"
                    ws.cell(row=row, column=4).value = event['date'].strftime("%Y-%m-%d")
                    ws.cell(row=row, column=5).value = (
                        f"Penalty Set (Total: ₦{event['new_penalty']:.2f})"
                        if not event.get('notes')
                        else f"Penalty Set (Total: ₦{event['new_penalty']:.2f}) - {event['notes']}"
                    )
            
            row += 1
            seq_num += 1
        
        # Total paid
        ws[f'A{row}'] = "TOTAL PAID"
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'] = total_paid
        ws[f'B{row}'].font = Font(bold=True)
        row += 2
        
        # Refund History Section (if any)
        if refunds:
            ws[f'A{row}'] = "REFUND HISTORY"
            ws[f'A{row}'].font = Font(bold=True, size=13)
            row += 1
            
            headers = ['S/N', 'Amount', 'Date', 'Status']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col)
                cell.value = header
                cell.font = header_font
                cell.fill = header_fill
                cell.border = border
            row += 1
            
            total_refunded = 0
            for idx, refund in enumerate(refunds, 1):
                total_refunded += refund.refund_amount
                ws.cell(row=row, column=1).value = idx
                ws.cell(row=row, column=2).value = refund.refund_amount
                ws.cell(row=row, column=3).value = refund.refund_date.strftime("%Y-%m-%d")
                ws.cell(row=row, column=4).value = refund.status
                row += 1
            
            ws[f'A{row}'] = "TOTAL REFUNDED"
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'] = total_refunded
            ws[f'B{row}'].font = Font(bold=True)
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 30
        
        # Save file
        downloads_path = os.path.expanduser("~/Downloads")
        os.makedirs(downloads_path, exist_ok=True)
        filename = f"Loan_Details_{borrower_name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = os.path.join(downloads_path, filename)
        
        wb.save(filepath)
        print(f"✓ Exported: {filepath}")
        
        # Show success message in container
        status_container.content = ft.Row([
            ft.Icon(ft.Icons.CHECK_CIRCLE, color=ft.Colors.GREEN_400),
            ft.Text(f"✓ Export Successful! Saved: {filename}", color=ft.Colors.GREEN_200, size=14),
        ], spacing=10)
        status_container.bgcolor = ft.Colors.GREEN_900
        status_container.visible = True
        try:
            page.update()
        except:
            pass
        
    except Exception as ex:
        print(f"Export Error: {str(ex)}")
        
        # Show error message in container
        status_container.content = ft.Row([
            ft.Icon(ft.Icons.ERROR, color=ft.Colors.RED_400),
            ft.Text(f"✗ Export Failed: {str(ex)}", color=ft.Colors.RED_200, size=14),
        ], spacing=10)
        status_container.bgcolor = ft.Colors.RED_900
        status_container.visible = True
        try:
            page.update()
        except:
            pass


def create_loan_details_dialog(loan, page):
    """
    Create a loan details dialog showing:
    - Left side: Payment history
    - Right side: Loan details
    """
    
    # Recalculate overdue interest for non-member loans before displaying
    if not loan.is_member:
        update_result = update_single_loan_overdue_interest(loan.id)
        if update_result:
            # Refresh the loan object with updated total_interest from DB
            refreshed_loan = get_loan_by_id(loan.id)
            if refreshed_loan:
                loan.total_interest = refreshed_loan.total_interest
    
    # Get repayment history for this loan
    repayments = get_repayments_by_loan(loan.id)
    
    # Get top-up history for this loan
    topups = get_topups_by_loan(loan.id)

    # Get overdue penalty history for this loan
    penalties = get_penalties_by_loan(loan.id)
    
    # Get borrower info
    borrower_name = "Unknown"
    borrower_ippis = "N/A"
    borrower_contact = "N/A"
    
    if loan.is_member and loan.member_id:
        member = get_member_by_id(loan.member_id)
        if member:
            borrower_name = member.name
            borrower_ippis = member.ippis_number or "N/A"
            borrower_contact = member.contact or "N/A"
    elif not loan.is_member and loan.non_member_id:
        non_member = get_non_member_by_id(loan.non_member_id)
        if non_member:
            borrower_name = non_member.name
            borrower_ippis = non_member.ippis_number or "N/A"
            borrower_contact = non_member.contact or "N/A"

    def is_overdue_loan():
        if loan.status == LoanStatus.PAID or not loan.end_date:
            return False
        end_date = loan.end_date.date() if hasattr(loan.end_date, "date") else loan.end_date
        return end_date < datetime.now().date()

    def get_total_due():
        return get_loan_total_due_amount(loan)

    def get_balance():
        return get_loan_balance_amount(loan)
    
    # ==================== PAYMENT HISTORY TABLE (LEFT SIDE) ====================
    repayment_rows = []
    total_paid = 0
    
    # Chronological sort of events for accurate running balance
    events = []
    for r in repayments:
        events.append({'type': 'repayment', 'amount': r.amount_paid, 'date': r.payment_date})
    for t in topups:
        events.append({'type': 'topup', 'amount': t.topup_amount, 'interest': t.interest_on_topup, 'rate': t.interest_rate, 'date': t.topup_date})
    for p in penalties:
        events.append(
            {
                'type': 'penalty',
                'amount': p.penalty_change,
                'new_penalty': p.new_penalty,
                'date': p.penalty_date,
                'notes': p.notes,
            }
        )
        
    sorted_events = sorted(events, key=lambda x: x['date'])
    
    # Starting balance is base principal + base interest, ignoring topups which happen later over time
    initial_principal = loan.amount - sum(t.topup_amount for t in topups)
    initial_interest = loan.total_interest - sum(t.interest_on_topup for t in topups)
    initial_penalty = (getattr(loan, "overdue_penalty", 0.0) or 0.0) - sum(
        p.penalty_change for p in penalties
    )
    running_balance = initial_principal + initial_interest + initial_penalty
    
    for idx, event in enumerate(sorted_events, 1):
        if event['type'] == 'repayment':
            total_paid += event['amount']
            running_balance -= event['amount']
            running_balance = max(0, running_balance)
            repayment_rows.append(
                ft.DataRow(
                    cells=[
                        ft.DataCell(ft.Text(str(idx), color=ft.Colors.WHITE, size=16, weight="bold")),
                        ft.DataCell(ft.Text(f"₦{event['amount']:.2f}", color=ft.Colors.GREEN_400, weight="bold", size=16)),
                        ft.DataCell(ft.Text(f"₦{running_balance:.2f}", color=ft.Colors.BLUE_200, weight="bold", size=16)),
                        ft.DataCell(ft.Text(event['date'].strftime("%Y-%m-%d"), color=ft.Colors.GREY, size=16)),
                        ft.DataCell(ft.Text("Repayment", color=ft.Colors.GREEN_300, size=15, weight="bold")),
                    ]
                )
            )
        elif event['type'] == 'topup':
            # Top-up increases running balance
            running_balance += event['amount'] + event['interest']
            repayment_rows.append(
                ft.DataRow(
                    cells=[
                        ft.DataCell(ft.Text(str(idx), color=ft.Colors.WHITE, size=16, weight="bold")),
                        ft.DataCell(ft.Text(f"₦{event['amount']:.2f} (+₦{event['interest']:.2f})", 
                                            color=ft.Colors.ORANGE_400, weight="bold", size=16)),
                        ft.DataCell(ft.Text(f"₦{running_balance:.2f}", color=ft.Colors.BLUE_200, weight="bold", size=16)),
                        ft.DataCell(ft.Text(event['date'].strftime("%Y-%m-%d"), color=ft.Colors.GREY, size=16)),
                        ft.DataCell(ft.Text(f"Top-up ({event['rate']}%)", color=ft.Colors.ORANGE_400, size=15, weight="bold")),
                    ]
                )
            )
        else:
            running_balance += event['amount']
            penalty_color = ft.Colors.RED_300 if event['amount'] >= 0 else ft.Colors.GREEN_300
            penalty_prefix = "+" if event['amount'] > 0 else ""
            penalty_notes = (
                f"Penalty Set (Total: ₦{event['new_penalty']:.2f})"
                if not event.get('notes')
                else f"Penalty Set (Total: ₦{event['new_penalty']:.2f}) - {event['notes']}"
            )
            repayment_rows.append(
                ft.DataRow(
                    cells=[
                        ft.DataCell(ft.Text(str(idx), color=ft.Colors.WHITE, size=16, weight="bold")),
                        ft.DataCell(ft.Text(f"{penalty_prefix}₦{event['amount']:.2f}", color=penalty_color, weight="bold", size=16)),
                        ft.DataCell(ft.Text(f"₦{running_balance:.2f}", color=ft.Colors.BLUE_200, weight="bold", size=16)),
                        ft.DataCell(ft.Text(event['date'].strftime("%Y-%m-%d"), color=ft.Colors.GREY, size=16)),
                        ft.DataCell(ft.Text(penalty_notes, color=ft.Colors.RED_300, size=15, weight="bold")),
                    ]
                )
            )
    
    # ==================== REFUND HISTORY ====================
    refunds = get_refunds_by_loan(loan.id)
    refund_rows = []
    total_refunded = 0
    total_pending_refund = 0
    
    for idx, refund in enumerate(refunds, 1):
        if refund.status == "PROCESSED":
            total_refunded += refund.refund_amount
        else:
            total_pending_refund += refund.refund_amount
        status_color = ft.Colors.ORANGE_400 if refund.status == "PENDING" else ft.Colors.GREEN_400
        
        # Process button for pending refunds
        process_btn = ft.Container(content=ft.Text(""), width=0)
        if refund.status == "PENDING":
            def _open_refund_dialog(e, r=refund):
                partial_field = ft.TextField(
                    label="Partial Amount (₦)",
                    keyboard_type=ft.KeyboardType.NUMBER,
                    width=220,
                    visible=False,
                    text_size=16,
                )
                error_text = ft.Text("", color=ft.Colors.RED_400, size=14, visible=False)

                def on_option_change(ev):
                    partial_field.visible = (ev.control.value == "partial")
                    error_text.visible = False
                    page.update()

                option_group = ft.RadioGroup(
                    content=ft.Column([
                        ft.Radio(value="full", label=f"Full Refund — ₦{r.refund_amount:,.2f}"),
                        ft.Radio(value="partial", label="Partial Refund"),
                    ], spacing=4),
                    value="full",
                    on_change=on_option_change,
                )

                def confirm_refund(ev):
                    try:
                        if option_group.value == "partial":
                            raw = (partial_field.value or "").strip().replace(",", "")
                            if not raw:
                                error_text.value = "Enter an amount"
                                error_text.visible = True
                                page.update()
                                return
                            amt = float(raw)
                            if amt <= 0:
                                error_text.value = "Amount must be greater than 0"
                                error_text.visible = True
                                page.update()
                                return
                            if amt >= r.refund_amount:
                                error_text.value = f"Must be less than ₦{r.refund_amount:,.2f} (use Full Refund)"
                                error_text.visible = True
                                page.update()
                                return
                            process_refund(r.id, partial_amount=amt)
                            ToastNotification.show(page, f"✓ Partial refund ₦{amt:,.2f} processed. ₦{r.refund_amount - amt:,.2f} still pending.", NotificationType.SUCCESS)
                        else:
                            process_refund(r.id)
                            ToastNotification.show(page, f"✓ Full refund ₦{r.refund_amount:,.2f} processed!", NotificationType.SUCCESS)
                        refund_dlg.open = False
                        page.update()
                    except ValueError:
                        error_text.value = "Enter a valid number"
                        error_text.visible = True
                        page.update()
                    except Exception as ex:
                        ToastNotification.show(page, f"Error: {str(ex)}", NotificationType.ERROR)
                        page.update()

                def cancel_refund(ev):
                    refund_dlg.open = False
                    page.update()

                refund_dlg = ft.AlertDialog(
                    modal=True,
                    title=ft.Text(f"Process Refund — Loan #{loan.id}", size=18, weight="bold"),
                    content=ft.Container(
                        content=ft.Column([
                            ft.Text(f"Refund Amount: ₦{r.refund_amount:,.2f}", size=16, weight="bold", color=ft.Colors.AMBER_400),
                            option_group,
                            partial_field,
                            error_text,
                        ], spacing=10, tight=True),
                        width=320,
                    ),
                    actions=[
                        ft.TextButton("Cancel", on_click=cancel_refund),
                        ft.ElevatedButton("Confirm", on_click=confirm_refund,
                                          bgcolor=ft.Colors.GREEN_700, color=ft.Colors.WHITE),
                    ],
                    actions_alignment=ft.MainAxisAlignment.END,
                )
                page.overlay.append(refund_dlg)
                refund_dlg.open = True
                page.update()

            process_btn = ft.TextButton("Process Refund", on_click=_open_refund_dialog, style=ft.ButtonStyle(
                color=ft.Colors.GREEN_400
            ))
        
        refund_rows.append(
            ft.DataRow(
                cells=[
                    ft.DataCell(ft.Text(str(idx), color=ft.Colors.WHITE, size=16, weight="bold")),
                    ft.DataCell(ft.Text(f"₦{refund.refund_amount:.2f}", color=status_color, weight="bold", size=16)),
                    ft.DataCell(ft.Text(refund.refund_date.strftime("%Y-%m-%d"), color=ft.Colors.GREY, size=16)),
                    ft.DataCell(ft.Text(refund.status, color=status_color, size=15, weight="bold")),
                    ft.DataCell(process_btn),
                ]
            )
        )
    
    payment_history_table = ft.DataTable(
        columns=[
            ft.DataColumn(ft.Text("S/N", color=ft.Colors.WHITE, weight="bold", size=15)),
            ft.DataColumn(ft.Text("AMOUNT", color=ft.Colors.WHITE, weight="bold", size=15)),
            ft.DataColumn(ft.Text("BALANCE", color=ft.Colors.WHITE, weight="bold", size=15)),
            ft.DataColumn(ft.Text("DATE", color=ft.Colors.WHITE, weight="bold", size=15)),
            ft.DataColumn(ft.Text("TYPE", color=ft.Colors.WHITE, weight="bold", size=15)),
        ],
        rows=repayment_rows,
        bgcolor="#1a1a1a",
        divider_thickness=1,
        vertical_lines=ft.border.BorderSide(1, "#2a2a2a"),
    )

    payment_history_total_paid_text = ft.Text(f"₦{total_paid:.2f}", size=16, weight="bold", color=ft.Colors.GREEN_400)
    payment_history_total_loan_text = ft.Text(f"₦{get_total_due():.2f}", size=16, weight="bold", color=ft.Colors.BLUE_200)
    payment_history_balance_text = ft.Text(
        f"₦{get_balance():.2f}",
        size=16,
        weight="bold",
        color=ft.Colors.ORANGE_400 if get_balance() > 0 else ft.Colors.GREEN_400,
    )

    def refresh_payment_history_table():
        refreshed_rows = []
        refreshed_total_paid = 0.0

        events = []
        for r in repayments:
            events.append({'type': 'repayment', 'amount': r.amount_paid, 'date': r.payment_date})
        for t in topups:
            events.append({'type': 'topup', 'amount': t.topup_amount, 'interest': t.interest_on_topup, 'rate': t.interest_rate, 'date': t.topup_date})
        for p in penalties:
            events.append(
                {
                    'type': 'penalty',
                    'amount': p.penalty_change,
                    'new_penalty': p.new_penalty,
                    'date': p.penalty_date,
                    'notes': p.notes,
                }
            )

        sorted_events = sorted(events, key=lambda x: x['date'])

        initial_principal = loan.amount - sum(t.topup_amount for t in topups)
        initial_interest = loan.total_interest - sum(t.interest_on_topup for t in topups)
        initial_penalty = (getattr(loan, "overdue_penalty", 0.0) or 0.0) - sum(
            p.penalty_change for p in penalties
        )
        running_balance = initial_principal + initial_interest + initial_penalty

        for idx, event in enumerate(sorted_events, 1):
            if event['type'] == 'repayment':
                refreshed_total_paid += event['amount']
                running_balance -= event['amount']
                running_balance = max(0, running_balance)
                refreshed_rows.append(
                    ft.DataRow(
                        cells=[
                            ft.DataCell(ft.Text(str(idx), color=ft.Colors.WHITE, size=16, weight="bold")),
                            ft.DataCell(ft.Text(f"₦{event['amount']:.2f}", color=ft.Colors.GREEN_400, weight="bold", size=16)),
                            ft.DataCell(ft.Text(f"₦{running_balance:.2f}", color=ft.Colors.BLUE_200, weight="bold", size=16)),
                            ft.DataCell(ft.Text(event['date'].strftime("%Y-%m-%d"), color=ft.Colors.GREY, size=16)),
                            ft.DataCell(ft.Text("Repayment", color=ft.Colors.GREEN_300, size=15, weight="bold")),
                        ]
                    )
                )
            elif event['type'] == 'topup':
                running_balance += event['amount'] + event['interest']
                refreshed_rows.append(
                    ft.DataRow(
                        cells=[
                            ft.DataCell(ft.Text(str(idx), color=ft.Colors.WHITE, size=16, weight="bold")),
                            ft.DataCell(ft.Text(f"₦{event['amount']:.2f} (+₦{event['interest']:.2f})", color=ft.Colors.ORANGE_400, weight="bold", size=16)),
                            ft.DataCell(ft.Text(f"₦{running_balance:.2f}", color=ft.Colors.BLUE_200, weight="bold", size=16)),
                            ft.DataCell(ft.Text(event['date'].strftime("%Y-%m-%d"), color=ft.Colors.GREY, size=16)),
                            ft.DataCell(ft.Text(f"Top-up ({event['rate']}%)", color=ft.Colors.ORANGE_400, size=15, weight="bold")),
                        ]
                    )
                )
            else:
                running_balance += event['amount']
                penalty_color = ft.Colors.RED_300 if event['amount'] >= 0 else ft.Colors.GREEN_300
                penalty_prefix = "+" if event['amount'] > 0 else ""
                penalty_notes = (
                    f"Penalty Set (Total: ₦{event['new_penalty']:.2f})"
                    if not event.get('notes')
                    else f"Penalty Set (Total: ₦{event['new_penalty']:.2f}) - {event['notes']}"
                )
                refreshed_rows.append(
                    ft.DataRow(
                        cells=[
                            ft.DataCell(ft.Text(str(idx), color=ft.Colors.WHITE, size=16, weight="bold")),
                            ft.DataCell(ft.Text(f"{penalty_prefix}₦{event['amount']:.2f}", color=penalty_color, weight="bold", size=16)),
                            ft.DataCell(ft.Text(f"₦{running_balance:.2f}", color=ft.Colors.BLUE_200, weight="bold", size=16)),
                            ft.DataCell(ft.Text(event['date'].strftime("%Y-%m-%d"), color=ft.Colors.GREY, size=16)),
                            ft.DataCell(ft.Text(penalty_notes, color=ft.Colors.RED_300, size=15, weight="bold")),
                        ]
                    )
                )

        payment_history_table.rows = refreshed_rows
        payment_history_total_paid_text.value = f"₦{refreshed_total_paid:.2f}"
        payment_history_total_loan_text.value = f"₦{get_total_due():.2f}"
        payment_history_balance_text.value = f"₦{get_balance():.2f}"
        payment_history_balance_text.color = ft.Colors.ORANGE_400 if get_balance() > 0 else ft.Colors.GREEN_400
    
    refund_history_table = ft.DataTable(
        columns=[
            ft.DataColumn(ft.Text("S/N", color=ft.Colors.WHITE, weight="bold", size=15)),
            ft.DataColumn(ft.Text("AMOUNT", color=ft.Colors.WHITE, weight="bold", size=15)),
            ft.DataColumn(ft.Text("DATE", color=ft.Colors.WHITE, weight="bold", size=15)),
            ft.DataColumn(ft.Text("STATUS", color=ft.Colors.WHITE, weight="bold", size=15)),
            ft.DataColumn(ft.Text("ACTION", color=ft.Colors.WHITE, weight="bold", size=15)),
        ],
        rows=refund_rows,
        bgcolor="#1a1a1a",
        divider_thickness=1,
        vertical_lines=ft.border.BorderSide(1, "#2a2a2a"),
    ) if refunds else None
    
    payment_history_section = ft.Container(
        content=ft.Column(
            controls=[
                ft.Text("Payment History", size=18, weight="bold", color=ft.Colors.BLUE_200),
                ft.Container(height=5),
                ft.Row([payment_history_table], scroll="always"),
                ft.Container(height=10),
                ft.Container(
                    content=ft.Column(
                        controls=[
                            ft.Row(
                                controls=[
                                    ft.Text("Total Paid:", size=16, weight="bold", color=ft.Colors.GREEN_400),
                                    payment_history_total_paid_text,
                                ],
                                alignment=ft.MainAxisAlignment.SPACE_BETWEEN,
                            ),
                            ft.Row(
                                controls=[
                                    ft.Text("Total Loan:", size=16, weight="bold", color=ft.Colors.BLUE_200),
                                    payment_history_total_loan_text,
                                ],
                                alignment=ft.MainAxisAlignment.SPACE_BETWEEN,
                            ),
                            ft.Row(
                                controls=[
                                    ft.Text("Balance:", size=16, weight="bold", color=ft.Colors.ORANGE_400 if get_balance() > 0 else ft.Colors.GREEN_400),
                                    payment_history_balance_text,
                                ],
                                alignment=ft.MainAxisAlignment.SPACE_BETWEEN,
                            ),
                        ] + ([
                            ft.Row(
                                controls=[
                                    ft.Text("Overpaid:", size=16, weight="bold", color=ft.Colors.AMBER_400),
                                    ft.Text(f"\u20a6{total_paid - get_total_due():.2f}", size=16, weight="bold", color=ft.Colors.AMBER_400),
                                ],
                                alignment=ft.MainAxisAlignment.SPACE_BETWEEN,
                            ),
                        ] if total_paid > get_total_due() else []),
                        spacing=4,
                    ),
                    padding=10,
                    bgcolor="#2a2a2a",
                    border_radius=5,
                ),
                # Refund section (if any refunds exist)
                ft.Container(height=15) if refunds else ft.Container(height=0),
                ft.Text("Refund History", size=18, weight="bold", color=ft.Colors.ORANGE_400) if refunds else ft.Container(height=0),
                ft.Row([refund_history_table], scroll="always") if refund_history_table else ft.Container(height=0),
                ft.Container(
                    content=ft.Column(
                        controls=[
                            ft.Row(
                                controls=[
                                    ft.Text("Total Refunded:", size=16, weight="bold", color=ft.Colors.GREEN_400),
                                    ft.Text(f"\u20a6{total_refunded:.2f}", size=16, weight="bold", color=ft.Colors.GREEN_400),
                                ],
                                alignment=ft.MainAxisAlignment.SPACE_BETWEEN,
                            ),
                        ] + ([
                            ft.Row(
                                controls=[
                                    ft.Text("Pending Refund:", size=16, weight="bold", color=ft.Colors.ORANGE_400),
                                    ft.Text(f"\u20a6{total_pending_refund:.2f}", size=16, weight="bold", color=ft.Colors.ORANGE_400),
                                ],
                                alignment=ft.MainAxisAlignment.SPACE_BETWEEN,
                            ),
                        ] if total_pending_refund > 0 else []),
                        spacing=4,
                    ),
                    padding=10,
                    bgcolor="#2a2a2a",
                    border_radius=5,
                ) if refunds else ft.Container(height=0),
            ],
            expand=True,
            spacing=5,
        ),
        padding=15,
        border_radius=8,
        bgcolor="#1e1e1e",
        expand=True,
    )
    
    # ==================== LOAN DETAILS (RIGHT SIDE) ====================
    total_amount = get_total_due()
    balance = get_balance()
    
    # ==================== EDIT LOAN INFORMATION DIALOG ====================
    edit_loan_amount_field = ft.TextField(
        label="Principal Amount (₦)",
        value=str(loan.amount),
        keyboard_type="number",
        width=400,
    )
    
    edit_interest_rate_field = ft.TextField(
        label="Interest Rate (%)",
        value=str(loan.interest_rate),
        keyboard_type="number",
        width=400,
    )
    
    edit_batch_number_field = ft.TextField(
        label="Batch Number",
        value=loan.batch_number or "",
        width=400,
    )
    
    edit_cheque_number_field = ft.TextField(
        label="Cheque Number",
        value=loan.cheque_number or "",
        width=400,
    )
    
    edit_start_date_field = ft.TextField(
        label="Start Date (YYYY-MM-DD)",
        value=loan.start_date.strftime("%Y-%m-%d"),
        width=400,
    )
    
    edit_end_date_field = ft.TextField(
        label="Due Date (YYYY-MM-DD)",
        value=loan.end_date.strftime("%Y-%m-%d") if loan.end_date else "",
        width=400,
    )
    
    # Create text controls for loan details that need updating
    principal_text = ft.Text(f"₦{loan.amount:.2f}", size=16, color=ft.Colors.GREEN_400, weight="bold")
    interest_rate_text = ft.Text(f"{loan.interest_rate}%", size=16, color=ft.Colors.ORANGE_400, weight="bold")
    total_interest_text = ft.Text(f"₦{loan.total_interest:.2f}", size=16, color=ft.Colors.ORANGE_400, weight="bold")
    total_amount_text = ft.Text(f"₦{get_total_due():.2f}", size=16, color=ft.Colors.BLUE_200, weight="bold")
    overdue_penalty_text = ft.Text(f"₦{(getattr(loan, 'overdue_penalty', 0.0) or 0.0):.2f}", size=16, color=ft.Colors.RED_300, weight="bold")
    batch_number_text = ft.Text(loan.batch_number or "N/A", size=16, color=ft.Colors.GREY, weight="bold")
    cheque_number_text = ft.Text(loan.cheque_number or "N/A", size=16, color=ft.Colors.GREY, weight="bold")
    start_date_text = ft.Text(loan.start_date.strftime("%Y-%m-%d"), size=16, color=ft.Colors.GREY, weight="bold")
    end_date_text = ft.Text(loan.end_date.strftime("%Y-%m-%d") if loan.end_date else "N/A", size=16, color=ft.Colors.GREY, weight="bold")
    
    def get_remaining_months():
        if loan.status == LoanStatus.PAID:
            return "Cleared"
        if loan.end_date:
            today = datetime.now()
            if loan.end_date.date() < today.date():
                months_overdue = (today.year - loan.end_date.year) * 12 + today.month - loan.end_date.month
                return f"Overdue ({months_overdue} Months)" if months_overdue > 0 else "Overdue (This Month)"
            months = (loan.end_date.year - today.year) * 12 + loan.end_date.month - today.month
            return f"{months} Months remaining" if months > 0 else "Due this month"
        return "N/A"
        
    remaining_months_text = ft.Text(get_remaining_months(), size=16, color=ft.Colors.WHITE, weight="bold")
    
    # Create text controls for guarantor information
    guarantor_name_text = ft.Text(loan.guarantor_name or "N/A", size=16, color=ft.Colors.WHITE, weight="bold")
    guarantor_phone_text = ft.Text(loan.guarantor_phone or "N/A", size=16, color=ft.Colors.GREY, weight="bold")
    
    # Create text control for balance (depends on principal and interest)
    def get_balance():
        return get_loan_balance_amount(loan)

    history_total_loan_text = ft.Text(f"₦{get_total_due():.2f}", size=16, weight="bold", color=ft.Colors.BLUE_200)
    history_balance_text = ft.Text(
        f"₦{get_balance():.2f}",
        size=16,
        weight="bold",
        color=ft.Colors.ORANGE_400 if get_balance() > 0 else ft.Colors.GREEN_400,
    )
    payment_total_loan_text = ft.Text(f"₦{get_total_due():.2f}", size=16, color=ft.Colors.BLUE_200, weight="bold")
    amount_paid_text = ft.Text(
        f"₦{loan.amount_repaid:.2f}",
        size=16,
        color=ft.Colors.GREEN_400 if loan.amount_repaid >= get_total_due() else ft.Colors.BLUE_200,
        weight="bold",
    )
    status_value_text = ft.Text(
        "Refund Due" if total_pending_refund > 0 else loan.status.value,
        size=16,
        color=ft.Colors.AMBER_400 if total_pending_refund > 0 else ft.Colors.GREY,
        weight="bold",
    )

    def refresh_financial_texts():
        total_due = get_total_due()
        balance = get_balance()
        penalty = getattr(loan, "overdue_penalty", 0.0) or 0.0

        overdue_penalty_text.value = f"₦{penalty:.2f}"
        total_amount_text.value = f"₦{total_due:.2f}"
        history_total_loan_text.value = f"₦{total_due:.2f}"
        history_balance_text.value = f"₦{balance:.2f}"
        history_balance_text.color = ft.Colors.ORANGE_400 if balance > 0 else ft.Colors.GREEN_400
        payment_total_loan_text.value = f"₦{total_due:.2f}"
        amount_paid_text.value = f"₦{loan.amount_repaid:.2f}"
        amount_paid_text.color = ft.Colors.GREEN_400 if loan.amount_repaid >= total_due else ft.Colors.BLUE_200
        balance_text.value = f"₦{balance:.2f}"
        balance_text.color = ft.Colors.ORANGE_400 if balance > 0 else ft.Colors.GREEN_400
        status_value_text.value = "Refund Due" if total_pending_refund > 0 else loan.status.value
        status_value_text.color = ft.Colors.AMBER_400 if total_pending_refund > 0 else ft.Colors.GREY
        remaining_months_text.value = get_remaining_months()
    
    balance_text = ft.Text(f"₦{get_balance():.2f}", size=16, weight="bold",
                           color=ft.Colors.ORANGE_400 if get_balance() > 0 else ft.Colors.GREEN_400)
    
    def save_loan_changes():
        """Save changes to loan information"""
        try:
            new_amount = float(edit_loan_amount_field.value or 0)
            new_interest_rate = float(edit_interest_rate_field.value or 0)
            new_batch = edit_batch_number_field.value or None
            new_cheque = edit_cheque_number_field.value or None
            
            # Parse dates
            try:
                new_start_date = datetime.strptime(edit_start_date_field.value, "%Y-%m-%d")
            except:
                new_start_date = loan.start_date
            
            try:
                new_end_date = datetime.strptime(edit_end_date_field.value, "%Y-%m-%d") if edit_end_date_field.value else None
            except:
                new_end_date = loan.end_date
            
            # Calculate duration in months
            if new_end_date and new_start_date:
                duration_days = (new_end_date - new_start_date).days
                duration_months = max(1, duration_days // 30)  # Convert days to months
            else:
                duration_months = 12  # Default to 12 months
            
            # Calculate new total interest based on member type
            # Use the loan's is_member flag to determine calculation method
            new_total_interest = calculate_interest(
                new_amount, 
                new_interest_rate, 
                loan.is_member,  # Use the existing loan's member type
                duration_months
            )
            
            # Update loan in database
            update_loan(
                loan_id=loan.id,
                amount=new_amount,
                interest_rate=new_interest_rate,
                total_interest=new_total_interest,
                batch_number=new_batch,
                cheque_number=new_cheque,
                start_date=new_start_date,
                end_date=new_end_date,
            )
            
            # Update the loan object in memory with new values
            loan.amount = new_amount
            loan.interest_rate = new_interest_rate
            loan.total_interest = new_total_interest
            loan.batch_number = new_batch
            loan.cheque_number = new_cheque
            loan.start_date = new_start_date
            loan.end_date = new_end_date
            
            # Update ALL text controls with new values
            principal_text.value = f"₦{new_amount:.2f}"
            interest_rate_text.value = f"{new_interest_rate}%"
            total_interest_text.value = f"₦{new_total_interest:.2f}"
            total_amount_text.value = f"₦{get_total_due():.2f}"
            batch_number_text.value = new_batch or "N/A"
            cheque_number_text.value = new_cheque or "N/A"
            start_date_text.value = new_start_date.strftime("%Y-%m-%d")
            end_date_text.value = new_end_date.strftime("%Y-%m-%d") if new_end_date else "N/A"
            remaining_months_text.value = get_remaining_months()
            refresh_financial_texts()
            
            edit_loan_dialog.open = False
            ToastNotification.show(page, "✓ Loan information updated successfully!", NotificationType.SUCCESS)
            page.update()
        except Exception as ex:
            ToastNotification.show(page, f"✗ Error updating loan: {str(ex)}", NotificationType.ERROR)
    
    def close_edit_loan_dialog():
        edit_loan_dialog.open = False
        page.update()
    
    def open_edit_loan_dialog(e):
        edit_loan_dialog.open = True
        page.update()
    
    # Edit Loan Dialog
    edit_loan_dialog = ft.AlertDialog(
        title=ft.Text(f"Edit Loan Information - Loan #{loan.id}"),
        content=ft.Column(
            controls=[
                edit_loan_amount_field,
                edit_interest_rate_field,
                edit_batch_number_field,
                edit_cheque_number_field,
                edit_start_date_field,
                edit_end_date_field,
            ],
            width=450,
            spacing=10,
            scroll=ft.ScrollMode.AUTO,
        ),
        actions=[
            ft.TextButton("Cancel", on_click=lambda e: close_edit_loan_dialog()),
            ft.TextButton("Save", on_click=lambda e: save_loan_changes(), style=ft.ButtonStyle(color=ft.Colors.GREEN_400)),
        ],
    )
    
    # ==================== EDIT GUARANTOR INFORMATION DIALOG ====================
    edit_guarantor_name_field = ft.TextField(
        label="Guarantor's Name",
        value=loan.guarantor_name or "",
        width=400,
    )
    
    edit_guarantor_phone_field = ft.TextField(
        label="Guarantor's Phone Number",
        value=loan.guarantor_phone or "",
        keyboard_type="phone",
        width=400,
    )
    
    def save_guarantor_changes():
        """Save changes to guarantor information"""
        try:
            new_guarantor_name = edit_guarantor_name_field.value or None
            new_guarantor_phone = edit_guarantor_phone_field.value or None
            
            if not new_guarantor_name or not new_guarantor_phone:
                ToastNotification.show(page, "✗ Please enter guarantor name and phone number!", NotificationType.ERROR)
                page.update()
                return
            
            # Update loan guarantor info in database
            update_loan(
                loan_id=loan.id,
                guarantor_name=new_guarantor_name,
                guarantor_phone=new_guarantor_phone,
            )
            
            # Update the loan object in memory
            loan.guarantor_name = new_guarantor_name
            loan.guarantor_phone = new_guarantor_phone
            
            # Update the guarantor text controls
            guarantor_name_text.value = new_guarantor_name
            guarantor_phone_text.value = new_guarantor_phone
            
            edit_guarantor_dialog.open = False
            ToastNotification.show(page, "✓ Guarantor information updated successfully!", NotificationType.SUCCESS)
            page.update()
        except Exception as ex:
            ToastNotification.show(page, f"✗ Error updating guarantor: {str(ex)}", NotificationType.ERROR)
    
    def close_edit_guarantor_dialog():
        edit_guarantor_dialog.open = False
        page.update()
    
    def open_edit_guarantor_dialog(e):
        edit_guarantor_dialog.open = True
        page.update()
    
    # Edit Guarantor Dialog
    edit_guarantor_dialog = ft.AlertDialog(
        title=ft.Text("Edit Guarantor Information"),
        content=ft.Column(
            controls=[
                edit_guarantor_name_field,
                edit_guarantor_phone_field,
            ],
            width=450,
            spacing=10,
        ),
        actions=[
            ft.TextButton("Cancel", on_click=lambda e: close_edit_guarantor_dialog()),
            ft.TextButton("Save", on_click=lambda e: save_guarantor_changes(), style=ft.ButtonStyle(color=ft.Colors.GREEN_400)),
        ],
    )

    edit_penalty_field = ft.TextField(
        label="New Overdue Penalty Total (₦)",
        value=f"{(getattr(loan, 'overdue_penalty', 0.0) or 0.0):.2f}",
        keyboard_type="number",
        width=400,
    )

    def save_penalty_changes():
        try:
            if loan.status == LoanStatus.PAID:
                ToastNotification.show(page, "✗ Paid loans cannot have overdue penalties.", NotificationType.ERROR)
                page.update()
                return
            if not is_overdue_loan():
                ToastNotification.show(page, "✗ Overdue penalty can only be set for overdue loans.", NotificationType.ERROR)
                page.update()
                return

            raw_value = (edit_penalty_field.value or "").strip().replace(",", "")
            new_penalty = float(raw_value) if raw_value else 0.0
            if new_penalty < 0:
                ToastNotification.show(page, "✗ Overdue penalty cannot be negative.", NotificationType.ERROR)
                page.update()
                return

            penalty_change = new_penalty - (getattr(loan, "overdue_penalty", 0.0) or 0.0)
            penalty_record = record_loan_penalty_change(
                loan_id=loan.id,
                new_penalty=new_penalty,
                notes=f"Manual overdue penalty updated by ₦{penalty_change:.2f}",
            )
            if not penalty_record:
                ToastNotification.show(page, "✗ Failed to save overdue penalty.", NotificationType.ERROR)
                page.update()
                return

            loan.overdue_penalty = new_penalty
            penalties.insert(0, penalty_record)
            refresh_payment_history_table()
            refresh_financial_texts()

            edit_penalty_dialog.open = False
            ToastNotification.show(page, "✓ Overdue penalty updated successfully!", NotificationType.SUCCESS)
            page.update()
        except ValueError:
            ToastNotification.show(page, "✗ Enter a valid penalty amount.", NotificationType.ERROR)
            page.update()
        except Exception as ex:
            ToastNotification.show(page, f"✗ Error updating overdue penalty: {str(ex)}", NotificationType.ERROR)
            page.update()

    def close_edit_penalty_dialog():
        edit_penalty_dialog.open = False
        page.update()

    def open_edit_penalty_dialog(e):
        if loan.status == LoanStatus.PAID:
            ToastNotification.show(page, "✗ Paid loans cannot have overdue penalties.", NotificationType.WARNING)
            page.update()
            return
        if not is_overdue_loan():
            ToastNotification.show(page, "✗ This loan is not overdue.", NotificationType.WARNING)
            page.update()
            return
        edit_penalty_field.value = f"{(getattr(loan, 'overdue_penalty', 0.0) or 0.0):.2f}"
        edit_penalty_dialog.open = True
        page.update()

    edit_penalty_dialog = ft.AlertDialog(
        title=ft.Text(f"Set Overdue Penalty - Loan #{loan.id}"),
        content=ft.Container(
            content=ft.Column(
                controls=[
                    ft.Text("Set the current total overdue penalty for this loan.", size=14, color=ft.Colors.GREY_400),
                    edit_penalty_field,
                ],
                spacing=10,
                tight=True,
            ),
            width=450,
            height=110,
            padding=ft.padding.only(top=4, bottom=4),
        ),
        actions=[
            ft.TextButton("Cancel", on_click=lambda e: close_edit_penalty_dialog()),
            ft.TextButton("Save", on_click=lambda e: save_penalty_changes(), style=ft.ButtonStyle(color=ft.Colors.GREEN_400)),
        ],
    )

    refresh_financial_texts()
    
    loan_details_section = ft.Container(
        content=ft.Column(
            controls=[
                ft.Text("Loan Details", size=18, weight="bold", color=ft.Colors.BLUE_200),
                ft.Divider(),
                
                # Borrower Info (READ-ONLY)
                ft.Text("Borrower Information", size=16, weight="bold", color=ft.Colors.BLUE_200),
                ft.Container(
                    content=ft.Column(
                        controls=[
                            ft.Row(
                                controls=[
                                    ft.Text("Name:", size=15, weight="bold", width=110),
                                    ft.Text(borrower_name, size=16, color=ft.Colors.WHITE, weight="bold"),
                                ],
                            ),
                            ft.Row(
                                controls=[
                                    ft.Text("IPPIS:", size=15, weight="bold", width=110),
                                    ft.Text(borrower_ippis, size=16, color=ft.Colors.GREY, weight="bold"),
                                ],
                            ),
                            ft.Row(
                                controls=[
                                    ft.Text("Contact:", size=15, weight="bold", width=110),
                                    ft.Text(borrower_contact, size=16, color=ft.Colors.GREY, weight="bold"),
                                ],
                            ),
                            ft.Row(
                                controls=[
                                    ft.Text("Type:", size=15, weight="bold", width=110),
                                    ft.Text("Member" if loan.is_member else "Non-Member", size=16, weight="bold",
                                           color=ft.Colors.BLUE_200 if loan.is_member else ft.Colors.ORANGE_400),
                                ],
                            ),
                        ],
                        spacing=5,
                    ),
                    padding=10,
                    bgcolor="#2a2a2a",
                    border_radius=5,
                ),
                ft.Container(height=10),
                
                # Guarantor Information with Edit Button
                ft.Row(
                    controls=[
                        ft.Text("Guarantor Information", size=16, weight="bold", color=ft.Colors.BLUE_200),
                        ft.IconButton(
                            ft.Icons.EDIT,
                            tooltip="Edit Guarantor Details",
                            on_click=open_edit_guarantor_dialog,
                            icon_size=20,
                        ),
                    ],
                    alignment=ft.MainAxisAlignment.SPACE_BETWEEN,
                ),
                ft.Container(
                    content=ft.Column(
                        controls=[
                            ft.Row(
                                controls=[
                                    ft.Text("Name:", size=15, weight="bold", width=110),
                                    guarantor_name_text,
                                ],
                            ),
                            ft.Row(
                                controls=[
                                    ft.Text("Phone:", size=15, weight="bold", width=110),
                                    guarantor_phone_text,
                                ],
                            ),
                        ],
                        spacing=5,
                    ),
                    padding=10,
                    bgcolor="#2a2a2a",
                    border_radius=5,
                ),
                ft.Container(height=10),
                
                # Loan Amount Info with Edit Button
                ft.Row(
                    controls=[
                        ft.Text("Loan Amount Details", size=16, weight="bold", color=ft.Colors.BLUE_200),
                        ft.Row(
                            controls=[
                                ft.IconButton(
                                    ft.Icons.WARNING_AMBER,
                                    tooltip="Set Overdue Penalty",
                                    on_click=open_edit_penalty_dialog,
                                    icon_size=20,
                                    icon_color=ft.Colors.RED_300,
                                ),
                                ft.IconButton(
                                    ft.Icons.EDIT,
                                    tooltip="Edit Loan Information",
                                    on_click=open_edit_loan_dialog,
                                    icon_size=20,
                                ),
                            ],
                            spacing=0,
                        ),
                    ],
                    alignment=ft.MainAxisAlignment.SPACE_BETWEEN,
                ),
                ft.Container(
                    content=ft.Column(
                        controls=[
                            ft.Row(
                                controls=[
                                    ft.Text("Principal:", size=15, weight="bold", width=110),
                                    principal_text,
                                ],
                            ),
                            ft.Row(
                                controls=[
                                    ft.Text("Interest Rate:", size=15, weight="bold", width=110),
                                    interest_rate_text,
                                ],
                            ),
                            ft.Row(
                                controls=[
                                    ft.Text("Total Interest:", size=15, weight="bold", width=110),
                                    total_interest_text,
                                ],
                            ),
                            ft.Row(
                                controls=[
                                    ft.Text("Overdue Penalty:", size=15, weight="bold", width=110),
                                    overdue_penalty_text,
                                ],
                            ),
                            ft.Row(
                                controls=[
                                    ft.Text("Total Amount:", size=15, weight="bold", width=110),
                                    total_amount_text,
                                ],
                            ),
                        ],
                        spacing=5,
                    ),
                    padding=10,
                    bgcolor="#2a2a2a",
                    border_radius=5,
                ),
                ft.Container(height=10),
                
                # Payment Progress
                ft.Text("Payment Progress", size=16, weight="bold", color=ft.Colors.BLUE_200),
                ft.Container(
                    content=ft.Column(
                        controls=[
                            ft.Row(
                                controls=[
                                    ft.Text("Total Loan:", size=15, weight="bold", width=110),
                                    payment_total_loan_text,
                                ],
                            ),
                            ft.Row(
                                controls=[
                                    ft.Text("Amount Paid:", size=15, weight="bold", width=110),
                                    amount_paid_text,
                                ],
                            ),
                            ft.Row(
                                controls=[
                                    ft.Text("Balance:", size=15, weight="bold", width=110),
                                    balance_text,
                                ],
                            ),
                        ] + ([
                            ft.Row(
                                controls=[
                                    ft.Text("Overpaid By:", size=15, weight="bold", width=110, color=ft.Colors.AMBER_400),
                                    ft.Text(f"₦{loan.amount_repaid - get_total_due():.2f}", size=16, color=ft.Colors.AMBER_400, weight="bold"),
                                ],
                            ),
                        ] if total_pending_refund > 0 else []) + [
                            ft.Row(
                                controls=[
                                    ft.Text("Status:", size=15, weight="bold", width=110),
                                    status_value_text,
                                ],
                            ),
                        ],
                        spacing=5,
                    ),
                    padding=10,
                    bgcolor="#2a2a2a",
                    border_radius=5,
                ),
                ft.Container(height=10),
                
                # Loan Metadata
                ft.Text("Loan Information", size=16, weight="bold", color=ft.Colors.BLUE_200),
                ft.Container(
                    content=ft.Column(
                        controls=[
                            ft.Row(
                                controls=[
                                    ft.Text("Loan ID:", size=15, weight="bold", width=110),
                                    ft.Text(f"#{loan.id}", size=16, color=ft.Colors.WHITE, weight="bold"),
                                ],
                            ),
                            ft.Row(
                                controls=[
                                    ft.Text("Batch No:", size=15, weight="bold", width=110),
                                    batch_number_text,
                                ],
                            ),
                            ft.Row(
                                controls=[
                                    ft.Text("Cheque No:", size=15, weight="bold", width=110),
                                    cheque_number_text,
                                ],
                            ),
                            ft.Row(
                                controls=[
                                    ft.Text("Start Date:", size=15, weight="bold", width=110),
                                    start_date_text,
                                ],
                            ),
                            ft.Row(
                                controls=[
                                    ft.Text("Due Date:", size=15, weight="bold", width=110),
                                    end_date_text,
                                ],
                            ),
                            ft.Row(
                                controls=[
                                    ft.Text("Time Remaining:", size=15, weight="bold", width=110),
                                    remaining_months_text,
                                ],
                            ),
                        ],
                        spacing=5,
                    ),
                    padding=10,
                    bgcolor="#2a2a2a",
                    border_radius=5,
                ),
            ],
            spacing=10,
            scroll=ft.ScrollMode.AUTO,
        ),
        padding=15,
        border_radius=8,
        bgcolor="#1e1e1e",
        expand=True,
    )
    
    # ==================== MAIN DIALOG WITH TWO-COLUMN LAYOUT ====================
    
    # Status message container
    status_container = ft.Container(
        content=ft.Text("", size=14),
        padding=10,
        border_radius=5,
        visible=False,
    )
    
    def close_dialog():
        dialog.open = False
        # Clean up overlay to prevent memory leaks
        for overlay_item in [edit_loan_dialog, edit_guarantor_dialog, edit_penalty_dialog, dialog]:
            try:
                page.overlay.remove(overlay_item)
            except (ValueError, AttributeError):
                pass
        page.update()
    
    def on_export_click(e):
        thread = threading.Thread(
            target=export_loan_to_excel,
            args=(loan, borrower_name, repayments, refunds, topups, penalties, page, status_container),
            daemon=True
        )
        thread.start()
    
    dialog = ft.AlertDialog(
        title=ft.Text(f"Loan Details - {borrower_name}", size=20, weight="bold"),
        content=ft.Container(
            content=ft.Column([
                ft.Row(
                    controls=[
                        # Left side: Payment History (scrollable)
                        ft.Container(
                            content=ft.Column(
                                controls=[payment_history_section],
                                scroll=ft.ScrollMode.AUTO,
                                expand=True,
                            ),
                            expand=1,
                        ),
                        # Right side: Loan Details (scrollable)
                        ft.Container(
                            content=ft.Column(
                                controls=[loan_details_section],
                                scroll=ft.ScrollMode.AUTO,
                                expand=True,
                            ),
                            expand=1,
                        ),
                    ],
                    spacing=15,
                    expand=True,
                ),
                status_container,
            ], spacing=10, expand=True),
            width=1200,
            height=600,
        ),
        actions=[
            ft.TextButton("Export .xlsx", on_click=on_export_click, style=ft.ButtonStyle(color=ft.Colors.GREEN_400)),
            ft.TextButton("Close", on_click=lambda e: close_dialog()),
        ],
    )
    
    # Add edit dialogs to page overlay so they appear above the main dialog
    page.overlay.append(edit_loan_dialog)
    page.overlay.append(edit_guarantor_dialog)
    page.overlay.append(edit_penalty_dialog)
    
    return dialog
