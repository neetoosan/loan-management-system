"""
Comprehensive import validation and processing system
Handles CSV and Excel imports with:
- Column validation and flexibility
- Header detection
- Duplicate checking
- Transaction rollback
- Data type conversion
"""

import csv
import os
import re
from typing import List, Dict, Tuple, Optional, Set
from datetime import datetime, timedelta
from enum import Enum
from dataclasses import dataclass, field

from components.error_handler import error_logger, UserFriendlyError
from database.connection import get_session
from database.models import Member, NonMember, Loan, LoanRepayment, LoanRefund, LoanTopUp, Contribution, MemberStatus, LoanStatus, ContributionType

try:
    from openpyxl import load_workbook
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


class ColumnMatchStrategy(Enum):
    """Column matching strategies"""
    EXACT = "exact"  # Exact column name match
    FUZZY = "fuzzy"  # Fuzzy matching with variations
    POSITION = "position"  # Position-based matching


@dataclass
class ImportColumn:
    """Represents a required import column"""
    name: str  # Primary column name
    aliases: List[str] = field(default_factory=list)  # Alternative column names
    required: bool = True  # Is this column required?
    type: str = "string"  # Expected data type
    validator: Optional[callable] = None  # Custom validation function


@dataclass
class ImportRowError(Exception):
    """Import error with row and field information"""
    row_number: int
    field: str
    error_message: str
    value: str = ""
    
    def __str__(self) -> str:
        return f"Row {self.row_number} ({self.field}): {self.error_message}"


@dataclass
class ImportSummary:
    """Summary of import results"""
    successful_count: int = 0
    failed_count: int = 0
    warning_count: int = 0
    duplicate_count: int = 0
    updated_count: int = 0
    errors: List[ImportRowError] = field(default_factory=list)
    warnings: List[str] = field(default_factory=list)
    processed_records: List[Dict] = field(default_factory=list)
    failed_payments: List[Dict] = field(default_factory=list)
    
    def add_error(self, row: int, field: str, message: str, value: str = ""):
        """Add an error"""
        self.errors.append(ImportRowError(row, field, message, value))
        self.failed_count += 1
    
    def add_warning(self, message: str):
        """Add a warning"""
        self.warnings.append(message)
        self.warning_count += 1
    
    def get_summary_text(self) -> str:
        """Get human-readable summary"""
        lines = [f"Imported {self.successful_count} records"]
        if self.updated_count > 0:
            lines.append(f"Updated {self.updated_count} records")
        if self.duplicate_count > 0:
            lines.append(f"Skipped {self.duplicate_count} duplicates")
        if self.failed_count > 0:
            lines.append(f"Failed: {self.failed_count} records")
        if self.warning_count > 0:
            lines.append(f"Warnings: {self.warning_count}")
        return " | ".join(lines)


class ColumnDetector:
    """Detect and validate columns in import files"""
    
    LOAN_COLUMNS = {
        'ippis': ImportColumn('IPPIS', ['IPPIS NUMBER', 'STAFF ID', 'MEMBER ID', 'STAFF NUMBER'], required=True),
        'name': ImportColumn('NAME', ['FULL NAME', 'MEMBER NAME', 'BORROWER NAME', 'EMPLOYEE NAME', 'FULLNAME', 'SURNAME'], required=True),
        'loan_type': ImportColumn('LOAN TYPE', ['TYPE', 'MEMBER TYPE', 'BORROWER TYPE', 'COLUMN3', 'STATUS'], required=False),
        'amount': ImportColumn('LOAN COLLECTED', ['LOAN COLLEC TED', 'LOAN AMOUNT', 'AMOUNT', 'LOAN AMT'], required=True),
        'interest': ImportColumn('INTEREST', ['INTEREST RATE', 'INT RATE'], required=False, type='float'),
        'duration': ImportColumn('NO OF MONTHS', ['LOAN DURATION', 'DURATION', 'MONTHS'], required=False, type='int'),
        'batch': ImportColumn('FORM NUMBER', ['BATCH NUMBER', 'BATCH', 'BATCH NO'], required=False),
        'cheque': ImportColumn('CHEQUE NUMBER', ['CHEQUE NO', 'CHQ NO'], required=False),
        'date': ImportColumn('COMMENCEMENT DATE', ['LOAN DATE', 'ISSUE DATE', 'START DATE', 'DATE', 'LOAN ISUUE DATE', 'LOAN ISSUE DATE'], required=False, type='date'),
        'amount_paid': ImportColumn('AMOUNT PAID', ['PAID', 'TOTAL PAID', 'AMT PAID'], required=False, type='float'),
        'total_repayment': ImportColumn('TOTAL REPAYMENT', ['TOTAL DUE', 'TOTAL AMOUNT'], required=False, type='float'),
        'balance': ImportColumn('LOAN OUTSTANDING BALANCE', ['BALANCE', 'OUTSTANDING', 'OUTSTANDING BALANCE'], required=False, type='float'),
        'remark': ImportColumn('REMARK', ['REMARKS', 'NOTE', 'NOTES', 'COMMENT'], required=False),
    }
    
    CONTRIBUTION_COLUMNS = {
        'ippis': ImportColumn('IPPIS', ['IPPIS NUMBER', 'STAFF ID', 'MEMBER ID'], required=True),
        'amount': ImportColumn('AMOUNT', ['CONTRIBUTION', 'CONTRIB AMT'], required=True, type='float'),
        'type': ImportColumn('TYPE', ['CONTRIB TYPE', 'CATEGORY'], required=False),
        'month': ImportColumn('MONTH', ['MONTH NAME', 'PERIOD'], required=False),
        'date': ImportColumn('DATE', ['CONTRIBUTION DATE', 'CONTRIB DATE'], required=False, type='date'),
    }
    
    PAYMENT_COLUMNS = {
        'ippis': ImportColumn('IPPIS', ['IPPIS NUMBER', 'STAFF ID', 'MEMBER ID', 'STAFF NUMBER'], required=True),
        'name': ImportColumn('FULL NAME', ['NAME', 'MEMBER NAME', 'BORROWER NAME', 'EMPLOYEE NAME', 'FULLNAME', 'SURNAME'], required=True),
        'amount': ImportColumn('AMOUNT', ['AMOUNT PAID', 'PAYMENT', 'PAID', 'MONTHLY PAYMENT', 'AMT PAID'], required=True, type='float'),
    }
    
    @staticmethod
    def detect_headers(headers: List[str], column_definitions: Dict[str, ImportColumn]) -> Tuple[Dict[str, int], List[str]]:
        """
        Detect column positions from headers
        
        Returns:
            (column_mapping, warnings) where column_mapping is {key: column_index}
        """
        column_map = {}
        warnings = []
        headers_upper = [h.upper().strip() if h else "" for h in headers]
        
        for key, col_def in column_definitions.items():
            found = False
            
            # Try exact match first
            for idx, header in enumerate(headers_upper):
                if header == col_def.name.upper():
                    column_map[key] = idx
                    found = True
                    break
            
            # Try aliases
            if not found:
                for alias in col_def.aliases:
                    for idx, header in enumerate(headers_upper):
                        if header == alias.upper():
                            column_map[key] = idx
                            found = True
                            warnings.append(f"Column '{alias}' mapped to '{col_def.name}'")
                            break
                    if found:
                        break
            
            # Check if required but not found
            if not found and col_def.required:
                warnings.append(f"Required column '{col_def.name}' not found. Aliases: {', '.join(col_def.aliases)}")
        
        return column_map, warnings
    
    @staticmethod
    def get_column_definitions(data_type: str) -> Dict[str, ImportColumn]:
        """Get column definitions for data type"""
        if data_type == "loans":
            return ColumnDetector.LOAN_COLUMNS
        elif data_type == "contributions":
            return ColumnDetector.CONTRIBUTION_COLUMNS
        elif data_type == "payments":
            return ColumnDetector.PAYMENT_COLUMNS
        else:
            raise ValueError(f"Unknown data type: {data_type}")


class DuplicateChecker:
    """Check for duplicates before import"""
    
    def __init__(self, session):
        """Initialize with database session"""
        self.session = session
        self.existing_ippis: Set[str] = set()
        self.existing_loans: Set[Tuple] = set()
        self._load_existing_records()
    
    def _load_existing_records(self):
        """Load existing records from database"""
        try:
            members = self.session.query(Member.ippis_number).all()
            self.existing_ippis = set(m[0] for m in members if m[0])
            
            loans = self.session.query(Loan.member_id, Loan.amount, Loan.start_date).all()
            self.existing_loans = set((l[0], l[1], l[2].date() if l[2] else None) for l in loans)
        except Exception as e:
            error_logger.warning(f"Failed to load existing records: {str(e)}")
    
    def check_member_duplicate(self, ippis: str) -> bool:
        """Check if member already exists"""
        return ippis in self.existing_ippis
    
    def check_loan_duplicate(self, member_id: int, amount: float, start_date: datetime) -> bool:
        """Check if loan already exists (same member, amount, date)"""
        return (member_id, amount, start_date.date()) in self.existing_loans
    
    def register_member(self, ippis: str):
        """Register new member to avoid duplicates in current batch"""
        self.existing_ippis.add(ippis)
    
    def register_loan(self, member_id: int, amount: float, start_date: datetime):
        """Register new loan to avoid duplicates in current batch"""
        self.existing_loans.add((member_id, amount, start_date.date()))


class DataTypeConverter:
    """Convert and validate data types"""
    
    @staticmethod
    def to_float(value, field_name: str = "Value") -> Tuple[Optional[float], Optional[str]]:
        """Convert to float with validation"""
        if value is None or value == "":
            return None, None
        
        try:
            if isinstance(value, (int, float)):
                return float(value), None
            
            # Remove currency symbols and commas
            str_val = str(value).strip().replace('â‚¦', '').replace(',', '').replace('%', '')
            return float(str_val), None
        except ValueError:
            return None, f"{field_name} must be a valid number (got: {value})"
    
    @staticmethod
    def to_int(value, field_name: str = "Value") -> Tuple[Optional[int], Optional[str]]:
        """Convert to int with validation"""
        if value is None or value == "":
            return None, None
        
        try:
            if isinstance(value, int):
                return value, None
            if isinstance(value, float):
                return int(value), None
            
            return int(str(value).strip()), None
        except ValueError:
            return None, f"{field_name} must be a valid integer (got: {value})"
    
    @staticmethod
    def to_percentage(value, field_name: str = "Interest Rate") -> Tuple[Optional[float], Optional[str]]:
        """Convert to percentage (0-100 range)"""
        float_val, error = DataTypeConverter.to_float(value, field_name)
        if error:
            return None, error
        
        if float_val is None:
            return None, None
        
        # If value is less than 1, assume it's already in percentage form (0.05 = 5%)
        if 0 < float_val < 1:
            return float_val * 100, None
        
        # If value is between 1 and 100, assume it's already a percentage
        if 1 <= float_val <= 100:
            return float_val, None
        
        # If value is greater than 100, it's invalid
        if float_val > 100:
            return None, f"{field_name} must be between 0 and 100 (got: {float_val})"
        
        return float_val, None
    
    @staticmethod
    def to_date(value, field_name: str = "Date") -> Tuple[Optional[datetime], Optional[str]]:
        """Convert to datetime with validation"""
        if value is None or value == "":
            return None, None
        
        # If already datetime
        if isinstance(value, datetime):
            return value, None
        
        # Try various date formats
        date_formats = [
            "%Y-%m-%d",
            "%d-%m-%Y",
            "%m-%d-%Y",
            "%d/%m/%Y",
            "%m/%d/%Y",
            "%Y/%m/%d",
            "%d-%b-%Y",
            "%d-%B-%Y",
            "%b-%d-%Y",
            "%B-%d-%Y",
        ]
        
        str_val = str(value).strip()
        for fmt in date_formats:
            try:
                return datetime.strptime(str_val, fmt), None
            except ValueError:
                continue
        
        return None, f"{field_name} has invalid format (got: {str_val}). Supported formats: YYYY-MM-DD, DD/MM/YYYY, etc."
    
    @staticmethod
    def to_string(value, field_name: str = "Value") -> Tuple[Optional[str], Optional[str]]:
        """Convert to string with validation"""
        if value is None:
            return None, None
        
        str_val = str(value).strip()
        if not str_val:
            return None, f"{field_name} cannot be empty"
        
        return str_val, None


class ImportValidator:
    """Main import validator orchestrating the import process"""
    
    # Counter for auto-generating IPPIS numbers
    _auto_ippis_counter = 0
    
    def __init__(self, data_type: str):
        """Initialize validator"""
        self.data_type = data_type
        self.column_defs = ColumnDetector.get_column_definitions(data_type)
        self.summary = ImportSummary()
        ImportValidator._auto_ippis_counter = 0
    
    def validate_csv_file(self, file_path: str) -> Tuple[bool, List[Dict], List[str]]:
        """
        Validate CSV file structure and return rows
        
        Returns:
            (is_valid, rows, warnings)
        """
        warnings = []
        
        try:
            # Check file exists
            if not os.path.exists(file_path):
                raise FileNotFoundError(f"File not found: {file_path}")
            
            with open(file_path, 'r', encoding='utf-8') as f:
                reader = csv.DictReader(f)
                
                if not reader.fieldnames:
                    return False, [], ["CSV file has no headers"]
                
                # Detect columns
                column_map, col_warnings = ColumnDetector.detect_headers(
                    list(reader.fieldnames),
                    self.column_defs
                )
                warnings.extend(col_warnings)
                
                # Check all required columns found
                missing_required = [
                    key for key, col in self.column_defs.items()
                    if col.required and key not in column_map
                ]
                
                if missing_required:
                    return False, [], [
                        f"Missing required columns: {', '.join([self.column_defs[k].name for k in missing_required])}"
                    ]
                
                # Read all rows with column mapping
                rows = []
                for row in reader:
                    mapped_row = {}
                    for key, col_idx_key in column_map.items():
                        col_name = list(reader.fieldnames)[col_idx_key] if isinstance(col_idx_key, int) else col_idx_key
                        mapped_row[key] = row.get(col_name)
                    rows.append(mapped_row)
                
                if not rows:
                    return False, [], ["CSV file has no data rows"]
                
                return True, rows, warnings
        
        except Exception as e:
            error_logger.error(f"CSV validation error: {str(e)}")
            return False, [], [str(e)]
    
    def validate_excel_file(self, file_path: str) -> Tuple[bool, List[Dict], List[str]]:
        """
        Validate Excel file structure and return rows.
        For loan imports, also detects monthly payment columns (e.g. JUNE-18, Jul-18).
        
        Returns:
            (is_valid, rows, warnings)
        """
        if not HAS_OPENPYXL:
            return False, [], ["openpyxl not installed. Install with: pip install openpyxl"]
        
        warnings = []
        
        try:
            # Check file exists
            if not os.path.exists(file_path):
                raise FileNotFoundError(f"File not found: {file_path}")
            
            workbook = load_workbook(file_path, data_only=True)
            sheet = workbook.active
            
            if not sheet:
                return False, [], ["Excel file has no sheets"]
            
            # Find header row by searching for column names
            header_row = None
            headers = None
            
            # Scan first 10 rows for headers
            for row_idx in range(1, min(11, sheet.max_row + 1)):
                row_values = [cell.value for cell in sheet[row_idx]]
                
                # Check if this looks like a header row
                if row_values and any(v for v in row_values):
                    # Try to detect column mapping
                    col_map, _ = ColumnDetector.detect_headers(
                        [str(v) if v else "" for v in row_values],
                        self.column_defs
                    )
                    
                    # If we found some required columns, this is probably the header
                    if len(col_map) > 0:
                        header_row = row_idx
                        headers = row_values
                        break
            
            if header_row is None:
                warnings.append("Could not auto-detect header row. Assuming row 1 contains headers.")
                header_row = 1
                headers = [cell.value for cell in sheet[1]]
            else:
                warnings.append(f"Auto-detected header row at row {header_row}")
            
            # Detect column mapping
            column_map, col_warnings = ColumnDetector.detect_headers(
                [str(h) if h else "" for h in headers],
                self.column_defs
            )
            warnings.extend(col_warnings)
            
            # Check all required columns found
            missing_required = [
                key for key, col in self.column_defs.items()
                if col.required and key not in column_map
            ]
            
            if missing_required:
                return False, [], [
                    f"Missing required columns: {', '.join([self.column_defs[k].name for k in missing_required])}"
                ]
            
            # For loan imports, detect monthly payment columns
            monthly_columns = []  # [(col_index, parsed_date), ...]
            if self.data_type == "loans":
                monthly_columns = self._detect_monthly_columns(headers)
                if monthly_columns:
                    warnings.append(f"Detected {len(monthly_columns)} monthly payment columns")
            
            # Read data rows
            rows = []
            for row_idx, row in enumerate(sheet.iter_rows(
                min_row=header_row + 1,
                values_only=True
            ), start=header_row + 1):
                if row is None or not any(row):
                    continue
                
                # Map columns
                mapped_row = {}
                for key, header_idx in column_map.items():
                    if header_idx < len(row):
                        mapped_row[key] = row[header_idx]
                    else:
                        mapped_row[key] = None
                
                # For loan imports, collect monthly payment data
                if monthly_columns:
                    monthly_payments = []
                    for col_idx, payment_date in monthly_columns:
                        if col_idx < len(row):
                            val = row[col_idx]
                            if val is not None and isinstance(val, (int, float)) and val > 0:
                                monthly_payments.append({
                                    'date': payment_date,
                                    'amount': float(val)
                                })
                    mapped_row['_monthly_payments'] = monthly_payments
                
                rows.append(mapped_row)
            
            workbook.close()
            
            if not rows:
                return False, [], ["Excel file has no data rows"]
            
            return True, rows, warnings
        
        except Exception as e:
            error_logger.error(f"Excel validation error: {str(e)}")
            return False, [], [str(e)]
    
    @staticmethod
    def _detect_monthly_columns(headers: List) -> List[Tuple[int, datetime]]:
        """Detect monthly payment columns from header names like JUNE-18, Jul-18, Jan-25, etc.
        
        Returns list of (column_index, payment_date) tuples.
        """
        monthly_cols = []
        
        # Month name patterns (full and abbreviated)
        month_map = {
            'JAN': 1, 'JANUARY': 1,
            'FEB': 2, 'FEBRUARY': 2,
            'MAR': 3, 'MARCH': 3,
            'APR': 4, 'APRIL': 4,
            'MAY': 5,
            'JUN': 6, 'JUNE': 6,
            'JUL': 7, 'JULY': 7,
            'AUG': 8, 'AUGUST': 8,
            'SEP': 9, 'SEPT': 9, 'SEPTEMBER': 9,
            'OCT': 10, 'OCTOBER': 10,
            'NOV': 11, 'NOVEMBER': 11,
            'DEC': 12, 'DECEMBER': 12,
        }
        
        for idx, header in enumerate(headers):
            if header is None:
                continue
            header_str = str(header).strip().upper()
            
            # Match patterns like "JUNE-18", "Jul-18", "JUNE- 18", "Jan-25", "Dec-26"
            # Strip all internal spaces first for robust matching
            header_clean = header_str.replace(' ', '')
            match = re.match(r'^([A-Z]+)[- ]?(\d{2,4})$', header_clean)
            if match:
                month_name = match.group(1)
                year_str = match.group(2)
                
                if month_name in month_map:
                    month_num = month_map[month_name]
                    year = int(year_str)
                    if year < 100:
                        year += 2000  # 18 -> 2018, 26 -> 2026
                    
                    # Use 1st of the month as payment date
                    try:
                        payment_date = datetime(year, month_num, 1)
                        monthly_cols.append((idx, payment_date))
                    except ValueError:
                        continue
        
        return monthly_cols
    
    def validate_and_parse_rows(self, rows: List[Dict]) -> Tuple[List[Dict], List[ImportRowError]]:
        """
        Validate and parse rows according to data type
        
        Returns:
            (valid_rows, errors)
        """
        valid_rows = []
        errors = []
        
        for row_idx, row in enumerate(rows, start=2):  # Start at 2 because headers are row 1
            try:
                parsed_row = self._parse_row(row, row_idx)
                if parsed_row:
                    valid_rows.append((row_idx, parsed_row))
            except ImportRowError as ie:
                errors.append(ie)
        
        return valid_rows, errors
    
    def _parse_row(self, row: Dict, row_idx: int) -> Optional[Dict]:
        """Parse a single row according to data type"""
        if self.data_type == "loans":
            return self._parse_loan_row(row, row_idx)
        elif self.data_type == "contributions":
            return self._parse_contribution_row(row, row_idx)
        elif self.data_type == "payments":
            return self._parse_payment_row(row, row_idx)
        else:
            raise ValueError(f"Unknown data type: {self.data_type}")
    
    def _parse_loan_row(self, row: Dict, row_idx: int) -> Optional[Dict]:
        """Parse a loan row - supports production format with auto IPPIS generation"""
        parsed = {}
        
        # IPPIS - auto-generate if missing
        ippis_raw = row.get('ippis')
        if ippis_raw is not None and str(ippis_raw).strip():
            ippis_str = str(ippis_raw).strip()
            # Convert numeric IPPIS to string without decimals
            if isinstance(ippis_raw, float) and ippis_raw == int(ippis_raw):
                ippis_str = str(int(ippis_raw))
            parsed['ippis'] = ippis_str
        else:
            # Auto-generate IPPIS for rows without one
            ImportValidator._auto_ippis_counter += 1
            parsed['ippis'] = f"AUTO-{ImportValidator._auto_ippis_counter:04d}"
            self.summary.add_warning(f"Row {row_idx}: No IPPIS - auto-generated {parsed['ippis']}")
        
        # Name (required)
        name, error = DataTypeConverter.to_string(row.get('name'), "Name")
        if error or not name or name.strip() == "":
            raise ImportRowError(row_idx, "NAME", "Name is required and cannot be empty", str(row.get('name')))
        parsed['name'] = name.strip()
        
        # Loan Type / Member Type (optional, default NON MEMBER)
        loan_type_raw = row.get('loan_type')
        if loan_type_raw:
            loan_type_str = str(loan_type_raw).strip().upper()
            parsed['is_member'] = loan_type_str == 'MEMBER'
        else:
            parsed['is_member'] = False  # Default to non-member
        
        # Amount / Loan Collected (required)
        amount, error = DataTypeConverter.to_float(row.get('amount'), "Loan Collected")
        if error:
            raise ImportRowError(row_idx, "LOAN COLLECTED", error, str(row.get('amount')))
        if amount is None or amount <= 0:
            raise ImportRowError(row_idx, "LOAN COLLECTED", "Amount must be greater than 0", str(row.get('amount')))
        parsed['amount'] = amount
        
        # Interest - treat as absolute amount (production data), not percentage
        interest_raw = row.get('interest')
        if interest_raw is not None:
            interest_val, error = DataTypeConverter.to_float(interest_raw, "Interest")
            if error:
                parsed['total_interest'] = 0.0
            else:
                parsed['total_interest'] = interest_val or 0.0
        else:
            parsed['total_interest'] = 0.0
        
        # Enforce fixed 3% interest rate standard
        parsed['interest_rate'] = 3.0
        
        # Duration / NO OF MONTHS (auto-calculate based on 3% fixed rate if interest provided)
        # Formula: Total Interest = Amount * 0.03 * Duration => Duration = Total Interest / (Amount * 0.03)
        if parsed['amount'] > 0 and parsed['total_interest'] > 0:
            calculated_duration = round(parsed['total_interest'] / (parsed['amount'] * 0.03))
            parsed['duration'] = calculated_duration
        else:
            # Fallback to column value if no interest is provided
            if row.get('duration'):
                duration, error = DataTypeConverter.to_int(row.get('duration'), "No of Months")
                if error:
                    parsed['duration'] = 0
                else:
                    parsed['duration'] = duration or 0
            else:
                parsed['duration'] = 0
        
        # Batch number / FORM NUMBER (optional)
        batch_raw = row.get('batch')
        if batch_raw is not None:
            if isinstance(batch_raw, float) and batch_raw == int(batch_raw):
                parsed['batch'] = str(int(batch_raw))
            else:
                parsed['batch'] = str(batch_raw).strip() if str(batch_raw).strip() else None
        else:
            parsed['batch'] = None
        
        # Cheque number (optional)
        cheque_raw = row.get('cheque')
        if cheque_raw is not None:
            if isinstance(cheque_raw, float) and cheque_raw == int(cheque_raw):
                parsed['cheque'] = str(int(cheque_raw))
            else:
                parsed['cheque'] = str(cheque_raw).strip() if str(cheque_raw).strip() else None
        else:
            parsed['cheque'] = None
        
        # Commencement Date (optional, default today)
        if row.get('date'):
            date_val, error = DataTypeConverter.to_date(row.get('date'), "Commencement Date")
            if error:
                parsed['date'] = datetime.now()
            else:
                parsed['date'] = date_val or datetime.now()
        else:
            parsed['date'] = datetime.now()
        
        # Remark / Notes (optional)
        remark_raw = row.get('remark')
        if remark_raw is not None and str(remark_raw).strip():
            parsed['notes'] = str(remark_raw).strip()
        else:
            parsed['notes'] = None
        
        # Balance (optional - for reference)
        if row.get('balance'):
            balance_val, _ = DataTypeConverter.to_float(row.get('balance'), "Balance")
            parsed['balance'] = balance_val or 0.0
        else:
            parsed['balance'] = 0.0
        
        # Monthly payments (from production Excel monthly columns)
        parsed['monthly_payments'] = row.get('_monthly_payments', [])
        
        # Calculate amount_paid from monthly payments if available
        if parsed['monthly_payments']:
            parsed['amount_paid'] = sum(mp['amount'] for mp in parsed['monthly_payments'])
        elif row.get('amount_paid'):
            amount_paid, _ = DataTypeConverter.to_float(row.get('amount_paid'), "Amount Paid")
            parsed['amount_paid'] = amount_paid or 0.0
        else:
            parsed['amount_paid'] = 0.0
        
        return parsed
    
    def _parse_contribution_row(self, row: Dict, row_idx: int) -> Optional[Dict]:
        """Parse a contribution row with flexible type mapping"""
        parsed = {}

        # IPPIS (required â€” must not be empty)
        ippis, error = DataTypeConverter.to_string(row.get('ippis'), "IPPIS")
        if error or not ippis or not ippis.strip():
            raise ImportRowError(row_idx, "IPPIS", "IPPIS number is required and cannot be empty", str(row.get('ippis')))
        parsed['ippis'] = ippis.strip()

        # Amount (required)
        amount, error = DataTypeConverter.to_float(row.get('amount'), "Amount")
        if error:
            raise ImportRowError(row_idx, "AMOUNT", error, str(row.get('amount')))
        if amount is None or amount <= 0:
            raise ImportRowError(row_idx, "AMOUNT", "Amount must be greater than 0", str(row.get('amount')))
        parsed['amount'] = amount

        # Type (optional, default MONTHLY)
        # Map various input types to valid ContributionType enum values
        type_mapping = {
            'MONTHLY': ContributionType.MONTHLY,
            'NORMAL': ContributionType.MONTHLY,
            'REGULAR': ContributionType.MONTHLY,
            'WEEKLY': ContributionType.WEEKLY,
            'SPECIAL': ContributionType.WEEKLY,
            'VOLUNTARY': ContributionType.VOLUNTARY,
            'EMERGENCY': ContributionType.VOLUNTARY,
            'DONATION': ContributionType.VOLUNTARY,
            'EXTRA': ContributionType.VOLUNTARY,
        }

        if row.get('type'):
            type_str = str(row.get('type')).strip().upper()

            if type_str in type_mapping:
                parsed['type'] = type_mapping[type_str]
            else:
                # Unknown type - log warning and use default
                self.summary.add_warning(
                    f"Row {row_idx}: Unknown contribution type '{type_str}', using MONTHLY. "
                    f"Valid types: {', '.join(type_mapping.keys())}"
                )
                parsed['type'] = ContributionType.MONTHLY
        else:
            # No type specified - use default
            parsed['type'] = ContributionType.MONTHLY

        # Month (optional, format YYYY-MM)
        month_value = row.get('month')
        if month_value:
            month_str = str(month_value).strip()
            # Validate format YYYY-MM
            if month_str and not re.match(r'^\d{4}-\d{2}$', month_str):
                # Try to parse and reformat
                try:
                    from datetime import datetime as _dt
                    # Try common formats
                    for fmt in ['%Y-%m', '%m-%Y', '%Y/%m', '%m/%Y', '%B %Y', '%b %Y']:
                        try:
                            dt = _dt.strptime(month_str, fmt)
                            month_str = dt.strftime('%Y-%m')
                            break
                        except ValueError:
                            continue
                except Exception:
                    pass
            parsed['month'] = month_str if month_str else None
        else:
            # Auto-generate month from date if not provided
            parsed['month'] = None

        # Date (optional, default today)
        if row.get('date'):
            date_val, error = DataTypeConverter.to_date(row.get('date'), "Date")
            if error:
                raise ImportRowError(row_idx, "DATE", error, str(row.get('date')))
            parsed['date'] = date_val or datetime.now()
        else:
            parsed['date'] = datetime.now()

        # Auto-generate month from date if month is None
        if parsed['month'] is None:
            parsed['month'] = parsed['date'].strftime('%Y-%m')

        return parsed

    def _parse_payment_row(self, row: Dict, row_idx: int) -> Optional[Dict]:
        """Parse a monthly payment row (S/N, IPPIS, Full Name, Amount)"""
        parsed = {}

        # IPPIS (required)
        ippis, error = DataTypeConverter.to_string(row.get('ippis'), "IPPIS")
        if error or not ippis or not ippis.strip():
            raise ImportRowError(row_idx, "IPPIS", "IPPIS number is required and cannot be empty", str(row.get('ippis')))
        parsed['ippis'] = ippis.strip()

        # Name (required)
        name, error = DataTypeConverter.to_string(row.get('name'), "Full Name")
        if error or not name or name.strip() == "":
            raise ImportRowError(row_idx, "FULL NAME", "Full name is required", str(row.get('name')))
        parsed['name'] = name.strip()

        # Amount (required, > 0)
        amount, error = DataTypeConverter.to_float(row.get('amount'), "Amount")
        if error:
            raise ImportRowError(row_idx, "AMOUNT", error, str(row.get('amount')))
        if amount is None or amount <= 0:
            raise ImportRowError(row_idx, "AMOUNT", "Payment amount must be greater than 0", str(row.get('amount')))
        parsed['amount'] = amount

        return parsed


class ImportProcessor:
    """Process validated import rows and save to database"""
    
    def __init__(self, data_type: str, validator: ImportValidator):
        """Initialize processor"""
        self.data_type = data_type
        self.validator = validator
        self.summary = validator.summary
        self.session = get_session()
        self.duplicate_checker = DuplicateChecker(self.session)
    
    def process_rows(self, rows: List[Tuple[int, Dict]]) -> bool:
        """
        Process and save rows to database
        Returns True if all successful, False if any failed
        """
        try:
            if self.data_type == "loans":
                return self._process_loans(rows)
            elif self.data_type == "contributions":
                return self._process_contributions(rows)
            elif self.data_type == "payments":
                return self._process_payments(rows)
            else:
                raise ValueError(f"Unknown data type: {self.data_type}")
        
        except Exception as e:
            error_logger.error(f"Processing error: {str(e)}")
            self.session.rollback()
            return False
        
        finally:
            self.session.close()
    
    def _process_loans(self, rows: List[Tuple[int, Dict]]) -> bool:
        """Process loan rows - supports production format with:
        - MEMBER and NON MEMBER types
        - Monthly payment columns as individual LoanRepayment records
        - Duplicate IPPIS+Name detection for top-up loans
        - REMARK stored as loan notes
        - FORM NUMBER stored as batch_number
        """
        try:
            # Track IPPIS occurrences to detect top-ups
            # Key: ippis -> first loan object (from current batch OR existing DB loan)
            seen_borrowers = {}
            
            # Pre-load existing loans by IPPIS so re-imports top-up instead of duplicating
            existing_member_loans = (
                self.session.query(Loan, Member.ippis_number)
                .join(Member, Loan.member_id == Member.id)
                .filter(Loan.is_member == True)
                .order_by(Loan.created_at.asc())
                .all()
            )
            for loan_obj, m_ippis in existing_member_loans:
                if m_ippis and m_ippis not in seen_borrowers:
                    seen_borrowers[m_ippis] = loan_obj
            
            existing_non_member_loans = (
                self.session.query(Loan, NonMember.ippis_number)
                .join(NonMember, Loan.non_member_id == NonMember.id)
                .filter(Loan.is_member == False)
                .order_by(Loan.created_at.asc())
                .all()
            )
            for loan_obj, nm_ippis in existing_non_member_loans:
                if nm_ippis and nm_ippis not in seen_borrowers:
                    seen_borrowers[nm_ippis] = loan_obj
            
            for row_idx, row in rows:
                try:
                    ippis = row['ippis']
                    is_member = row.get('is_member', False)
                    borrower_key = ippis
                    
                    member_id = None
                    non_member_id = None
                    
                    if is_member:
                        # ---- MEMBER loan ----
                        member = self.session.query(Member).filter(Member.ippis_number == ippis).first()
                        if member:
                            if not member.name or member.name.strip() == "" or member.name == ippis:
                                member.name = row['name']
                                self.session.add(member)
                                self.session.flush()
                        else:
                            member = Member(
                                ippis_number=ippis,
                                name=row['name'],
                                status=MemberStatus.ACTIVE
                            )
                            self.session.add(member)
                            self.session.flush()
                            self.duplicate_checker.register_member(ippis)
                        member_id = member.id
                    else:
                        # ---- NON MEMBER loan ----
                        non_member = None
                        if ippis:
                            non_member = self.session.query(NonMember).filter(NonMember.ippis_number == ippis).first()
                        if not non_member and row['name'] and row['name'].strip() != "":
                            non_member = self.session.query(NonMember).filter(NonMember.name == row['name']).first()
                        
                        if not non_member:
                            non_member = NonMember(
                                name=row['name'],
                                ippis_number=ippis,
                            )
                            self.session.add(non_member)
                            self.session.flush()
                        else:
                            if (not non_member.name or non_member.name.strip() == "") and row['name']:
                                non_member.name = row['name']
                                self.session.flush()
                        non_member_id = non_member.id
                    
                    # Calculate loan values
                    total_interest = row.get('total_interest', 0.0)
                    interest_rate = row.get('interest_rate', 0.0)
                    amount_paid = row.get('amount_paid', 0.0)
                    total_due = row['amount'] + total_interest
                    duration = row.get('duration', 0)
                    end_date = row['date'] + timedelta(days=30 * duration) if duration > 0 else row['date']
                    
                    if amount_paid >= total_due and total_due > 0:
                        loan_status = LoanStatus.PAID
                    else:
                        loan_status = LoanStatus.ACTIVE
                    
                    # Check if this borrower already appeared (top-up detection)
                    is_topup = borrower_key in seen_borrowers
                    
                    if is_topup:
                        # This is a TOP-UP — merge into the FIRST loan, no new loan created
                        first_loan = seen_borrowers[borrower_key]
                        
                        # Record the top-up audit trail
                        topup_record = LoanTopUp(
                            loan_id=first_loan.id,
                            topup_amount=row['amount'],
                            interest_rate=interest_rate,
                            interest_on_topup=total_interest,
                            topup_date=row['date'],
                            notes=f"Top-up loan imported from spreadsheet. {row.get('notes') or ''}"
                        )
                        self.session.add(topup_record)
                        
                        # Update the first loan's totals
                        first_loan.amount += row['amount']
                        first_loan.total_interest += total_interest
                        first_loan.interest_rate = round(
                            (first_loan.total_interest / first_loan.amount) * 100, 2
                        ) if first_loan.amount > 0 else 0.0
                        first_loan.amount_repaid += amount_paid
                        
                        # Append notes
                        topup_note = f"[TOP-UP {row['date'].strftime('%Y-%m-%d')}] +{row['amount']:,.0f}"
                        if row.get('notes'):
                            topup_note += f" ({row['notes']})"
                        first_loan.notes = (
                            (first_loan.notes + " | " if first_loan.notes else "") + topup_note
                        )
                        
                        # Update batch/cheque if the top-up has newer values
                        if row.get('batch'):
                            first_loan.batch_number = (
                                (first_loan.batch_number + ", " if first_loan.batch_number else "")
                                + row['batch']
                            )
                        if row.get('cheque'):
                            first_loan.cheque_number = row['cheque']
                        
                        # Recalculate status
                        new_total_due = first_loan.amount + first_loan.total_interest + (getattr(first_loan, "overdue_penalty", 0.0) or 0.0)
                        if first_loan.amount_repaid >= new_total_due and new_total_due > 0:
                            first_loan.status = LoanStatus.PAID
                        else:
                            first_loan.status = LoanStatus.ACTIVE
                        
                        # Use later end_date if this top-up is newer
                        if end_date and (first_loan.end_date is None or end_date > first_loan.end_date):
                            first_loan.end_date = end_date
                        
                        self.session.flush()
                        
                        # Record monthly payments against the FIRST loan
                        monthly_payments = row.get('monthly_payments', [])
                        if monthly_payments:
                            for mp in monthly_payments:
                                repayment = LoanRepayment(
                                    loan_id=first_loan.id,
                                    amount_paid=mp['amount'],
                                    payment_date=mp['date'],
                                    notes=f"Imported (top-up): {mp['date'].strftime('%B %Y')} payment"
                                )
                                self.session.add(repayment)
                        elif amount_paid > 0:
                            repayment = LoanRepayment(
                                loan_id=first_loan.id,
                                amount_paid=min(amount_paid, total_due),
                                payment_date=row['date'],
                                notes="Imported top-up payment (from spreadsheet)"
                            )
                            self.session.add(repayment)
                        
                        self.summary.add_warning(
                            f"Row {row_idx}: {row['name']} (IPPIS {ippis}) appears again - merged as TOP-UP into loan #{first_loan.id}"
                        )
                        self.summary.updated_count += 1
                        self.summary.processed_records.append(row)
                        continue
                    
                    # ---- Create NEW loan (first occurrence) ----
                    loan = Loan(
                        member_id=member_id,
                        non_member_id=non_member_id,
                        is_member=is_member,
                        amount=row['amount'],
                        interest_rate=interest_rate,
                        total_interest=total_interest,
                        amount_repaid=amount_paid,
                        start_date=row['date'],
                        end_date=end_date,
                        status=loan_status,
                        batch_number=row.get('batch'),
                        cheque_number=row.get('cheque'),
                        notes=row.get('notes'),
                    )
                    self.session.add(loan)
                    self.session.flush()
                    
                    # Register this borrower for top-up detection
                    seen_borrowers[borrower_key] = loan
                    
                    # Record individual monthly payments as LoanRepayment records
                    monthly_payments = row.get('monthly_payments', [])
                    if monthly_payments:
                        for mp in monthly_payments:
                            repayment = LoanRepayment(
                                loan_id=loan.id,
                                amount_paid=mp['amount'],
                                payment_date=mp['date'],
                                notes=f"Imported: {mp['date'].strftime('%B %Y')} payment"
                            )
                            self.session.add(repayment)
                    elif amount_paid > 0:
                        # No monthly breakdown - record as single bulk payment
                        repayment = LoanRepayment(
                            loan_id=loan.id,
                            amount_paid=min(amount_paid, total_due),
                            payment_date=row['date'],
                            notes="Imported payment (from spreadsheet)"
                        )
                        self.session.add(repayment)
                    
                    # Handle overpayment refund
                    if amount_paid > total_due and total_due > 0:
                        overpayment = round(amount_paid - total_due, 2)
                        refund = LoanRefund(
                            loan_id=loan.id,
                            refund_amount=overpayment,
                            refund_date=row['date'],
                            status="PENDING",
                            notes=f"Overpayment detected during import"
                        )
                        self.session.add(refund)
                        self.summary.add_warning(
                            f"Row {row_idx}: Overpayment of {overpayment:,.2f} - refund record created"
                        )
                    
                    self.summary.successful_count += 1
                    self.summary.processed_records.append(row)
                    if member_id:
                        self.duplicate_checker.register_loan(member_id, row['amount'], row['date'])
                
                except Exception as e:
                    self.session.rollback()
                    error_logger.debug(f"Loan processing error at row {row_idx}: {str(e)}")
                    self.summary.add_error(row_idx, "Loan", str(e))
                    continue
            
            # Commit all successful rows
            if self.summary.successful_count > 0 or self.summary.updated_count > 0:
                self.session.commit()
                topup_count = sum(1 for _ in self.summary.warnings if 'TOP-UP' in _)
                error_logger.info(
                    f"Import complete: {self.summary.successful_count} loans "
                    f"({topup_count} top-ups detected)"
                )
                return True
            
            return self.summary.failed_count == 0
        
        except Exception as e:
            error_logger.error(f"Loan import failed: {str(e)}")
            self.session.rollback()
            return False
    
    def _process_contributions(self, rows: List[Tuple[int, Dict]]) -> bool:
        """Process contribution rows â€” updates existing contributions for same IPPIS + month"""
        try:
            for row_idx, row in rows:
                try:
                    ippis = row['ippis']

                    # Get member
                    member = self.session.query(Member).filter(Member.ippis_number == ippis).first()
                    if not member:
                        self.summary.add_error(row_idx, "Member", f"Member with IPPIS {ippis} not found. Please add member first.")
                        continue

                    # Check for existing contribution (same member + month)
                    existing = self.session.query(Contribution).filter(
                        Contribution.member_id == member.id,
                        Contribution.month == row.get('month')
                    ).first()

                    if existing:
                        # UPDATE existing contribution instead of skipping
                        existing.amount = row['amount']
                        existing.contribution_type = row['type']
                        existing.contribution_date = row['date']
                        self.session.flush()
                        self.summary.updated_count += 1
                        self.summary.processed_records.append(row)
                        continue

                    # Create new contribution
                    contrib = Contribution(
                        member_id=member.id,
                        amount=row['amount'],
                        contribution_type=row['type'],
                        contribution_date=row['date'],
                        month=row.get('month')
                    )
                    self.session.add(contrib)
                    self.session.flush()

                    self.summary.successful_count += 1
                    self.summary.processed_records.append(row)

                except Exception as e:
                    error_logger.debug(f"Contribution processing error at row {row_idx}: {str(e)}")
                    self.summary.add_error(row_idx, "Contribution", str(e))
                    continue

            # Commit all successful rows
            if self.summary.successful_count > 0 or self.summary.updated_count > 0:
                self.session.commit()
                error_logger.info(f"Import complete: {self.summary.successful_count} new, {self.summary.updated_count} updated contributions")
                return True

            return self.summary.failed_count == 0

        except Exception as e:
            error_logger.error(f"Contribution import failed: {str(e)}")
            self.session.rollback()
            return False

    def _process_payments(self, rows: List[Tuple[int, Dict]]) -> bool:
        """Process monthly payment rows — record repayments against loans.
        Each row: IPPIS + Name + Amount. Finds the borrower's loan and
        records the payment. Handles overpayment detection."""
        current_month = datetime.now().strftime('%Y-%m')
        try:
            for row_idx, row in rows:
                try:
                    ippis = str(row['ippis']).strip()
                    amount = row['amount']

                    # Find member or non-member by IPPIS
                    member = self.session.query(Member).filter(Member.ippis_number == ippis).first()
                    non_member = None
                    if not member:
                        non_member = self.session.query(NonMember).filter(NonMember.ippis_number == ippis).first()

                    if not member and not non_member:
                        self.summary.add_error(row_idx, "IPPIS", f"No member/non-member found with IPPIS {ippis}")
                        self.summary.failed_payments.append({
                            'row': row_idx,
                            'ippis': ippis,
                            'name': row.get('name', ''),
                            'amount': amount,
                            'reason': f"No member/non-member found with IPPIS {ippis}",
                        })
                        continue

                    # Find loan — prioritise active/defaulted, fall back to any with balance
                    if member:
                        active_loan = self.session.query(Loan).filter(
                            Loan.member_id == member.id,
                            Loan.is_member == True,
                        ).order_by(Loan.created_at.desc()).first()
                    else:
                        active_loan = self.session.query(Loan).filter(
                            Loan.non_member_id == non_member.id,
                            Loan.is_member == False,
                        ).order_by(Loan.created_at.desc()).first()

                    if not active_loan:
                        borrower_name = member.name if member else non_member.name
                        self.summary.add_error(
                            row_idx, "Loan",
                            f"No loan found for {borrower_name} (IPPIS {ippis})"
                        )
                        self.summary.failed_payments.append({
                            'row': row_idx,
                            'ippis': ippis,
                            'name': row.get('name', ''),
                            'db_name': borrower_name,
                            'amount': amount,
                            'reason': f"No loan found for {borrower_name}",
                        })
                        continue

                    # Name mismatch warning (IPPIS matched, name differs — still process)
                    db_name = member.name if member else non_member.name
                    file_name = row.get('name', '').strip()
                    if db_name.upper() != file_name.upper():
                        self.summary.add_warning(
                            f"Row {row_idx}: Name mismatch — file has '{file_name}' "
                            f"but DB has '{db_name}' (IPPIS {ippis}). Payment recorded."
                        )

                    # Calculate balance before payment
                    total_due = active_loan.amount + (active_loan.total_interest or 0) + (getattr(active_loan, "overdue_penalty", 0.0) or 0.0)
                    balance_before = total_due - (active_loan.amount_repaid or 0)

                    # Record the repayment
                    repayment = LoanRepayment(
                        loan_id=active_loan.id,
                        amount_paid=amount,
                        payment_date=datetime.now(),
                        notes=f"Monthly payment import ({current_month})"
                    )
                    self.session.add(repayment)

                    # Update loan amount_repaid
                    active_loan.amount_repaid = (active_loan.amount_repaid or 0) + amount
                    balance_after = total_due - active_loan.amount_repaid

                    # Check if fully paid or overpaid
                    if active_loan.amount_repaid >= total_due:
                        active_loan.status = LoanStatus.PAID
                        active_loan.end_date = datetime.now()

                        if active_loan.amount_repaid > total_due:
                            # Overpayment — create refund record
                            overpayment = round(active_loan.amount_repaid - total_due, 2)
                            refund = LoanRefund(
                                loan_id=active_loan.id,
                                refund_amount=overpayment,
                                refund_date=datetime.now(),
                                status="PENDING",
                                notes=f"Overpayment from monthly import ({current_month}) — paid \u20a6{amount:,.2f}, excess \u20a6{overpayment:,.2f}"
                            )
                            self.session.add(refund)
                            self.summary.add_warning(
                                f"Row {row_idx}: {row['name']} OVERPAID by \u20a6{overpayment:,.2f} — refund record created"
                            )

                    self.session.flush()
                    self.summary.successful_count += 1
                    db_name_for_rec = member.name if member else non_member.name
                    self.summary.processed_records.append({
                        **row,
                        'db_name': db_name_for_rec,
                        'loan_id': active_loan.id,
                        'balance_before': balance_before,
                        'balance_after': max(balance_after, 0),
                        'status': 'OVERPAID' if balance_after < 0 else ('PAID' if balance_after == 0 else 'ACTIVE'),
                    })

                except Exception as e:
                    self.session.rollback()
                    error_logger.debug(f"Payment processing error at row {row_idx}: {str(e)}")
                    self.summary.add_error(row_idx, "Payment", str(e))
                    self.summary.failed_payments.append({
                        'row': row_idx,
                        'ippis': str(row.get('ippis', '')).strip(),
                        'name': row.get('name', ''),
                        'amount': row.get('amount', 0),
                        'reason': str(e),
                    })
                    continue

            # Commit all
            if self.summary.successful_count > 0:
                self.session.commit()
                error_logger.info(f"Payment import complete: {self.summary.successful_count} payments recorded")
                return True

            return self.summary.failed_count == 0

        except Exception as e:
            error_logger.error(f"Payment import failed: {str(e)}")
            self.session.rollback()
            return False


class ImportReportGenerator:
    """Generate a detailed Excel report log after any import operation.
    
    Includes:
    - Summary statistics (imported, updated, failed, warnings)
    - Detailed error list with row numbers and values
    - Warning list (auto-IPPIS, top-ups, overpayments, etc.)
    - Successfully imported records overview
    """

    @staticmethod
    def generate_report(summary: ImportSummary, data_type: str, source_file: str,
                        export_dir: str = None) -> str:
        """Generate an import report Excel file.
        
        Args:
            summary: The ImportSummary from the completed import
            data_type: 'loans', 'contributions', or 'payments'
            source_file: Original import file path
            export_dir: Directory to save report (defaults to ~/Downloads)
        
        Returns:
            Full path to the generated report file
        """
        if not HAS_OPENPYXL:
            raise RuntimeError("openpyxl is required for report generation")

        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        if export_dir is None:
            export_dir = os.path.join(os.path.expanduser('~'), 'Downloads')
        os.makedirs(export_dir, exist_ok=True)

        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"import_report_{data_type}_{timestamp}.xlsx"
        full_path = os.path.join(export_dir, filename)

        wb = Workbook()

        # --- Styles ---
        title_font = Font(bold=True, size=14, color="2F5496")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
        success_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        warning_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        error_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        info_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        bold_font = Font(bold=True)
        stat_value_font = Font(bold=True, size=12)

        # ====================================================================
        # SHEET 1: SUMMARY
        # ====================================================================
        ws_summary = wb.active
        ws_summary.title = "Summary"

        # Title
        ws_summary.merge_cells('A1:D1')
        ws_summary['A1'] = f"Import Report — {data_type.upper()}"
        ws_summary['A1'].font = title_font
        ws_summary['A1'].alignment = Alignment(horizontal='center')

        # Metadata
        ws_summary['A3'] = "Generated:"
        ws_summary['A3'].font = bold_font
        ws_summary['B3'] = datetime.now().strftime('%d %B %Y, %I:%M %p')

        ws_summary['A4'] = "Source File:"
        ws_summary['A4'].font = bold_font
        ws_summary['B4'] = os.path.basename(source_file)

        ws_summary['A5'] = "Import Type:"
        ws_summary['A5'].font = bold_font
        ws_summary['B5'] = data_type.title()

        # Statistics table
        stats_start = 7
        stats = [
            ("Successfully Imported", summary.successful_count, success_fill),
            ("Updated / Top-ups", summary.updated_count, info_fill),
            ("Duplicates Skipped", summary.duplicate_count, info_fill),
            ("Failed", summary.failed_count, error_fill),
            ("Warnings", summary.warning_count, warning_fill),
            ("Total Processed", summary.successful_count + summary.updated_count + summary.failed_count, None),
        ]

        ws_summary['A' + str(stats_start)] = "METRIC"
        ws_summary['B' + str(stats_start)] = "COUNT"
        for col in ['A', 'B']:
            cell = ws_summary[col + str(stats_start)]
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')

        for i, (label, value, fill) in enumerate(stats):
            row = stats_start + 1 + i
            label_cell = ws_summary.cell(row=row, column=1, value=label)
            label_cell.font = bold_font
            label_cell.border = thin_border
            if fill:
                label_cell.fill = fill

            value_cell = ws_summary.cell(row=row, column=2, value=value)
            value_cell.font = stat_value_font
            value_cell.border = thin_border
            value_cell.alignment = Alignment(horizontal='center')
            if fill:
                value_cell.fill = fill

        # Overall result
        result_row = stats_start + len(stats) + 2
        ws_summary['A' + str(result_row)] = "RESULT:"
        ws_summary['A' + str(result_row)].font = Font(bold=True, size=13)
        if summary.failed_count == 0:
            ws_summary['B' + str(result_row)] = "SUCCESS"
            ws_summary['B' + str(result_row)].font = Font(bold=True, size=13, color="006100")
        elif summary.successful_count > 0:
            ws_summary['B' + str(result_row)] = "PARTIAL SUCCESS"
            ws_summary['B' + str(result_row)].font = Font(bold=True, size=13, color="9C5700")
        else:
            ws_summary['B' + str(result_row)] = "FAILED"
            ws_summary['B' + str(result_row)].font = Font(bold=True, size=13, color="9C0006")

        ws_summary.column_dimensions['A'].width = 28
        ws_summary.column_dimensions['B'].width = 20

        # ====================================================================
        # SHEET 2: ERRORS (if any)
        # ====================================================================
        if summary.errors:
            ws_errors = wb.create_sheet("Errors")

            ws_errors.merge_cells('A1:E1')
            ws_errors['A1'] = f"Import Errors — {summary.failed_count} records failed"
            ws_errors['A1'].font = Font(bold=True, size=13, color="9C0006")

            error_headers = ['#', 'Row', 'Field', 'Error Message', 'Value']
            for col_idx, header in enumerate(error_headers, 1):
                cell = ws_errors.cell(row=3, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center')

            for i, err in enumerate(summary.errors, 1):
                row = i + 3
                ws_errors.cell(row=row, column=1, value=i).border = thin_border
                ws_errors.cell(row=row, column=2, value=err.row_number).border = thin_border
                ws_errors.cell(row=row, column=3, value=err.field).border = thin_border
                ws_errors.cell(row=row, column=4, value=err.error_message).border = thin_border
                ws_errors.cell(row=row, column=5, value=err.value).border = thin_border
                for c in range(1, 6):
                    ws_errors.cell(row=row, column=c).fill = error_fill

            ws_errors.column_dimensions['A'].width = 6
            ws_errors.column_dimensions['B'].width = 8
            ws_errors.column_dimensions['C'].width = 18
            ws_errors.column_dimensions['D'].width = 55
            ws_errors.column_dimensions['E'].width = 25

        # ====================================================================
        # SHEET: FAILED PAYMENTS (if any — payment imports only)
        # ====================================================================
        if summary.failed_payments:
            ws_fp = wb.create_sheet("Failed Payments")

            ws_fp.merge_cells('A1:F1')
            ws_fp['A1'] = f"Failed Payment Records — {len(summary.failed_payments)} entries"
            ws_fp['A1'].font = Font(bold=True, size=13, color="9C0006")
            ws_fp['A2'] = "These payments could not be recorded. Review and adjust manually."
            ws_fp['A2'].font = Font(italic=True, color="9C0006")

            fp_headers = ['#', 'Row', 'IPPIS', 'Name', 'Amount', 'Reason']
            for col_idx, header in enumerate(fp_headers, 1):
                cell = ws_fp.cell(row=4, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center')

            currency_fmt_fp = '#,##0.00'
            for i, fp in enumerate(summary.failed_payments, 1):
                row = i + 4
                ws_fp.cell(row=row, column=1, value=i).border = thin_border
                ws_fp.cell(row=row, column=2, value=fp.get('row', '')).border = thin_border
                ws_fp.cell(row=row, column=3, value=str(fp.get('ippis', ''))).border = thin_border
                name_val = fp.get('db_name') or fp.get('name', '')
                ws_fp.cell(row=row, column=4, value=name_val).border = thin_border
                amt_cell = ws_fp.cell(row=row, column=5, value=fp.get('amount', 0))
                amt_cell.number_format = currency_fmt_fp
                amt_cell.border = thin_border
                ws_fp.cell(row=row, column=6, value=fp.get('reason', '')).border = thin_border
                for c in range(1, 7):
                    ws_fp.cell(row=row, column=c).fill = error_fill

            ws_fp.column_dimensions['A'].width = 6
            ws_fp.column_dimensions['B'].width = 8
            ws_fp.column_dimensions['C'].width = 16
            ws_fp.column_dimensions['D'].width = 30
            ws_fp.column_dimensions['E'].width = 18
            ws_fp.column_dimensions['F'].width = 50

        # ====================================================================
        # SHEET 3: WARNINGS (if any)
        # ====================================================================
        if summary.warnings:
            ws_warnings = wb.create_sheet("Warnings")

            ws_warnings.merge_cells('A1:C1')
            ws_warnings['A1'] = f"Import Warnings — {summary.warning_count} issues"
            ws_warnings['A1'].font = Font(bold=True, size=13, color="9C5700")

            # Categorize warnings
            auto_ippis = [w for w in summary.warnings if 'auto-generated' in w.lower() or 'AUTO-' in w]
            topups = [w for w in summary.warnings if 'TOP-UP' in w or 'top-up' in w.lower()]
            overpayments = [w for w in summary.warnings if 'overpay' in w.lower() or 'Overpayment' in w]
            other = [w for w in summary.warnings if w not in auto_ippis and w not in topups and w not in overpayments]

            warn_headers = ['#', 'Category', 'Details']
            for col_idx, header in enumerate(warn_headers, 1):
                cell = ws_warnings.cell(row=3, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = PatternFill(start_color="BF8F00", end_color="BF8F00", fill_type="solid")
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center')

            row_num = 4
            categories = [
                ("Auto-Generated IPPIS", auto_ippis, PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")),
                ("Top-Up Detected", topups, PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")),
                ("Overpayment", overpayments, PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")),
                ("Other", other, warning_fill),
            ]

            sn = 0
            for category_name, items, cat_fill in categories:
                for warning_text in items:
                    sn += 1
                    ws_warnings.cell(row=row_num, column=1, value=sn).border = thin_border
                    cat_cell = ws_warnings.cell(row=row_num, column=2, value=category_name)
                    cat_cell.border = thin_border
                    cat_cell.fill = cat_fill
                    cat_cell.font = Font(bold=True)
                    detail_cell = ws_warnings.cell(row=row_num, column=3, value=warning_text)
                    detail_cell.border = thin_border
                    row_num += 1

            ws_warnings.column_dimensions['A'].width = 6
            ws_warnings.column_dimensions['B'].width = 24
            ws_warnings.column_dimensions['C'].width = 80

        # ====================================================================
        # SHEET 4: IMPORTED RECORDS
        # ====================================================================
        if summary.processed_records:
            ws_records = wb.create_sheet("Imported Records")

            ws_records.merge_cells('A1:F1')
            ws_records['A1'] = f"Successfully Processed Records — {len(summary.processed_records)}"
            ws_records['A1'].font = Font(bold=True, size=13, color="006100")

            if data_type == "loans":
                rec_headers = ['#', 'IPPIS', 'Name', 'Type', 'Loan Amount', 'Interest',
                               'Amount Paid', 'Duration', 'Batch', 'Cheque', 'Date', 'Notes']
            elif data_type == "payments":
                rec_headers = ['#', 'IPPIS', 'Name', 'Amount Paid', 'Status']
            else:
                rec_headers = ['#', 'IPPIS', 'Amount', 'Type', 'Month']

            for col_idx, header in enumerate(rec_headers, 1):
                cell = ws_records.cell(row=3, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = PatternFill(start_color="375623", end_color="375623", fill_type="solid")
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center')

            currency_fmt = '#,##0.00'
            for i, rec in enumerate(summary.processed_records, 1):
                row = i + 3
                ws_records.cell(row=row, column=1, value=i).border = thin_border

                if data_type == "loans":
                    ws_records.cell(row=row, column=2, value=str(rec.get('ippis', ''))).border = thin_border
                    ws_records.cell(row=row, column=3, value=rec.get('name', '')).border = thin_border
                    ws_records.cell(row=row, column=4, value='Member' if rec.get('is_member') else 'Non-Member').border = thin_border
                    amt_cell = ws_records.cell(row=row, column=5, value=rec.get('amount', 0))
                    amt_cell.number_format = currency_fmt
                    amt_cell.border = thin_border
                    int_cell = ws_records.cell(row=row, column=6, value=rec.get('total_interest', 0))
                    int_cell.number_format = currency_fmt
                    int_cell.border = thin_border
                    paid_cell = ws_records.cell(row=row, column=7, value=rec.get('amount_paid', 0))
                    paid_cell.number_format = currency_fmt
                    paid_cell.border = thin_border
                    ws_records.cell(row=row, column=8, value=rec.get('duration', 0)).border = thin_border
                    ws_records.cell(row=row, column=9, value=rec.get('batch', '')).border = thin_border
                    ws_records.cell(row=row, column=10, value=rec.get('cheque', '')).border = thin_border
                    date_val = rec.get('date')
                    ws_records.cell(row=row, column=11, value=date_val.strftime('%Y-%m-%d') if date_val else '').border = thin_border
                    ws_records.cell(row=row, column=12, value=rec.get('notes', '') or '').border = thin_border
                elif data_type == "payments":
                    ws_records.cell(row=row, column=2, value=str(rec.get('ippis', ''))).border = thin_border
                    ws_records.cell(row=row, column=3, value=rec.get('name', '')).border = thin_border
                    paid_cell = ws_records.cell(row=row, column=4, value=rec.get('amount', 0))
                    paid_cell.number_format = currency_fmt
                    paid_cell.border = thin_border
                    ws_records.cell(row=row, column=5, value=rec.get('status', '')).border = thin_border
                else:
                    ws_records.cell(row=row, column=2, value=str(rec.get('ippis', ''))).border = thin_border
                    amt_cell = ws_records.cell(row=row, column=3, value=rec.get('amount', 0))
                    amt_cell.number_format = currency_fmt
                    amt_cell.border = thin_border
                    ws_records.cell(row=row, column=4, value=str(rec.get('type', ''))).border = thin_border
                    ws_records.cell(row=row, column=5, value=rec.get('month', '')).border = thin_border

            # Auto-size columns
            for col_idx in range(1, len(rec_headers) + 1):
                ws_records.column_dimensions[chr(64 + min(col_idx, 26))].width = 16

        wb.save(full_path)
        error_logger.info(f"Import report saved to: {full_path}")
        return full_path


class MonthlySummaryExporter:
    """Generate monthly payment summary report.
    Shows all active borrowers with their payment status for the month:
    - Who paid and how much
    - Who was missed (no payment)
    - Who overpaid
    - Who still has balance remaining
    """

    @staticmethod
    def generate_summary(processed_records: List[Dict] = None) -> Dict:
        """Generate monthly summary comparing active borrowers vs payments received.

        Returns only MISSED, OVERPAID, and FULLY PAID borrowers (the actionable items).

        Args:
            processed_records: Records from a payment import (optional).

        Returns:
            Dict with keys:
                'rows': List of dicts {sn, ippis, name, amount, balance, note, category}
                'stats': Dict {total_borrowers, payments_received, missed, overpaid, fully_paid, total_collected}
        """
        from database.connection import get_all_active_borrowers, get_monthly_repayments

        current_month = datetime.now().strftime('%Y-%m')

        # Build lookup of payments made this import (by IPPIS)
        imported_payments = {}
        if processed_records:
            for rec in processed_records:
                ippis = str(rec.get('ippis', '')).strip()
                if ippis:
                    imported_payments[ippis] = rec

        # Get all active borrowers
        all_borrowers = get_all_active_borrowers()

        # Also get borrowers whose loans were just marked PAID by this import
        if processed_records:
            for rec in processed_records:
                ippis = str(rec.get('ippis', '')).strip()
                if not any(b['ippis'] == ippis for b in all_borrowers):
                    all_borrowers.append({
                        'ippis': ippis,
                        'name': rec.get('db_name', rec.get('name', '')),
                        'is_member': True,
                        'loan_id': rec.get('loan_id'),
                        'loan_amount': 0,
                        'total_interest': 0,
                        'total_due': 0,
                        'amount_repaid': 0,
                        'balance': 0,
                    })

        missed_rows = []
        overpaid_rows = []
        fully_paid_rows = []
        total_collected = 0
        payments_received = 0

        for borrower in all_borrowers:
            ippis = borrower['ippis']
            name = borrower['name']

            if ippis in imported_payments:
                rec = imported_payments[ippis]
                amount_paid = rec.get('amount', 0)
                status = rec.get('status', 'ACTIVE')
                balance_after = rec.get('balance_after', 0)
                total_collected += amount_paid
                payments_received += 1

                if status == 'OVERPAID':
                    overpayment = abs(balance_after) if balance_after < 0 else round(amount_paid - borrower.get('balance', 0), 2)
                    overpaid_rows.append({
                        'ippis': ippis, 'name': name, 'amount': amount_paid,
                        'balance': 0, 'note': f"OVERPAID by \u20a6{overpayment:,.2f}",
                        'category': 'OVERPAID',
                    })
                elif status == 'PAID':
                    fully_paid_rows.append({
                        'ippis': ippis, 'name': name, 'amount': amount_paid,
                        'balance': 0, 'note': 'FULLY PAID',
                        'category': 'FULLY PAID',
                    })
                # ACTIVE status = normal payment, not an exception — skip from report
            else:
                balance = borrower.get('balance', 0)
                missed_rows.append({
                    'ippis': ippis, 'name': name, 'amount': 0,
                    'balance': balance,
                    'note': f"No payment received (Balance: \u20a6{balance:,.2f})",
                    'category': 'MISSED',
                })

        # Number rows sequentially across all categories
        all_exception_rows = []
        sn = 0
        for row in missed_rows + overpaid_rows + fully_paid_rows:
            sn += 1
            row['sn'] = sn
            all_exception_rows.append(row)

        stats = {
            'total_borrowers': len(all_borrowers),
            'payments_received': payments_received,
            'missed': len(missed_rows),
            'overpaid': len(overpaid_rows),
            'fully_paid': len(fully_paid_rows),
            'total_collected': total_collected,
        }

        return {'rows': all_exception_rows, 'stats': stats}

    @staticmethod
    def export_to_excel(summary_data: Dict, file_path: str) -> str:
        """Export monthly payment exception report to Excel.

        Only includes MISSED, OVERPAID, and FULLY PAID borrowers — grouped by section.

        Args:
            summary_data: Output from generate_summary() — dict with 'rows' and 'stats'
            file_path: Directory to save the file

        Returns:
            Full path to the exported file
        """
        if not HAS_OPENPYXL:
            raise RuntimeError("openpyxl is required for Excel export. Install with: pip install openpyxl")

        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        summary_rows = summary_data.get('rows', [])
        stats = summary_data.get('stats', {})

        current_month = datetime.now().strftime('%Y-%m')
        filename = f"monthly_payment_summary_{current_month}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        full_path = os.path.join(file_path, filename)
        os.makedirs(file_path, exist_ok=True)

        wb = Workbook()
        ws = wb.active
        ws.title = f"Payment Summary {current_month}"

        # Styles
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
        title_font = Font(bold=True, size=14, color="2F5496")
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        missed_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        overpaid_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        paid_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        section_font = Font(bold=True, size=12, color="FFFFFF")
        currency_format = '#,##0.00'

        # Title
        ws.merge_cells('A1:E1')
        ws['A1'] = f"Monthly Payment Exception Report — {current_month}"
        ws['A1'].font = title_font
        ws['A1'].alignment = Alignment(horizontal='center')

        ws.merge_cells('A2:E2')
        ws['A2'] = f"Generated: {datetime.now().strftime('%d %B %Y, %I:%M %p')}"
        ws['A2'].alignment = Alignment(horizontal='center')
        ws['A2'].font = Font(italic=True, color="666666")

        # --- Stats overview (row 4-10) ---
        ws.cell(row=4, column=1, value="OVERVIEW").font = Font(bold=True, size=12)
        stat_items = [
            ("Total Active Borrowers:", stats.get('total_borrowers', 0)),
            ("Payments Received:", stats.get('payments_received', 0)),
            ("Missed Payments:", stats.get('missed', 0)),
            ("Overpaid:", stats.get('overpaid', 0)),
            ("Fully Paid Off:", stats.get('fully_paid', 0)),
            ("Total Collected:", stats.get('total_collected', 0)),
        ]
        for i, (label, val) in enumerate(stat_items):
            ws.cell(row=5 + i, column=1, value=label).font = Font(bold=True)
            val_cell = ws.cell(row=5 + i, column=2, value=val)
            if label == "Total Collected:":
                val_cell.number_format = currency_format
                val_cell.font = Font(bold=True)
            elif label == "Missed Payments:" and val > 0:
                val_cell.font = Font(color="FF0000", bold=True)

        current_row = 12  # Start sections after stats

        # Helper: write a section
        def write_section(title, rows, fill, section_fill_color):
            nonlocal current_row
            if not rows:
                return

            # Section header
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=5)
            section_cell = ws.cell(row=current_row, column=1,
                                   value=f"{title} ({len(rows)})")
            section_cell.font = section_font
            section_cell.fill = PatternFill(start_color=section_fill_color,
                                            end_color=section_fill_color, fill_type="solid")
            section_cell.alignment = Alignment(horizontal='center')
            current_row += 1

            # Column headers
            col_headers = ['S/N', 'IPPIS', 'NAME', 'AMOUNT', 'NOTE']
            for col_idx, h in enumerate(col_headers, 1):
                cell = ws.cell(row=current_row, column=col_idx, value=h)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')
                cell.border = thin_border
            current_row += 1

            # Data rows
            for sn, row_data in enumerate(rows, 1):
                ws.cell(row=current_row, column=1, value=sn).border = thin_border
                ws.cell(row=current_row, column=2, value=row_data['ippis']).border = thin_border
                ws.cell(row=current_row, column=3, value=row_data['name']).border = thin_border
                amt_cell = ws.cell(row=current_row, column=4, value=row_data['amount'])
                amt_cell.number_format = currency_format
                amt_cell.border = thin_border
                ws.cell(row=current_row, column=5, value=row_data['note']).border = thin_border
                for col in range(1, 6):
                    ws.cell(row=current_row, column=col).fill = fill
                current_row += 1

            current_row += 1  # Blank row between sections

        # Group rows by category
        missed = [r for r in summary_rows if r.get('category') == 'MISSED']
        overpaid = [r for r in summary_rows if r.get('category') == 'OVERPAID']
        fully_paid = [r for r in summary_rows if r.get('category') == 'FULLY PAID']

        write_section("MISSED PAYMENTS", missed, missed_fill, "C00000")
        write_section("OVERPAID", overpaid, overpaid_fill, "BF8F00")
        write_section("FULLY PAID", fully_paid, paid_fill, "375623")

        # Legend
        current_row += 1
        ws.cell(row=current_row, column=1, value="LEGEND").font = Font(bold=True, size=11)
        current_row += 1
        ws.cell(row=current_row, column=1, value="").fill = missed_fill
        ws.cell(row=current_row, column=2, value="MISSED — No payment received this month")
        current_row += 1
        ws.cell(row=current_row, column=1, value="").fill = overpaid_fill
        ws.cell(row=current_row, column=2, value="OVERPAID — Payment exceeded total loan balance")
        current_row += 1
        ws.cell(row=current_row, column=1, value="").fill = paid_fill
        ws.cell(row=current_row, column=2, value="FULLY PAID — Loan completely paid off")

        # Column widths
        ws.column_dimensions['A'].width = 6
        ws.column_dimensions['B'].width = 14
        ws.column_dimensions['C'].width = 30
        ws.column_dimensions['D'].width = 18
        ws.column_dimensions['E'].width = 50

        wb.save(full_path)
        return full_path
