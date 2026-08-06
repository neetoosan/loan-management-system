import flet as ft
from database.connection import get_member_by_id, get_contributions_by_member
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
import os
import threading
import logging

# Logger for error reporting
error_logger = logging.getLogger(__name__)


def export_contribution_to_excel(member, contributions, total_contributed, total_withdrawn, page, status_container):
    """
    Export contribution details to Excel file
    """
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Contribution Details"
        
        # Define styles
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=14)
        title_font = Font(bold=True, size=16, color="1F4E78")
        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Title
        ws['A1'] = f"CONTRIBUTION DETAILS - {member.name}"
        ws['A1'].font = title_font
        ws.merge_cells('A1:E1')
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # Member Info
        row = 3
        ws[f'A{row}'] = "MEMBER INFORMATION"
        ws[f'A{row}'].font = Font(bold=True, size=13)
        row += 1
        
        ws[f'A{row}'] = "Name:"
        ws[f'B{row}'] = member.name
        row += 1
        
        ws[f'A{row}'] = "IPPIS:"
        ws[f'B{row}'] = member.ippis_number or "N/A"
        row += 1
        
        ws[f'A{row}'] = "Contact:"
        ws[f'B{row}'] = member.contact or "N/A"
        row += 2
        
        # Summary Section
        ws[f'A{row}'] = "SUMMARY"
        ws[f'A{row}'].font = Font(bold=True, size=13)
        row += 1
        
        ws[f'A{row}'] = "Total Contributions:"
        ws[f'B{row}'] = total_contributed
        ws[f'B{row}'].fill = green_fill
        ws[f'B{row}'].font = Font(bold=True)
        row += 1
        
        ws[f'A{row}'] = "Total Withdrawals:"
        ws[f'B{row}'] = total_withdrawn
        ws[f'B{row}'].fill = red_fill
        ws[f'B{row}'].font = Font(bold=True)
        row += 1
        
        ws[f'A{row}'] = "Net Balance:"
        ws[f'B{row}'] = total_contributed - total_withdrawn
        ws[f'B{row}'].font = Font(bold=True)
        row += 2
        
        # Contribution History
        ws[f'A{row}'] = "CONTRIBUTION HISTORY"
        ws[f'A{row}'].font = Font(bold=True, size=13)
        row += 1
        
        # Headers
        headers = ['S/N', 'Amount', 'Type', 'Category', 'Date', 'Notes']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
        row += 1
        
        # Data
        for idx, contrib in enumerate(contributions, 1):
            ws.cell(row=row, column=1).value = idx
            ws.cell(row=row, column=2).value = abs(contrib.amount)
            ws.cell(row=row, column=3).value = "Contribution" if contrib.amount > 0 else "Withdrawal"
            ws.cell(row=row, column=4).value = contrib.contribution_type.value
            ws.cell(row=row, column=5).value = contrib.contribution_date.strftime("%Y-%m-%d")
            ws.cell(row=row, column=6).value = contrib.notes or "-"
            
            # Color code rows
            if contrib.amount > 0:
                ws.cell(row=row, column=3).fill = green_fill
            else:
                ws.cell(row=row, column=3).fill = red_fill
            
            row += 1
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 18
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 25

        # Prepare filename and downloads path
        filename = f"{member.name.replace(' ', '_')}_contributions_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        downloads_path = os.path.join(os.path.expanduser("~"), "Downloads")
        filepath = os.path.join(downloads_path, filename)
        
        wb.save(filepath)
        print(f"✓ Exported: {filepath}")
        
        # Update UI on main thread using page's update mechanism
        def update_success_ui():
            status_container.content = ft.Row([
                ft.Icon(ft.Icons.CHECK_CIRCLE, color=ft.Colors.GREEN_400, size=22),
                ft.Text(f"✓ Export Successful! Saved to Downloads: {filename}", color=ft.Colors.GREEN_200, size=14),
            ], spacing=10)
            status_container.bgcolor = ft.Colors.with_opacity(0.2, ft.Colors.GREEN)
            status_container.border = ft.border.all(1, ft.Colors.GREEN_400)
            status_container.visible = True
            page.update()
        
        # Update UI directly (running on main thread)
        try:
            update_success_ui()
        except Exception:
            pass
        
    except Exception as ex:
        print(f"Export Error: {str(ex)}")
        
        # Update UI on main thread using page's update mechanism
        def update_error_ui():
            status_container.content = ft.Row([
                ft.Icon(ft.Icons.ERROR, color=ft.Colors.RED_400, size=22),
                ft.Text(f"✗ Export Failed: {str(ex)}", color=ft.Colors.RED_200, size=14),
            ], spacing=10)
            status_container.bgcolor = ft.Colors.with_opacity(0.2, ft.Colors.RED)
            status_container.border = ft.border.all(1, ft.Colors.RED_400)
            status_container.visible = True
            page.update()
        
        # Update UI directly (running on main thread)
        try:
            update_error_ui()
        except Exception:
            pass


def create_contribution_details_dialog(member_id, page):
    """
    Create a contribution details dialog showing all contributions for a member
    """
    
    # Get member info
    member = get_member_by_id(member_id)
    if not member:
        return ft.AlertDialog(
            title=ft.Text("Error"),
            content=ft.Text("Member not found"),
            actions=[ft.TextButton("Close")],
        )
    
    # Get all contributions for this member
    contributions = get_contributions_by_member(member_id)
    
    # Build contribution history table
    contribution_rows = []
    total_contributed = 0
    total_withdrawn = 0
    
    for idx, contrib in enumerate(contributions, 1):
        if contrib.amount > 0:
            total_contributed += contrib.amount
            amount_color = ft.Colors.GREEN_400
            type_label = "Contribution"
        else:
            total_withdrawn += abs(contrib.amount)
            amount_color = ft.Colors.RED_400
            type_label = "Withdrawal"
        
        contribution_rows.append(
            ft.DataRow(
                cells=[
                    ft.DataCell(ft.Text(str(idx), color=ft.Colors.WHITE, size=16, weight="bold")),
                    ft.DataCell(ft.Text(f"₦{abs(contrib.amount):.2f}", color=amount_color, weight="bold", size=16)),
                    ft.DataCell(ft.Text(type_label, color=amount_color, size=16, weight="bold")),
                    ft.DataCell(ft.Text(contrib.contribution_type.value, color=ft.Colors.GREY, size=15, weight="bold")),
                    ft.DataCell(ft.Text(contrib.contribution_date.strftime("%Y-%m-%d"), color=ft.Colors.GREY, size=16, weight="bold")),
                    ft.DataCell(ft.Text(contrib.notes or "-", color=ft.Colors.WHITE, size=15)),
                ]
            )
        )
    
    # Contribution history table
    contribution_history_table = ft.DataTable(
        columns=[
            ft.DataColumn(ft.Text("S/N", color=ft.Colors.WHITE, weight="bold", size=15)),
            ft.DataColumn(ft.Text("AMOUNT", color=ft.Colors.WHITE, weight="bold", size=15)),
            ft.DataColumn(ft.Text("TYPE", color=ft.Colors.WHITE, weight="bold", size=15)),
            ft.DataColumn(ft.Text("CATEGORY", color=ft.Colors.WHITE, weight="bold", size=15)),
            ft.DataColumn(ft.Text("DATE", color=ft.Colors.WHITE, weight="bold", size=15)),
            ft.DataColumn(ft.Text("NOTES", color=ft.Colors.WHITE, weight="bold", size=15)),
        ],
        rows=contribution_rows,
        bgcolor="#1a1a1a",
        divider_thickness=1,
    )
    
    # Summary section
    net_balance = total_contributed - total_withdrawn
    
    summary_section = ft.Column(
        controls=[
            ft.Text("Summary", size=18, weight="bold", color=ft.Colors.BLUE_200),
            ft.Row(
                controls=[
                    ft.Container(
                        content=ft.Column(
                            controls=[
                                ft.Text("Total Contributions", size=15, color=ft.Colors.GREY, weight="bold"),
                                ft.Text(f"₦{total_contributed:.2f}", size=20, weight="bold", color=ft.Colors.GREEN_400),
                            ],
                            spacing=5,
                        ),
                        padding=10,
                        bgcolor="#252525",
                        border_radius=8,
                        expand=True,
                    ),
                    ft.Container(
                        content=ft.Column(
                            controls=[
                                ft.Text("Total Withdrawals", size=15, color=ft.Colors.GREY, weight="bold"),
                                ft.Text(f"₦{total_withdrawn:.2f}", size=20, weight="bold", color=ft.Colors.RED_400),
                            ],
                            spacing=5,
                        ),
                        padding=10,
                        bgcolor="#252525",
                        border_radius=8,
                        expand=True,
                    ),
                    ft.Container(
                        content=ft.Column(
                            controls=[
                                ft.Text("Net Balance", size=15, color=ft.Colors.GREY, weight="bold"),
                                ft.Text(f"₦{net_balance:.2f}", size=20, weight="bold", color=ft.Colors.BLUE_200),
                            ],
                            spacing=5,
                        ),
                        padding=10,
                        bgcolor="#252525",
                        border_radius=8,
                        expand=True,
                    ),
                ],
                spacing=10,
            ),
        ],
        spacing=10,
    )
    
    # Create dialog
    def close_dialog():
        dialog.open = False
        page.update()
    
    # Status container for export messages
    status_container = ft.Container(
        content=ft.Text("", size=14, color=ft.Colors.WHITE),
        padding=15,
        border_radius=8,
        visible=False
    )
    
    def on_export_click(e):
        # Show loading state immediately
        status_container.content = ft.Row([
            ft.ProgressRing(width=16, height=16, stroke_width=2, color=ft.Colors.BLUE_400),
            ft.Text("Exporting to Excel...", color=ft.Colors.BLUE_200, size=14),
        ], spacing=10)
        status_container.bgcolor = ft.Colors.with_opacity(0.2, ft.Colors.BLUE)
        status_container.border = ft.border.all(1, ft.Colors.BLUE_400)
        status_container.visible = True
        page.update()
        
        # Run export synchronously on UI thread so updates are visible
        try:
            export_contribution_to_excel(member, contributions, total_contributed, total_withdrawn, page, status_container)
        except Exception as ex:
            error_logger.exception(f"Export failed: {ex}")
            try:
                status_container.content = ft.Row([
                    ft.Icon(ft.Icons.ERROR, color=ft.Colors.RED_400, size=18),
                    ft.Text(f"✗ Export failed: {str(ex)}", color=ft.Colors.RED_200, size=14),
                ], spacing=10)
                status_container.bgcolor = ft.Colors.with_opacity(0.2, ft.Colors.RED)
                status_container.border = ft.border.all(1, ft.Colors.RED_400)
                status_container.visible = True
                page.update()
            except Exception:
                pass
    
    dialog = ft.AlertDialog(
        title=ft.Text(f"Contribution History - {member.name}", size=20, weight="bold", color=ft.Colors.BLUE_200),
        content=ft.Column(
            controls=[
                ft.Container(
                    content=ft.Column(
                        controls=[
                            ft.Text(f"IPPIS: {member.ippis_number or 'N/A'}", color=ft.Colors.GREY, size=16, weight="bold"),
                            ft.Text(f"Contact: {member.contact or 'N/A'}", color=ft.Colors.GREY, size=16, weight="bold"),
                        ],
                        spacing=5,
                    ),
                    padding=10,
                    bgcolor="#252525",
                    border_radius=8,
                ),
                ft.Divider(),
                summary_section,
                ft.Divider(),
                ft.Text("Contribution History", size=18, weight="bold", color=ft.Colors.BLUE_200),
                ft.Container(
                    content=contribution_history_table,
                    height=400,
                    expand=True,
                ),
                status_container,
            ],
            width=1000,
            spacing=10,
            scroll=ft.ScrollMode.AUTO,
        ),
        actions=[
            ft.TextButton(
                "📥 Export .xlsx",
                on_click=on_export_click,
                style=ft.ButtonStyle(color=ft.Colors.GREEN_400)
            ),
            ft.TextButton("Close", on_click=lambda e: close_dialog()),
        ],
    )
    
    return dialog
