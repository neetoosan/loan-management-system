"""
Advanced reporting system for LMS application
Supports: filtering, date ranges, multiple export formats (CSV, Excel, PDF), scheduled reports
"""

import csv
import os
import sys
from datetime import datetime, timedelta
from typing import Dict, List, Tuple, Optional
from enum import Enum

# Add parent directory to path for imports
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

try:
    from components.error_handler import error_logger
except ImportError:
    # Fallback for direct imports
    from app.components.error_handler import error_logger

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

try:
    from reportlab.lib.pagesizes import letter, A4, A3, landscape
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
    from reportlab.pdfgen import canvas
    HAS_REPORTLAB = True
except ImportError:
    HAS_REPORTLAB = False


def _loan_total_due(loan) -> float:
    return (loan.amount or 0.0) + (loan.total_interest or 0.0) + (getattr(loan, "overdue_penalty", 0.0) or 0.0)


def _loan_balance(loan) -> float:
    return max(0.0, _loan_total_due(loan) - (loan.amount_repaid or 0.0))


class ReportType(Enum):
    """Report types available"""
    MEMBER_SUMMARY = "member_summary"
    LOAN_SUMMARY = "loan_summary"
    CONTRIBUTION_SUMMARY = "contribution_summary"
    DETAILED_MEMBER = "detailed_member"
    LOAN_STATUS = "loan_status"
    OVERDUE_LOANS = "overdue_loans"


class ExportFormat(Enum):
    """Export format types"""
    CSV = "csv"
    EXCEL = "excel"
    PDF = "pdf"


class DateRange:
    """Date range for filtering"""
    
    def __init__(self, start_date: datetime = None, end_date: datetime = None):
        """
        Initialize date range
        
        Args:
            start_date: Start date (default: 90 days ago)
            end_date: End date (default: today)
        """
        self.end_date = end_date or datetime.now()
        self.start_date = start_date or (self.end_date - timedelta(days=90))
    
    @staticmethod
    def last_30_days():
        """Last 30 days"""
        end = datetime.now()
        start = end - timedelta(days=30)
        return DateRange(start, end)
    
    @staticmethod
    def last_90_days():
        """Last 90 days"""
        end = datetime.now()
        start = end - timedelta(days=90)
        return DateRange(start, end)
    
    @staticmethod
    def last_year():
        """Last year"""
        end = datetime.now()
        start = end - timedelta(days=365)
        return DateRange(start, end)
    
    @staticmethod
    def this_month():
        """This month"""
        now = datetime.now()
        start = datetime(now.year, now.month, 1)
        end = datetime.now()
        return DateRange(start, end)
    
    @staticmethod
    def this_year():
        """This year"""
        start = datetime(datetime.now().year, 1, 1)
        return DateRange(start, datetime.now())
    
    def is_in_range(self, date: datetime) -> bool:
        """Check if date is in range"""
        if date is None:
            return False
        return self.start_date <= date <= self.end_date


class ReportFilter:
    """Filter criteria for reports"""
    
    def __init__(self):
        """Initialize filter"""
        self.date_range = DateRange.last_90_days()
        self.member_ids = None  # None = all members
        self.loan_statuses = None  # None = all statuses
        self.include_paid_loans = True
        self.include_pending_loans = True
        self.include_overdue_loans = True
        self.borrower_type = "all"  # "all", "member", or "non_member"
    
    def matches_loan(self, loan) -> bool:
        """Check if loan matches filter criteria"""
        # Date range check
        if not self.date_range.is_in_range(loan.start_date):
            return False
        
        # Borrower type filter
        if self.borrower_type == "member" and not loan.is_member:
            return False
        if self.borrower_type == "non_member" and loan.is_member:
            return False
        
        # Member filter
        if self.member_ids and loan.member_id not in self.member_ids:
            return False
        
        # Status filter
        status = loan.status.value.upper()
        
        if status == "PAID" and not self.include_paid_loans:
            return False
        if status == "PENDING" and not self.include_pending_loans:
            return False
        if status == "ACTIVE" and not self.include_overdue_loans:
            return False
        
        return True
    
    def matches_contribution(self, contribution) -> bool:
        """Check if contribution matches filter criteria"""
        # Date range check
        if not self.date_range.is_in_range(contribution.contribution_date):
            return False
        
        # Member filter
        if self.member_ids and contribution.member_id not in self.member_ids:
            return False
        
        return True


class ReportGenerator:
    """Generate reports in various formats"""
    
    @staticmethod
    def generate_member_summary(members: List, loans: List, contributions: List,
                               report_filter: ReportFilter) -> Dict:
        """
        Generate member summary report (includes both members and non-members)
        
        Returns:
            Dict with headers and rows
        """
        from database.models import NonMember
        from database.connection import get_session, get_all_non_members
        
        report = {
            "title": "Borrower Summary Report",
            "date_generated": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "date_range": f"{report_filter.date_range.start_date.strftime('%Y-%m-%d')} to {report_filter.date_range.end_date.strftime('%Y-%m-%d')}",
            "headers": ["Name", "IPPIS", "Borrower Type", "Total Contributed", "Total Borrowed", "Active Loans", "Avg Loan Amount"],
            "rows": []
        }
        
        # Process members
        for member in members:
            # Check if member should be included
            if report_filter.member_ids and member.id not in report_filter.member_ids:
                continue
            
            # Calculate totals
            member_contribs = [c for c in contributions if c.member_id == member.id and report_filter.matches_contribution(c)]
            member_loans = [l for l in loans if l.member_id == member.id and report_filter.matches_loan(l)]
            
            total_contributed = sum(c.amount for c in member_contribs)
            total_borrowed = sum(l.amount for l in member_loans)
            active_loans = sum(1 for l in member_loans if l.status.value.upper() != "PAID")
            avg_loan = (total_borrowed / len(member_loans)) if member_loans else 0
            
            # Skip if no activity
            if total_contributed > 0 or total_borrowed > 0:
                report["rows"].append([
                    member.name,
                    member.ippis_number or "N/A",
                    "Member",
                    f"₦{total_contributed:,.2f}",
                    f"₦{total_borrowed:,.2f}",
                    str(active_loans),
                    f"₦{avg_loan:,.2f}",
                ])
        
        # Process non-members
        try:
            non_members = get_all_non_members()
            for non_member in non_members:
                # Get non-member loans
                nm_loans = [l for l in loans if l.non_member_id == non_member.id and report_filter.matches_loan(l)]
                
                total_borrowed = sum(l.amount for l in nm_loans)
                active_loans = sum(1 for l in nm_loans if l.status.value.upper() != "PAID")
                avg_loan = (total_borrowed / len(nm_loans)) if nm_loans else 0
                
                # Skip if no loans
                if total_borrowed > 0:
                    report["rows"].append([
                        non_member.name,
                        non_member.ippis_number or "N/A",
                        "Non-Member",
                        "₦0.00",  # Non-members don't contribute
                        f"₦{total_borrowed:,.2f}",
                        str(active_loans),
                        f"₦{avg_loan:,.2f}",
                    ])
        except Exception as e:
            error_logger.warning(f"Could not load non-members for report: {str(e)}")
        
        return report
    
    @staticmethod
    def generate_loan_summary(loans: List, report_filter: ReportFilter) -> Dict:
        """Generate loan summary report (includes both member and non-member loans)"""
        from database.connection import get_all_non_members
        
        filtered_loans = [l for l in loans if report_filter.matches_loan(l)]
        
        # Create a borrower lookup
        borrower_names = {}
        
        # Add member names
        if hasattr(report_filter, '_members_cache'):
            for member in report_filter._members_cache:
                borrower_names[(member.id, True)] = member.name
        
        # Add non-member names
        try:
            non_members = get_all_non_members()
            for nm in non_members:
                borrower_names[(nm.id, False)] = nm.name
        except Exception:
            pass
        
        report = {
            "title": "Loan Summary Report",
            "date_generated": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "date_range": f"{report_filter.date_range.start_date.strftime('%Y-%m-%d')} to {report_filter.date_range.end_date.strftime('%Y-%m-%d')}",
            "headers": ["Loan ID", "Borrower Name", "Type", "Amount", "Interest Rate", "Total Due", "Amount Repaid", "Balance", "Status", "Start Date", "End Date"],
            "rows": []
        }
        
        for loan in filtered_loans:
            total_due = _loan_total_due(loan)
            balance = _loan_balance(loan)
            
            # Get borrower name
            if loan.is_member:
                borrower_name = borrower_names.get((loan.member_id, True), f"Member #{loan.member_id}")
                borrower_type = "Member"
            else:
                borrower_name = borrower_names.get((loan.non_member_id, False), f"Non-Member #{loan.non_member_id}")
                borrower_type = "Non-Member"
            
            report["rows"].append([
                str(loan.id),
                borrower_name,
                borrower_type,
                f"₦{loan.amount:,.2f}",
                f"{loan.interest_rate:.1f}%",
                f"₦{total_due:,.2f}",
                f"₦{loan.amount_repaid:,.2f}",
                f"₦{balance:,.2f}",
                loan.status.value,
                loan.start_date.strftime("%Y-%m-%d"),
                loan.end_date.strftime("%Y-%m-%d") if loan.end_date else "N/A",
            ])
        
        # Add summary statistics
        report["summary"] = {
            "total_loans": len(filtered_loans),
            "total_amount": sum(l.amount for l in filtered_loans),
            "total_borrowed": sum(l.amount for l in filtered_loans),
            "total_repaid": sum(l.amount_repaid for l in filtered_loans),
            "total_due": sum(_loan_total_due(l) for l in filtered_loans),
            "avg_interest_rate": (sum(l.interest_rate for l in filtered_loans) / len(filtered_loans)) if filtered_loans else 0,
        }
        
        return report
    
    @staticmethod
    def generate_loan_status_report(loans: List, report_filter: ReportFilter) -> Dict:
        """Generate loan status report organized by status category"""
        from database.connection import get_all_non_members
        from database.models import LoanStatus
        
        filtered_loans = [l for l in loans if report_filter.matches_loan(l)]
        
        # Create a borrower lookup
        borrower_names = {}
        
        # Add member names
        if hasattr(report_filter, '_members_cache'):
            for member in report_filter._members_cache:
                borrower_names[(member.id, True)] = member.name
        
        # Add non-member names
        try:
            non_members = get_all_non_members()
            for nm in non_members:
                borrower_names[(nm.id, False)] = nm.name
        except Exception:
            pass
        
        # Organize loans by status
        loans_by_status = {
            "pending": [],
            "active": [],
            "paid": [],
            "overdue": [],
            "defaulted": []
        }
        
        # Categorize loans
        today = datetime.now().date()
        for loan in filtered_loans:
            # Check if loan is overdue (active, has end_date, and past due date)
            if loan.status == LoanStatus.ACTIVE and loan.end_date:
                end_date = loan.end_date.date() if isinstance(loan.end_date, datetime) else loan.end_date
                if today > end_date:
                    loans_by_status["overdue"].append(loan)
                else:
                    loans_by_status["active"].append(loan)
            else:
                status_key = loan.status.value.lower()
                if status_key in loans_by_status:
                    loans_by_status[status_key].append(loan)
                else:
                    loans_by_status["active"].append(loan)  # Default category
        
        # Create report
        report = {
            "title": "Loan Status Report",
            "date_generated": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "date_range": f"{report_filter.date_range.start_date.strftime('%Y-%m-%d')} to {report_filter.date_range.end_date.strftime('%Y-%m-%d')}",
            "headers": ["Loan ID", "Borrower Name", "Type", "Amount", "Balance", "Status", "Start Date", "Days Outstanding"],
            "rows": []
        }
        
        # Add loans organized by status
        for status, loans_in_status in loans_by_status.items():
            if loans_in_status:
                # Add status section header
                report["rows"].append([f"--- {status.upper()} LOANS ({len(loans_in_status)}) ---", "", "", "", "", "", "", ""])
                
                for loan in loans_in_status:
                    total_due = _loan_total_due(loan)
                    balance = _loan_balance(loan)
                    days_outstanding = (datetime.now().date() - loan.start_date.date()).days if hasattr(loan.start_date, 'date') else 0
                    
                    # Get borrower name
                    if loan.is_member:
                        borrower_name = borrower_names.get((loan.member_id, True), f"Member #{loan.member_id}")
                        borrower_type = "Member"
                    else:
                        borrower_name = borrower_names.get((loan.non_member_id, False), f"Non-Member #{loan.non_member_id}")
                        borrower_type = "Non-Member"
                    
                    report["rows"].append([
                        str(loan.id),
                        borrower_name,
                        borrower_type,
                        f"₦{loan.amount:,.2f}",
                        f"₦{balance:,.2f}",
                        loan.status.value,
                        loan.start_date.strftime("%Y-%m-%d"),
                        str(days_outstanding),
                    ])
        
        # Add summary statistics
        report["summary"] = {
            "total_loans": len(filtered_loans),
            "pending_loans": len(loans_by_status["pending"]),
            "active_loans": len(loans_by_status["active"]),
            "overdue_loans": len(loans_by_status["overdue"]),
            "paid_loans": len(loans_by_status["paid"]),
            "defaulted_loans": len(loans_by_status["defaulted"]),
            "total_outstanding": sum(_loan_balance(l) for l in filtered_loans if l.status != LoanStatus.PAID),
            "total_repaid": sum(l.amount_repaid for l in filtered_loans),
        }
        
        return report
    
    @staticmethod
    def generate_detailed_member_report(member, loans: List, contributions: List) -> Dict:
        """Generate detailed report for specific member"""
        member_loans = [l for l in loans if l.member_id == member.id]
        member_contribs = [c for c in contributions if c.member_id == member.id]
        
        report = {
            "title": f"Detailed Member Report: {member.name}",
            "date_generated": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "member_info": {
                "ID": member.id,
                "Name": member.name,
                "IPPIS": member.ippis_number or "N/A",
                "Contact": member.contact or "N/A",
                "Email": member.email or "N/A",
                "Status": member.status.value,
                "Join Date": member.join_date.strftime("%Y-%m-%d"),
            },
            "loans": {
                "headers": ["Loan ID", "Amount", "Interest %", "Status", "Start Date", "End Date", "Balance"],
                "rows": []
            },
            "contributions": {
                "headers": ["Date", "Amount", "Type", "Month"],
                "rows": []
            }
        }
        
        # Add loan details
        for loan in member_loans:
            total_due = _loan_total_due(loan)
            balance = _loan_balance(loan)
            report["loans"]["rows"].append([
                str(loan.id),
                f"₦{loan.amount:,.2f}",
                f"{loan.interest_rate:.1f}%",
                loan.status.value,
                loan.start_date.strftime("%Y-%m-%d"),
                loan.end_date.strftime("%Y-%m-%d") if loan.end_date else "N/A",
                f"₦{balance:,.2f}",
            ])
        
        # Add contribution details
        for contrib in member_contribs:
            report["contributions"]["rows"].append([
                contrib.contribution_date.strftime("%Y-%m-%d"),
                f"₦{contrib.amount:,.2f}",
                contrib.contribution_type.value,
                contrib.month or "N/A",
            ])
        
        # Summaries
        report["summary"] = {
            "total_loans": len(member_loans),
            "active_loans": sum(1 for l in member_loans if l.status.value.upper() != "PAID"),
            "total_borrowed": sum(l.amount for l in member_loans),
            "total_repaid": sum(l.amount_repaid for l in member_loans),
            "total_contributed": sum(c.amount for c in member_contribs),
        }
        
        return report

    @staticmethod
    def generate_ippis_ledger(members: List, loans: List, contributions: List) -> Dict:
        """
        Generate IPPIS Ledger report for ALL members and non-members.
        One row per person with aggregated financial data.
        """
        from database.connection import (
            get_all_non_members, get_contributions_by_member,
            get_repayments_by_loan, get_refunds_by_loan, get_topups_by_loan,
        )
        from database.models import LoanStatus

        report = {
            "title": "IPPIS Ledger Report",
            "date_generated": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "headers": [
                "SN", "IPPIS", "Name", "Contribution Amount",
                "Monthly Deduction", "Loan Refund", "Loan Balance",
                "Loan Paid", "Equity",
            ],
            "rows": [],
        }

        sn = 0

        # Totals accumulators
        t_contrib = 0.0
        t_deduction = 0.0
        t_refund = 0.0
        t_balance = 0.0
        t_paid = 0.0
        t_equity = 0.0

        # --- Members ---
        for member in members:
            member_contribs = get_contributions_by_member(member.id)
            member_loans = [l for l in loans if l.is_member and l.member_id == member.id]

            contribution_amount = sum(c.amount for c in member_contribs if c.amount > 0)

            # Monthly deduction: for each active loan, (total_due / duration_months)
            monthly_deduction = 0.0
            loan_balance = 0.0
            loan_paid_total = 0.0
            loan_refund_total = 0.0

            for loan in member_loans:
                total_due = _loan_total_due(loan)

                # Repayments
                repayments = get_repayments_by_loan(loan.id)
                loan_paid_total += sum(r.amount_paid for r in repayments)

                # Refunds (PENDING = still owed back to borrower)
                refunds = get_refunds_by_loan(loan.id)
                loan_refund_total += sum(rf.refund_amount for rf in refunds if rf.status == "PENDING")

                # Balance (active/overdue loans only)
                if loan.status == LoanStatus.ACTIVE:
                    loan_balance += max(0, total_due - loan.amount_repaid)

                    # Monthly deduction
                    if loan.start_date and loan.end_date:
                        months = max(1, round((loan.end_date - loan.start_date).days / 30))
                        monthly_deduction += total_due / months

            equity = contribution_amount

            # Only include if there's any financial activity
            if contribution_amount > 0 or member_loans:
                sn += 1
                report["rows"].append([
                    str(sn),
                    member.ippis_number or "N/A",
                    member.name,
                    f"₦{contribution_amount:,.2f}",
                    f"₦{monthly_deduction:,.2f}",
                    f"₦{loan_refund_total:,.2f}",
                    f"₦{loan_balance:,.2f}",
                    f"₦{loan_paid_total:,.2f}",
                    f"₦{equity:,.2f}",
                ])
                t_contrib += contribution_amount
                t_deduction += monthly_deduction
                t_refund += loan_refund_total
                t_balance += loan_balance
                t_paid += loan_paid_total
                t_equity += equity

        # --- Non-Members ---
        try:
            non_members = get_all_non_members()
            for nm in non_members:
                nm_loans = [l for l in loans if not l.is_member and l.non_member_id == nm.id]
                if not nm_loans:
                    continue

                monthly_deduction = 0.0
                loan_balance = 0.0
                loan_paid_total = 0.0
                loan_refund_total = 0.0

                for loan in nm_loans:
                    total_due = _loan_total_due(loan)

                    repayments = get_repayments_by_loan(loan.id)
                    loan_paid_total += sum(r.amount_paid for r in repayments)

                    refunds = get_refunds_by_loan(loan.id)
                    loan_refund_total += sum(rf.refund_amount for rf in refunds if rf.status == "PENDING")

                    if loan.status == LoanStatus.ACTIVE:
                        loan_balance += max(0, total_due - loan.amount_repaid)

                        if loan.start_date and loan.end_date:
                            months = max(1, round((loan.end_date - loan.start_date).days / 30))
                            monthly_deduction += total_due / months

                sn += 1
                report["rows"].append([
                    str(sn),
                    nm.ippis_number or "N/A",
                    nm.name,
                    "₦0.00",  # Non-members don't contribute
                    f"₦{monthly_deduction:,.2f}",
                    f"₦{loan_refund_total:,.2f}",
                    f"₦{loan_balance:,.2f}",
                    f"₦{loan_paid_total:,.2f}",
                    "₦0.00",  # Non-members have no equity
                ])
                t_deduction += monthly_deduction
                t_refund += loan_refund_total
                t_balance += loan_balance
                t_paid += loan_paid_total
        except Exception as e:
            error_logger.warning(f"Could not load non-members for IPPIS ledger: {str(e)}")

        report["summary"] = {
            "total_people": sn,
            "total_contributions": t_contrib,
            "total_monthly_deductions": t_deduction,
            "total_refunds": t_refund,
            "total_loan_balance": t_balance,
            "total_loan_paid": t_paid,
            "total_equity": t_equity,
        }

        return report


class ReportExporter:
    """Export reports in different formats"""
    
    @staticmethod
    def export_csv(report: Dict, file_path: str) -> bool:
        """
        Export report to CSV
        
        Returns:
            True if successful, False otherwise
        """
        try:
            with open(file_path, 'w', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)
                
                # Write header info
                writer.writerow([report.get("title", "Report")])
                writer.writerow([f"Generated: {report.get('date_generated', '')}"])
                if "date_range" in report:
                    writer.writerow([f"Date Range: {report['date_range']}"])
                writer.writerow([])  # Blank line
                
                # Write table
                writer.writerow(report["headers"])
                for row in report["rows"]:
                    writer.writerow(row)
                
                # Write summary if available
                if "summary" in report:
                    writer.writerow([])
                    writer.writerow(["SUMMARY"])
                    for key, value in report["summary"].items():
                        if isinstance(value, float):
                            writer.writerow([key, f"{value:,.2f}"])
                        else:
                            writer.writerow([key, value])
            
            error_logger.info(f"CSV report exported to {file_path}")
            return True
        
        except Exception as e:
            error_logger.error(f"Failed to export CSV report: {str(e)}")
            return False
    
    @staticmethod
    def export_excel(report: Dict, file_path: str) -> bool:
        """
        Export report to Excel with formatting
        
        Returns:
            True if successful, False otherwise
        """
        if not HAS_OPENPYXL:
            error_logger.error("openpyxl not installed, cannot export to Excel")
            return False
        
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Report"
            
            # Define styles
            title_font = Font(name='Calibri', size=14, bold=True, color="FFFFFF")
            title_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
            header_font = Font(name='Calibri', size=11, bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            alt_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            currency_format = '₦#,##0.00'
            
            row = 1
            
            # Write title
            ws.merge_cells(f'A{row}:H{row}')
            title_cell = ws[f'A{row}']
            title_cell.value = report.get("title", "Report")
            title_cell.font = title_font
            title_cell.fill = title_fill
            title_cell.alignment = Alignment(horizontal='center', vertical='center')
            ws.row_dimensions[row].height = 25
            row += 1
            
            # Write metadata
            ws.merge_cells(f'A{row}:H{row}')
            meta_cell = ws[f'A{row}']
            meta_cell.value = f"Generated: {report.get('date_generated', '')}"
            meta_cell.font = Font(italic=True, size=9)
            row += 1
            
            if "date_range" in report:
                ws.merge_cells(f'A{row}:H{row}')
                range_cell = ws[f'A{row}']
                range_cell.value = f"Date Range: {report['date_range']}"
                range_cell.font = Font(italic=True, size=9)
                row += 1
            
            row += 1  # Blank row
            
            # Write headers
            for col_idx, header in enumerate(report["headers"], 1):
                cell = ws.cell(row=row, column=col_idx)
                cell.value = header
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border
            
            row += 1
            
            # Write data rows
            for row_idx, data_row in enumerate(report["rows"]):
                for col_idx, value in enumerate(data_row, 1):
                    cell = ws.cell(row=row, column=col_idx)
                    cell.value = value
                    
                    # Alternate row colors
                    if row_idx % 2 == 1:
                        cell.fill = alt_fill
                    
                    cell.border = border
                    cell.alignment = Alignment(horizontal='right' if isinstance(value, (int, float)) else 'left', vertical='center')
                    
                    # Format currency
                    if isinstance(value, str) and value.startswith('₦'):
                        cell.number_format = currency_format
                
                row += 1
            
            # Write summary if available
            if "summary" in report:
                row += 1
                ws.merge_cells(f'A{row}:B{row}')
                summary_cell = ws[f'A{row}']
                summary_cell.value = "SUMMARY"
                summary_cell.font = Font(bold=True, size=11, color="FFFFFF")
                summary_cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
                row += 1
                
                for key, value in report["summary"].items():
                    key_cell = ws.cell(row=row, column=1)
                    key_cell.value = key.replace("_", " ").title()
                    key_cell.font = Font(bold=True)
                    
                    val_cell = ws.cell(row=row, column=2)
                    if isinstance(value, float):
                        val_cell.value = value
                        val_cell.number_format = currency_format
                    else:
                        val_cell.value = value
                    
                    row += 1
            
            # Adjust column widths
            for col_idx in range(1, len(report["headers"]) + 1):
                ws.column_dimensions[get_column_letter(col_idx)].width = 15
            
            wb.save(file_path)
            error_logger.info(f"Excel report exported to {file_path}")
            return True
        
        except Exception as e:
            error_logger.error(f"Failed to export Excel report: {str(e)}")
            return False
    
    @staticmethod
    def export_pdf(report: Dict, file_path: str) -> bool:
        """
        Export report to PDF
        
        Returns:
            True if successful, False otherwise
        """
        if not HAS_REPORTLAB:
            error_logger.error("reportlab not installed, cannot export to PDF")
            return False
        
        try:
            # Use A3 landscape for wider tables
            page_size = landscape(A3)
            doc = SimpleDocTemplate(file_path, pagesize=page_size, leftMargin=0.5*inch, rightMargin=0.5*inch, topMargin=0.5*inch, bottomMargin=0.5*inch)
            story = []
            styles = getSampleStyleSheet()
            
            # Custom styles
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=16,
                textColor=colors.HexColor('#1F4E78'),
                spaceAfter=6,
                alignment=1  # Center
            )
            
            meta_style = ParagraphStyle(
                'Meta',
                parent=styles['Normal'],
                fontSize=9,
                textColor=colors.grey,
                spaceAfter=3,
                alignment=1  # Center
            )
            
            # Add title
            story.append(Paragraph(report.get("title", "Report"), title_style))
            story.append(Paragraph(f"Generated: {report.get('date_generated', '')}", meta_style))
            
            if "date_range" in report:
                story.append(Paragraph(f"Date Range: {report['date_range']}", meta_style))
            
            story.append(Spacer(1, 0.3 * inch))
            
            # Create table with A3 landscape width
            table_data = [report["headers"]] + report["rows"]
            # Calculate column width based on available width in landscape A3
            available_width = page_size[0] - 1 * inch  # Subtract margins
            col_width = available_width / len(report["headers"])
            
            table = Table(table_data, colWidths=[col_width] * len(report["headers"]))
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#D9E1F2')]),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 9),
            ]))
            
            story.append(table)
            
            # Add summary if available
            if "summary" in report:
                story.append(Spacer(1, 0.3 * inch))
                story.append(Paragraph("SUMMARY", styles['Heading3']))
                
                summary_data = [["Metric", "Value"]]
                for key, value in report["summary"].items():
                    key_text = key.replace("_", " ").title()
                    if isinstance(value, float):
                        value_text = f"₦{value:,.2f}"
                    else:
                        value_text = str(value)
                    summary_data.append([key_text, value_text])
                
                summary_table = Table(summary_data, colWidths=[2.5 * inch, 2.5 * inch])
                summary_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#70AD47')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black),
                    ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#E2EFD9')]),
                ]))
                
                story.append(summary_table)
            
            doc.build(story)
            error_logger.info(f"PDF report exported to {file_path}")
            return True
        
        except Exception as e:
            error_logger.error(f"Failed to export PDF report: {str(e)}")
            return False


class ScheduledReport:
    """Configuration for scheduled reports"""
    
    def __init__(self, name: str, report_type: ReportType, export_format: ExportFormat,
                frequency: str = "daily", email_recipients: List[str] = None):
        """
        Initialize scheduled report
        
        Args:
            name: Report name
            report_type: Type of report
            export_format: Export format
            frequency: "daily", "weekly", "monthly"
            email_recipients: Email addresses for delivery
        """
        self.name = name
        self.report_type = report_type
        self.export_format = export_format
        self.frequency = frequency
        self.email_recipients = email_recipients or []
        self.last_run = None
        self.next_run = self._calculate_next_run()
    
    def _calculate_next_run(self) -> datetime:
        """Calculate next run time"""
        now = datetime.now()
        
        if self.frequency == "daily":
            return now + timedelta(days=1)
        elif self.frequency == "weekly":
            return now + timedelta(weeks=1)
        elif self.frequency == "monthly":
            # Next month, same day
            if now.month == 12:
                return datetime(now.year + 1, 1, now.day)
            else:
                return datetime(now.year, now.month + 1, now.day)
        
        return now + timedelta(days=1)
    
    def should_run(self) -> bool:
        """Check if report should run now"""
        return datetime.now() >= self.next_run
    
    def mark_as_run(self):
        """Mark report as run"""
        self.last_run = datetime.now()
        self.next_run = self._calculate_next_run()
