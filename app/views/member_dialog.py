import flet as ft
import csv
import os
from datetime import datetime
from database.connection import (
    get_all_members,
    create_member,
    update_member,
    delete_member,
    get_contributions_by_member,
    get_loans_by_member,
    get_member_summary_stats,
    get_repayments_by_loan,
    get_refunds_by_loan,
    get_topups_by_loan,
    process_refund,
)
from database.models import MemberStatus
from components.navigation import create_app_bar
from components.error_handler import error_logger, UserFriendlyError
from components.ui_components import (
    ToastNotification, NotificationType, ConfirmDialog, UndoManager
)


def navigate_to(page: ft.Page, route: str):
    """Navigate to a different route (replaces deprecated page.go)"""
    page.route = route
    page.update()
from components.burger_menu import create_sidebar_overlay, create_burger_menu


def MemberScreen(page: ft.Page):
    """Members management screen with DataTable and dialogs"""
    
    # Undo manager for member operations
    undo_manager = UndoManager()
    
    # State management
    members_list = get_all_members()
    
    # State for edit mode
    edit_member_mode = {"enabled": False, "member_id": None}
    
    # State for search, filter, sort, and pagination
    active_status_filter = {"value": "All"}
    sort_state = {"column": "name", "ascending": True}
    current_search_term = {"value": ""}
    pagination_state = {"page": 1, "per_page": 20}
    filtered_members_cache = {"list": []}  # cache for export
    
    # Dialog for creating/editing member
    member_dialog = ft.AlertDialog(
        title=ft.Text("Add New Member"),
        content=ft.Column(
            controls=[
                ft.TextField(label="Full Name", width=400),
                ft.TextField(label="IPPIS Number", width=400),
                ft.TextField(label="Contact Number", width=400),
                ft.TextField(label="Email Address", width=400),
                ft.Dropdown(
                    label="Status",
                    options=[
                        ft.dropdown.Option("ACTIVE", "ACTIVE"),
                        ft.dropdown.Option("INACTIVE", "INACTIVE"),
                        ft.dropdown.Option("SUSPENDED", "SUSPENDED"),
                    ],
                    value="ACTIVE",
                    width=400,
                ),
            ],
            width=500,
            spacing=10,
        ),
        actions=[
            ft.TextButton("Cancel", on_click=lambda e: close_member_dialog()),
            ft.TextButton("Save", on_click=lambda e: save_member()),
        ],
    )
    
    # Dialog for viewing member details (ledger view)
    details_dialog = ft.AlertDialog(
        title=ft.Text("Member Ledger"),
        content=ft.Column(
            controls=[],
            width=750,
            height=550,
            scroll=ft.ScrollMode.AUTO,
            spacing=10,
        ),
        actions=[
            ft.TextButton("Close", on_click=lambda e: close_details_dialog()),
        ],
    )
    
    def close_member_dialog():
        member_dialog.open = False
        page.update()
    
    def close_details_dialog():
        details_dialog.open = False
        page.update()
    
    def _check_duplicate(name, ippis_number, exclude_id=None):
        """Check if a member with the same name or IPPIS already exists.
        Returns a warning message string, or None if no duplicate."""
        for m in members_list:
            if exclude_id and m.id == exclude_id:
                continue
            if name and m.name and m.name.strip().lower() == name.strip().lower():
                return f"A member with the name '{m.name}' already exists."
            if ippis_number and m.ippis_number and m.ippis_number.strip().lower() == ippis_number.strip().lower():
                return f"A member with IPPIS '{m.ippis_number}' already exists."
        return None
    
    def save_member():
        """Save member with comprehensive error handling, validation, and duplicate detection"""
        try:
            # Extract form values
            name = member_dialog.content.controls[0].value
            ippis_number = member_dialog.content.controls[1].value
            contact = member_dialog.content.controls[2].value
            email = member_dialog.content.controls[3].value
            status = member_dialog.content.controls[4].value or "ACTIVE"
            
            # Validate required name field
            if not name or not name.strip():
                error_logger.warning("Member save: name field is empty")
                error_msg = UserFriendlyError.get_message('missing_required_field', "Member name is required")
                ToastNotification.show(page, error_msg, NotificationType.ERROR)
                return
            
            # Validate status
            if status not in ["ACTIVE", "INACTIVE", "SUSPENDED"]:
                error_logger.warning(f"Invalid member status: {status}")
                error_msg = UserFriendlyError.get_message('validation_error', f"Invalid status: {status}")
                ToastNotification.show(page, error_msg, NotificationType.ERROR)
                return
            
            # --- Duplicate detection ---
            exclude_id = edit_member_mode["member_id"] if edit_member_mode["enabled"] else None
            dup_warning = _check_duplicate(name.strip(), ippis_number.strip() if ippis_number else None, exclude_id)
            if dup_warning:
                ToastNotification.show(page, f"⚠ {dup_warning}", NotificationType.WARNING)
                # Still allow save — it's a warning, not a block
            
            # Get member status enum
            member_status = MemberStatus[status]
            
            # Save member
            if edit_member_mode["enabled"]:
                # Update existing member
                error_logger.info(f"Updating member {edit_member_mode['member_id']}: name={name}")
                update_member(
                    edit_member_mode["member_id"],
                    name=name.strip(),
                    ippis_number=ippis_number.strip() if ippis_number else None,
                    contact=contact.strip() if contact else None,
                    email=email.strip() if email else None,
                    status=member_status
                )
                error_logger.info(f"Member {edit_member_mode['member_id']} updated successfully")
                ToastNotification.show(page, f"✓ Member '{name}' updated successfully!", NotificationType.SUCCESS)
                edit_member_mode["enabled"] = False
                edit_member_mode["member_id"] = None
            else:
                # Create new member
                error_logger.info(f"Creating new member: name={name}")
                create_member(
                    name.strip(),
                    contact.strip() if contact else None,
                    email.strip() if email else None,
                    ippis_number.strip() if ippis_number else None,
                    member_status
                )
                error_logger.info(f"New member created: {name}")
                ToastNotification.show(page, f"✓ Member '{name}' added successfully!", NotificationType.SUCCESS)
            
            close_member_dialog()
            refresh_members()
            page.update()
        
        except Exception as e:
            error_logger.exception(f"Error saving member: {str(e)}")
            error_msg = UserFriendlyError.get_message('db_commit_error', str(e))
            ToastNotification.show(page, f"✗ {error_msg}", NotificationType.ERROR)
            page.update()
    
    def refresh_members():
        nonlocal members_list
        members_list = get_all_members()
        filter_members_table(current_search_term["value"])
    
    def _get_sort_key(member, stats, col, ascending):
        """Return sort value for a given column."""
        member_stats = stats.get(member.id, {'total_contributions': 0.0, 'active_loans': 0})
        if col == "name":
            return (member.name or "").lower()
        elif col == "ippis":
            return (member.ippis_number or "").lower()
        elif col == "contact":
            return (member.contact or "").lower()
        elif col == "email":
            return (member.email or "").lower()
        elif col == "status":
            return (member.status.value or "").lower()
        elif col == "contributions":
            return member_stats['total_contributions']
        elif col == "active_loans":
            return member_stats['active_loans']
        elif col == "join_date":
            return member.join_date or datetime.min
        return (member.name or "").lower()
    
    def on_column_sort(col_key):
        """Handle click-to-sort on column headers."""
        if sort_state["column"] == col_key:
            sort_state["ascending"] = not sort_state["ascending"]
        else:
            sort_state["column"] = col_key
            sort_state["ascending"] = True
        update_column_headers()
        filter_members_table(current_search_term["value"])
    
    def on_filter_tab_change(e):
        """Handle status filter segment change."""
        selected = e.control.selected
        if selected:
            active_status_filter["value"] = list(selected)[0]
        filter_members_table(current_search_term["value"])
    
    def on_search_change(e):
        """Handle search text field change."""
        current_search_term["value"] = e.control.value or ""
        filter_members_table(current_search_term["value"])
    
    def on_search_clear(e):
        """Clear search field."""
        search_field.value = ""
        current_search_term["value"] = ""
        filter_members_table("")
        page.update()
    
    def on_page_change(delta):
        """Handle pagination page change."""
        pagination_state["page"] += delta
        if pagination_state["page"] < 1:
            pagination_state["page"] = 1
        filter_members_table(current_search_term["value"])
    
    def on_per_page_change(e):
        """Handle rows-per-page dropdown change."""
        pagination_state["per_page"] = int(e.control.value)
        pagination_state["page"] = 1
        filter_members_table(current_search_term["value"])
    
    def filter_members_table(search_term=""):
        """Main function: filter by status, search, sort, paginate, and render table rows."""
        filtered = list(members_list)
        
        # --- Status filter ---
        status_filter = active_status_filter["value"]
        if status_filter == "Active":
            filtered = [m for m in filtered if m.status == MemberStatus.ACTIVE]
        elif status_filter == "Inactive":
            filtered = [m for m in filtered if m.status == MemberStatus.INACTIVE]
        elif status_filter == "Suspended":
            filtered = [m for m in filtered if m.status == MemberStatus.SUSPENDED]
        # "All" → no filter
        
        # --- Search filter ---
        if search_term.strip():
            term = search_term.strip().lower()
            filtered = [
                m for m in filtered
                if term in (m.name or "").lower()
                or term in (m.ippis_number or "").lower()
                or term in (m.contact or "").lower()
                or term in (m.email or "").lower()
            ]
        
        # --- Batch stats query ---
        stats = get_member_summary_stats()
        
        # --- Sort ---
        col = sort_state["column"]
        asc = sort_state["ascending"]
        try:
            filtered.sort(
                key=lambda m: _get_sort_key(m, stats, col, asc),
                reverse=not asc
            )
        except TypeError:
            pass
        
        # Cache filtered list for export
        filtered_members_cache["list"] = filtered
        
        total_filtered = len(filtered)
        
        # --- Pagination ---
        per_page = pagination_state["per_page"]
        total_pages = max(1, (total_filtered + per_page - 1) // per_page)
        if pagination_state["page"] > total_pages:
            pagination_state["page"] = total_pages
        current_page = pagination_state["page"]
        start_idx = (current_page - 1) * per_page
        end_idx = start_idx + per_page
        page_members = filtered[start_idx:end_idx]
        
        # --- Build rows ---
        rows = []
        for member in page_members:
            member_stats = stats.get(member.id, {'total_contributions': 0.0, 'active_loans': 0})
            total_contrib = member_stats['total_contributions']
            active_loan_count = member_stats['active_loans']
            
            # Status color
            if member.status == MemberStatus.ACTIVE:
                status_color = ft.Colors.GREEN_400
            elif member.status == MemberStatus.INACTIVE:
                status_color = ft.Colors.GREY_400
            else:  # SUSPENDED
                status_color = ft.Colors.RED_400
            
            # Row background: dim suspended members
            row_color = None
            if member.status == MemberStatus.SUSPENDED:
                row_color = "#3d1f1f"
            elif member.status == MemberStatus.INACTIVE:
                row_color = "#2a2a2a"
            
            # Join date formatted
            join_date_str = member.join_date.strftime("%d %b %Y") if member.join_date else "N/A"
            
            rows.append(
                ft.DataRow(
                    color=row_color,
                    cells=[
                        ft.DataCell(
                            ft.Row(
                                [ft.Icon(ft.Icons.FLAG, color=ft.Colors.RED_400, size=16, tooltip="Flagged: Overdue 90+ days"),
                                 ft.Text(member.name, color=ft.Colors.WHITE)],
                                spacing=4, vertical_alignment=ft.CrossAxisAlignment.CENTER,
                            ) if member.is_flagged else ft.Text(member.name, color=ft.Colors.WHITE)
                        ),
                        ft.DataCell(ft.Text(member.ippis_number or "N/A", color=ft.Colors.GREY_400)),
                        ft.DataCell(ft.Text(member.contact or "N/A", color=ft.Colors.GREY)),
                        ft.DataCell(ft.Text(member.email or "N/A", color=ft.Colors.GREY)),
                        ft.DataCell(ft.Text(member.status.value, color=status_color, weight="bold")),
                        ft.DataCell(ft.Text(join_date_str, color=ft.Colors.GREY_400)),
                        ft.DataCell(ft.Text(f"₦{total_contrib:,.2f}", color=ft.Colors.GREEN_400, weight="bold")),
                        ft.DataCell(ft.Text(str(active_loan_count), color=ft.Colors.ORANGE_400)),
                        ft.DataCell(
                            ft.Row(
                                controls=[
                                    ft.IconButton(
                                        ft.Icons.VISIBILITY,
                                        tooltip="View Details",
                                        on_click=lambda e, m=member: view_member_details(m),
                                        icon_size=20,
                                    ),
                                    ft.IconButton(
                                        ft.Icons.EDIT,
                                        tooltip="Edit Member",
                                        on_click=lambda e, m=member: edit_member(m),
                                        icon_size=20,
                                    ),
                                    ft.IconButton(
                                        ft.Icons.DELETE,
                                        tooltip="Delete",
                                        on_click=lambda e, m=member: delete_member_record(m),
                                        icon_size=20,
                                    ),
                                ],
                                spacing=0,
                            )
                        ),
                    ]
                )
            )
        
        members_table.rows = rows
        
        # Update member count indicator
        showing_start = start_idx + 1 if total_filtered > 0 else 0
        showing_end = min(end_idx, total_filtered)
        member_count_text.value = f"Showing {showing_start}-{showing_end} of {total_filtered} members (Page {current_page}/{total_pages})"
        
        # Update pagination buttons
        btn_prev.disabled = current_page <= 1
        btn_next.disabled = current_page >= total_pages
        
        # Update stats and filter counts
        update_summary_stats(stats)
        update_filter_tab_counts()
        
        page.update()
    
    def update_summary_stats(stats=None):
        """Update the 4 summary stat card values."""
        if stats is None:
            stats = get_member_summary_stats()
        
        total_members = len(members_list)
        active_members = sum(1 for m in members_list if m.status == MemberStatus.ACTIVE)
        total_contributions = sum(s['total_contributions'] for s in stats.values())
        total_active_loans = sum(s['active_loans'] for s in stats.values())
        
        stat_total_members.value = str(total_members)
        stat_active_members.value = str(active_members)
        stat_total_contributions.value = f"₦{total_contributions:,.2f}"
        stat_active_loans.value = str(total_active_loans)
    
    def update_filter_tab_counts():
        """Update segment labels with counts."""
        counts = {"All": 0, "Active": 0, "Inactive": 0, "Suspended": 0}
        for m in members_list:
            counts["All"] += 1
            if m.status == MemberStatus.ACTIVE:
                counts["Active"] += 1
            elif m.status == MemberStatus.INACTIVE:
                counts["Inactive"] += 1
            elif m.status == MemberStatus.SUSPENDED:
                counts["Suspended"] += 1
        
        seg_all.label = ft.Text(f"All ({counts['All']})")
        seg_active.label = ft.Text(f"Active ({counts['Active']})")
        seg_inactive.label = ft.Text(f"Inactive ({counts['Inactive']})")
        seg_suspended.label = ft.Text(f"Suspended ({counts['Suspended']})")
    
    def update_column_headers():
        """Update sort arrows on column headers."""
        col = sort_state["column"]
        asc = sort_state["ascending"]
        arrow = " ▲" if asc else " ▼"
        
        for key, dc in sortable_columns.items():
            label = column_labels[key]
            if key == col:
                dc.label = ft.Text(f"{label}{arrow}", color=ft.Colors.BLUE_200, weight="bold")
            else:
                dc.label = ft.Text(label, color=ft.Colors.WHITE, weight="bold")
    
    def export_members_csv(e):
        """Export the currently filtered member list to CSV."""
        try:
            stats = get_member_summary_stats()
            filtered = filtered_members_cache.get("list", members_list)
            
            downloads_dir = os.path.join(os.path.expanduser("~"), "Downloads")
            os.makedirs(downloads_dir, exist_ok=True)
            
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filepath = os.path.join(downloads_dir, f"members_{timestamp}.csv")
            
            with open(filepath, "w", newline="", encoding="utf-8") as f:
                writer = csv.writer(f)
                writer.writerow(["Name", "IPPIS", "Contact", "Email", "Status", "Join Date", "Total Contributions", "Active Loans"])
                for m in filtered:
                    ms = stats.get(m.id, {'total_contributions': 0.0, 'active_loans': 0})
                    join_str = m.join_date.strftime("%d %b %Y") if m.join_date else "N/A"
                    writer.writerow([
                        m.name,
                        m.ippis_number or "N/A",
                        m.contact or "N/A",
                        m.email or "N/A",
                        m.status.value,
                        join_str,
                        f"{ms['total_contributions']:.2f}",
                        ms['active_loans'],
                    ])
            
            filename = os.path.basename(filepath)
            error_logger.info(f"Exported {len(filtered)} members to {filepath}")
            ToastNotification.show(page, f"✓ Exported {len(filtered)} members to Downloads/{filename}", NotificationType.SUCCESS)
        except Exception as ex:
            error_logger.exception(f"CSV export failed: {ex}")
            ToastNotification.show(page, f"Export failed: {ex}", NotificationType.ERROR)
    
    def export_members_excel(e):
        """Export the currently filtered member list to Excel (.xlsx)."""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            
            stats = get_member_summary_stats()
            filtered = filtered_members_cache.get("list", members_list)
            
            downloads_dir = os.path.join(os.path.expanduser("~"), "Downloads")
            os.makedirs(downloads_dir, exist_ok=True)
            
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filepath = os.path.join(downloads_dir, f"members_{timestamp}.xlsx")
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Members"
            
            # Header row
            headers = ["Name", "IPPIS", "Contact", "Email", "Status", "Join Date", "Total Contributions", "Active Loans"]
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="1a237e", end_color="1a237e", fill_type="solid")
            
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
            
            # Data rows
            for row_idx, m in enumerate(filtered, 2):
                ms = stats.get(m.id, {'total_contributions': 0.0, 'active_loans': 0})
                join_str = m.join_date.strftime("%d %b %Y") if m.join_date else "N/A"
                ws.cell(row=row_idx, column=1, value=m.name)
                ws.cell(row=row_idx, column=2, value=m.ippis_number or "N/A")
                ws.cell(row=row_idx, column=3, value=m.contact or "N/A")
                ws.cell(row=row_idx, column=4, value=m.email or "N/A")
                ws.cell(row=row_idx, column=5, value=m.status.value)
                ws.cell(row=row_idx, column=6, value=join_str)
                ws.cell(row=row_idx, column=7, value=round(ms['total_contributions'], 2))
                ws.cell(row=row_idx, column=8, value=ms['active_loans'])
            
            # Auto-fit column widths
            for col in ws.columns:
                max_len = 0
                col_letter = col[0].column_letter
                for cell in col:
                    try:
                        if cell.value:
                            max_len = max(max_len, len(str(cell.value)))
                    except:
                        pass
                ws.column_dimensions[col_letter].width = min(max_len + 3, 40)
            
            wb.save(filepath)
            filename = os.path.basename(filepath)
            error_logger.info(f"Exported {len(filtered)} members to {filepath}")
            ToastNotification.show(page, f"✓ Exported {len(filtered)} members to Downloads/{filename}", NotificationType.SUCCESS)
        except Exception as ex:
            error_logger.exception(f"Excel export failed: {ex}")
            ToastNotification.show(page, f"Export failed: {ex}", NotificationType.ERROR)
    
    def view_member_details(member):
        """Show comprehensive member ledger — like an Excel sheet per member."""
        from collections import defaultdict
        from datetime import datetime as dt

        contributions = get_contributions_by_member(member.id)
        loans = get_loans_by_member(member.id)

        # Gather all repayments, refunds, topups across all loans
        all_repayments = []   # (loan, repayment)
        all_refunds = []      # (loan, refund)
        all_topups = []       # (loan, topup)
        for loan in loans:
            for r in get_repayments_by_loan(loan.id):
                all_repayments.append((loan, r))
            for rf in get_refunds_by_loan(loan.id):
                all_refunds.append((loan, rf))
            for tu in get_topups_by_loan(loan.id):
                all_topups.append((loan, tu))

        # ---- Build month-keyed ledger ----
        # Each month: {savings, loans_taken, deductions, topups, refunds, events}
        ledger = defaultdict(lambda: {
            "savings": 0.0,
            "loans_taken": 0.0,
            "deductions": 0.0,
            "topups": 0.0,
            "refunds": 0.0,
            "events": [],
        })

        # Contributions (savings)
        for c in contributions:
            d = c.contribution_date or c.created_at
            if d:
                mk = d.strftime("%Y-%m")
                if c.amount > 0:
                    ledger[mk]["savings"] += c.amount
                    ledger[mk]["events"].append(("Saving", c.amount, None))
                else:
                    ledger[mk]["events"].append(("Withdrawal", c.amount, None))

        # Loans taken
        for loan in loans:
            d = loan.start_date or loan.created_at
            if d:
                mk = d.strftime("%Y-%m")
                ledger[mk]["loans_taken"] += loan.amount + loan.total_interest + (getattr(loan, "overdue_penalty", 0.0) or 0.0)
                ledger[mk]["events"].append(("Loan Taken", loan.amount, f"#{loan.id} + Interest ₦{loan.total_interest:,.2f}"))

        # Repayments (deductions from member)
        for loan, r in all_repayments:
            d = r.payment_date or r.created_at
            if d:
                mk = d.strftime("%Y-%m")
                ledger[mk]["deductions"] += r.amount_paid
                ledger[mk]["events"].append(("Repayment", r.amount_paid, f"Loan #{loan.id}"))

        # Top-ups
        for loan, tu in all_topups:
            d = tu.topup_date or tu.created_at
            if d:
                mk = d.strftime("%Y-%m")
                ledger[mk]["topups"] += tu.topup_amount + tu.interest_on_topup
                ledger[mk]["events"].append(("Top-Up", tu.topup_amount, f"Loan #{loan.id} + Interest ₦{tu.interest_on_topup:,.2f}"))

        # Refunds (overpayment refunds back to member)
        for loan, rf in all_refunds:
            d = rf.refund_date or rf.created_at
            if d:
                mk = d.strftime("%Y-%m")
                ledger[mk]["refunds"] += rf.refund_amount
                ledger[mk]["events"].append(("Refund", rf.refund_amount, f"Loan #{loan.id} ({rf.status})"))

        # Sort months chronologically
        sorted_months = sorted(ledger.keys())

        # Compute running totals
        running_equity = 0.0
        running_loan_balance = 0.0
        ledger_rows = []
        for mk in sorted_months:
            entry = ledger[mk]
            running_equity += entry["savings"]
            running_loan_balance += entry["loans_taken"] + entry["topups"]
            running_loan_balance -= entry["deductions"]
            running_loan_balance = max(0, running_loan_balance)

            # Format month
            try:
                month_label = dt.strptime(mk, "%Y-%m").strftime("%b %Y")
            except Exception:
                month_label = mk

            ledger_rows.append({
                "month": month_label,
                "savings": entry["savings"],
                "loan_taken": entry["loans_taken"] + entry["topups"],
                "deduction": entry["deductions"],
                "refund": entry["refunds"],
                "loan_balance": running_loan_balance,
                "equity": running_equity,
                "events": entry["events"],
            })

        # ---- Summary stats ----
        total_savings = sum(c.amount for c in contributions if c.amount > 0)
        total_loans = sum(l.amount + l.total_interest + (getattr(l, "overdue_penalty", 0.0) or 0.0) for l in loans) + sum(tu.topup_amount + tu.interest_on_topup for _, tu in all_topups)
        total_repaid = sum(r.amount_paid for _, r in all_repayments)
        outstanding = max(0, total_loans - total_repaid)
        active_loans = [l for l in loans if l.status.value == "Active"]

        # Monthly deduction: sum of monthly payment for active loans
        monthly_deduction = 0.0
        for loan in active_loans:
            total_due = loan.amount + loan.total_interest + (getattr(loan, "overdue_penalty", 0.0) or 0.0)
            duration = 12
            if loan.end_date and loan.start_date:
                duration = max(1, (loan.end_date - loan.start_date).days // 30)
            monthly_deduction += total_due / duration if duration > 0 else 0

        # Pending refunds
        pending_refunds = [(loan, rf) for loan, rf in all_refunds if rf.status == "PENDING"]

        join_str = member.join_date.strftime("%d %b %Y") if member.join_date else "N/A"

        # ---- Build UI ----
        def _stat_card(label, value, color):
            return ft.Container(
                content=ft.Column([
                    ft.Text(label, size=14, color=ft.Colors.GREY_400),
                    ft.Text(value, size=17, weight="bold", color=color),
                ], spacing=2, horizontal_alignment=ft.CrossAxisAlignment.CENTER),
                bgcolor="#2a2a2a",
                border_radius=8,
                padding=ft.padding.symmetric(horizontal=12, vertical=8),
                expand=True,
            )

        stats_row = ft.Row(
            controls=[
                _stat_card("Total Equity", f"₦{total_savings:,.2f}", ft.Colors.GREEN_400),
                _stat_card("Total Loans", f"₦{total_loans:,.2f}", ft.Colors.RED_400),
                _stat_card("Total Repaid", f"₦{total_repaid:,.2f}", ft.Colors.BLUE_400),
                _stat_card("Outstanding", f"₦{outstanding:,.2f}", ft.Colors.ORANGE_400),
                _stat_card("Monthly Deduction", f"₦{monthly_deduction:,.2f}", ft.Colors.AMBER_400),
            ],
            spacing=8,
        )

        # Ledger DataTable
        ledger_columns = [
            ft.DataColumn(ft.Text("DATE", size=15, weight="bold")),
            ft.DataColumn(ft.Text("AMOUNT\n(Savings)", size=14, weight="bold"), numeric=True),
            ft.DataColumn(ft.Text("LOAN/\nDEDUCTION", size=14, weight="bold"), numeric=True),
            ft.DataColumn(ft.Text("LOAN\nREFUND", size=14, weight="bold"), numeric=True),
            ft.DataColumn(ft.Text("LOAN\nBALANCE", size=14, weight="bold"), numeric=True),
            ft.DataColumn(ft.Text("EQUITY", size=14, weight="bold"), numeric=True),
        ]

        ledger_table_rows = []
        for row in ledger_rows:
            ledger_table_rows.append(ft.DataRow(
                cells=[
                    ft.DataCell(ft.Text(row["month"], size=15, color=ft.Colors.WHITE)),
                    ft.DataCell(ft.Text(
                        f"₦{row['savings']:,.2f}" if row["savings"] else "-",
                        size=15, color=ft.Colors.GREEN_400 if row["savings"] else ft.Colors.GREY_600,
                    )),
                    ft.DataCell(ft.Text(
                        f"₦{row['deduction']:,.2f}" if row["deduction"] else "-",
                        size=15, color=ft.Colors.RED_400 if row["deduction"] else ft.Colors.GREY_600,
                    )),
                    ft.DataCell(ft.Text(
                        f"₦{row['refund']:,.2f}" if row["refund"] else "-",
                        size=15, color=ft.Colors.ORANGE_400 if row["refund"] else ft.Colors.GREY_600,
                    )),
                    ft.DataCell(ft.Text(
                        f"₦{row['loan_balance']:,.2f}",
                        size=15, color=ft.Colors.RED_300 if row["loan_balance"] > 0 else ft.Colors.GREEN_400,
                    )),
                    ft.DataCell(ft.Text(
                        f"₦{row['equity']:,.2f}",
                        size=15, color=ft.Colors.GREEN_400,
                    )),
                ],
            ))

        if not ledger_table_rows:
            ledger_table_rows.append(ft.DataRow(cells=[
                ft.DataCell(ft.Text("No records", color=ft.Colors.GREY_400, italic=True)),
                ft.DataCell(ft.Text("-")), ft.DataCell(ft.Text("-")),
                ft.DataCell(ft.Text("-")), ft.DataCell(ft.Text("-")),
                ft.DataCell(ft.Text("-")),
            ]))

        ledger_table = ft.DataTable(
            columns=ledger_columns,
            rows=ledger_table_rows,
            border=ft.border.all(1, ft.Colors.GREY_800),
            border_radius=8,
            heading_row_color=ft.Colors.with_opacity(0.1, ft.Colors.BLUE_400),
            data_row_max_height=44,
            column_spacing=18,
        )

        # ---- Refund section ----
        refund_controls = []
        if all_refunds:
            refund_controls.append(ft.Text("Refunds", weight="bold", color=ft.Colors.AMBER_400, size=18))
            for loan, rf in all_refunds:
                is_pending = rf.status == "PENDING"
                status_color = ft.Colors.ORANGE_400 if is_pending else ft.Colors.GREEN_400
                status_icon = ft.Icons.HOURGLASS_BOTTOM if is_pending else ft.Icons.CHECK_CIRCLE
                date_str = rf.refund_date.strftime("%d %b %Y") if rf.refund_date else "N/A"

                row_controls = [
                    ft.Icon(status_icon, color=status_color, size=20),
                    ft.Text(f"Loan #{loan.id}", size=15, color=ft.Colors.GREY_400),
                    ft.Text(f"₦{rf.refund_amount:,.2f}", size=16, weight="bold", color=ft.Colors.WHITE),
                    ft.Text(rf.status, size=15, color=status_color, weight="bold"),
                    ft.Text(date_str, size=15, color=ft.Colors.GREY_400),
                ]

                if is_pending:
                    def _show_refund_options(e, refund_id=rf.id, refund_amt=rf.refund_amount, loan_id=loan.id):
                        partial_field = ft.TextField(
                            label="Partial Amount (₦)",
                            keyboard_type=ft.KeyboardType.NUMBER,
                            width=220,
                            visible=False,
                            text_size=16,
                        )
                        error_text = ft.Text("", color=ft.Colors.RED_400, size=14, visible=False)

                        def on_option_change(ev):
                            partial_field.visible = (ev.control.value == "partial")
                            error_text.visible = False
                            page.update()

                        option_group = ft.RadioGroup(
                            content=ft.Column([
                                ft.Radio(value="full", label=f"Full Refund — ₦{refund_amt:,.2f}"),
                                ft.Radio(value="partial", label="Partial Refund"),
                            ], spacing=4),
                            value="full",
                            on_change=on_option_change,
                        )

                        def confirm_refund(ev):
                            try:
                                if option_group.value == "partial":
                                    raw = (partial_field.value or "").strip().replace(",", "")
                                    if not raw:
                                        error_text.value = "Enter an amount"
                                        error_text.visible = True
                                        page.update()
                                        return
                                    amt = float(raw)
                                    if amt <= 0:
                                        error_text.value = "Amount must be greater than 0"
                                        error_text.visible = True
                                        page.update()
                                        return
                                    if amt >= refund_amt:
                                        error_text.value = f"Must be less than ₦{refund_amt:,.2f} (use Full Refund instead)"
                                        error_text.visible = True
                                        page.update()
                                        return
                                    process_refund(refund_id, partial_amount=amt)
                                    ToastNotification.show(page, f"✓ Partial refund ₦{amt:,.2f} processed. Remaining ₦{refund_amt - amt:,.2f} still pending.", NotificationType.SUCCESS)
                                else:
                                    process_refund(refund_id)
                                    ToastNotification.show(page, f"✓ Full refund ₦{refund_amt:,.2f} processed!", NotificationType.SUCCESS)
                                refund_dialog.open = False
                                details_dialog.open = False
                                page.update()
                                view_member_details(member)
                            except ValueError:
                                error_text.value = "Enter a valid number"
                                error_text.visible = True
                                page.update()
                            except Exception as ex:
                                ToastNotification.show(page, f"Error: {str(ex)}", NotificationType.ERROR)
                                page.update()

                        def cancel_refund(ev):
                            refund_dialog.open = False
                            page.update()

                        refund_dialog = ft.AlertDialog(
                            modal=True,
                            title=ft.Text(f"Process Refund — Loan #{loan_id}", size=18, weight="bold"),
                            content=ft.Container(
                                content=ft.Column([
                                    ft.Text(f"Refund Amount: ₦{refund_amt:,.2f}", size=16, weight="bold", color=ft.Colors.AMBER_400),
                                    option_group,
                                    partial_field,
                                    error_text,
                                ], spacing=10, tight=True),
                                width=320,
                            ),
                            actions=[
                                ft.TextButton("Cancel", on_click=cancel_refund),
                                ft.ElevatedButton("Confirm", on_click=confirm_refund,
                                                  bgcolor=ft.Colors.GREEN_700, color=ft.Colors.WHITE),
                            ],
                            actions_alignment=ft.MainAxisAlignment.END,
                        )
                        page.overlay.append(refund_dialog)
                        refund_dialog.open = True
                        page.update()

                    row_controls.append(
                        ft.TextButton("Process Refund", on_click=_show_refund_options,
                                      style=ft.ButtonStyle(color=ft.Colors.GREEN_400))
                    )

                refund_controls.append(ft.Container(
                    content=ft.Row(row_controls, spacing=10, vertical_alignment=ft.CrossAxisAlignment.CENTER),
                    bgcolor="#2a2a2a",
                    border_radius=6,
                    padding=ft.padding.symmetric(horizontal=10, vertical=6),
                ))

        # ---- Assemble dialog content ----
        def export_ledger_excel(e):
            try:
                from openpyxl import Workbook
                from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

                downloads_dir = os.path.join(os.path.expanduser("~"), "Downloads")
                os.makedirs(downloads_dir, exist_ok=True)
                safe_name = "".join(c for c in member.name if c.isalnum() or c in " _-").strip().replace(" ", "_")
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                filename = f"Ledger_{safe_name}_{timestamp}.xlsx"
                filepath = os.path.join(downloads_dir, filename)

                wb = Workbook()
                ws = wb.active
                ws.title = "Member Ledger"

                # ---- Member Info Header ----
                info_font = Font(bold=True, size=14)
                ws.merge_cells('A1:F1')
                ws['A1'] = f"Member Ledger - {member.name}"
                ws['A1'].font = Font(bold=True, size=16)
                ws['A2'] = "IPPIS:"
                ws['A2'].font = info_font
                ws['B2'] = member.ippis_number or "N/A"
                ws['C2'] = "Contact:"
                ws['C2'].font = info_font
                ws['D2'] = member.contact or "N/A"
                ws['E2'] = "Email:"
                ws['E2'].font = info_font
                ws['F2'] = member.email or "N/A"
                ws['A3'] = "Status:"
                ws['A3'].font = info_font
                ws['B3'] = member.status.value
                ws['C3'] = "Joined:"
                ws['C3'].font = info_font
                ws['D3'] = join_str

                # ---- Summary Row ----
                summary_fill = PatternFill(start_color="1a237e", end_color="1a237e", fill_type="solid")
                summary_font = Font(bold=True, color="FFFFFF", size=13)
                summary_labels = ["Total Equity", "Total Loans", "Total Repaid", "Outstanding", "Monthly Deduction"]
                summary_values = [
                    f"\u20a6{total_savings:,.2f}", f"\u20a6{total_loans:,.2f}",
                    f"\u20a6{total_repaid:,.2f}", f"\u20a6{outstanding:,.2f}",
                    f"\u20a6{monthly_deduction:,.2f}",
                ]
                for col_idx, lbl in enumerate(summary_labels, 1):
                    cell = ws.cell(row=5, column=col_idx, value=lbl)
                    cell.font = summary_font
                    cell.fill = summary_fill
                    cell.alignment = Alignment(horizontal="center")
                for col_idx, val in enumerate(summary_values, 1):
                    cell = ws.cell(row=6, column=col_idx, value=val)
                    cell.alignment = Alignment(horizontal="center")
                    cell.font = Font(bold=True, size=13)

                # ---- Ledger Table ----
                headers = ["DATE", "AMOUNT (Savings)", "LOAN/DEDUCTION", "LOAN REFUND", "LOAN BALANCE", "EQUITY"]
                header_font = Font(bold=True, color="FFFFFF", size=13)
                header_fill = PatternFill(start_color="283593", end_color="283593", fill_type="solid")
                thin_border = Border(
                    left=Side(style="thin"), right=Side(style="thin"),
                    top=Side(style="thin"), bottom=Side(style="thin"),
                )

                header_row = 8
                for col_idx, h in enumerate(headers, 1):
                    cell = ws.cell(row=header_row, column=col_idx, value=h)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal="center")
                    cell.border = thin_border

                for r_idx, row in enumerate(ledger_rows, header_row + 1):
                    values = [
                        row["month"],
                        row["savings"] if row["savings"] else None,
                        row["deduction"] if row["deduction"] else None,
                        row["refund"] if row["refund"] else None,
                        row["loan_balance"],
                        row["equity"],
                    ]
                    for col_idx, val in enumerate(values, 1):
                        cell = ws.cell(row=r_idx, column=col_idx)
                        if col_idx == 1:
                            cell.value = val
                        elif val is not None:
                            cell.value = round(val, 2)
                            cell.number_format = '#,##0.00'
                        else:
                            cell.value = "-"
                        cell.border = thin_border
                        cell.alignment = Alignment(horizontal="center" if col_idx == 1 else "right")

                # ---- Refunds section ----
                if all_refunds:
                    refund_start = header_row + len(ledger_rows) + 2
                    ws.cell(row=refund_start, column=1, value="REFUNDS").font = Font(bold=True, size=14, color="FF8F00")
                    ref_headers = ["Loan ID", "Amount", "Status", "Date"]
                    for col_idx, rh in enumerate(ref_headers, 1):
                        cell = ws.cell(row=refund_start + 1, column=col_idx, value=rh)
                        cell.font = Font(bold=True)
                        cell.border = thin_border
                    for ri, (loan, rf) in enumerate(all_refunds, refund_start + 2):
                        ws.cell(row=ri, column=1, value=f"#{loan.id}").border = thin_border
                        c = ws.cell(row=ri, column=2, value=round(rf.refund_amount, 2))
                        c.number_format = '#,##0.00'
                        c.border = thin_border
                        ws.cell(row=ri, column=3, value=rf.status).border = thin_border
                        ws.cell(row=ri, column=4, value=rf.refund_date.strftime("%d %b %Y") if rf.refund_date else "N/A").border = thin_border

                # Auto-fit column widths
                for col in ws.columns:
                    max_len = 0
                    col_letter = col[0].column_letter
                    for cell in col:
                        if cell.value:
                            max_len = max(max_len, len(str(cell.value)))
                    ws.column_dimensions[col_letter].width = min(max_len + 4, 30)

                wb.save(filepath)
                ToastNotification.show(page, f"\u2713 Exported to Downloads/{filename}", NotificationType.SUCCESS)
            except Exception as ex:
                error_logger.exception(f"Ledger export failed: {ex}")
                ToastNotification.show(page, f"Export failed: {ex}", NotificationType.ERROR)
            page.update()

        content_controls = [
            # Header
            ft.Text(member.name, size=22, weight="bold", color=ft.Colors.WHITE),
            ft.Text(
                f"IPPIS: {member.ippis_number or 'N/A'}  |  Contact: {member.contact or 'N/A'}  |  Email: {member.email or 'N/A'}",
                size=16, color=ft.Colors.GREY_400,
            ),
            ft.Text(f"Status: {member.status.value}  |  Joined: {join_str}", size=16, color=ft.Colors.GREY_400),
            ft.Row([
                ft.Container(expand=True),
                ft.IconButton(
                    icon=ft.Icons.DOWNLOAD,
                    tooltip="Export Ledger to Excel",
                    icon_color=ft.Colors.GREEN_400,
                    icon_size=22,
                    on_click=export_ledger_excel,
                ),
            ]),
            ft.Divider(height=10),
            # Summary cards
            stats_row,
            ft.Divider(height=10),
            # Pending refund alert
        ]

        if pending_refunds:
            total_pending = sum(rf.refund_amount for _, rf in pending_refunds)
            content_controls.append(ft.Container(
                content=ft.Row([
                    ft.Icon(ft.Icons.WARNING_AMBER, color=ft.Colors.AMBER_400, size=22),
                    ft.Text(f"Pending Refund: ₦{total_pending:,.2f} ({len(pending_refunds)} refund{'s' if len(pending_refunds) > 1 else ''})",
                            size=17, weight="bold", color=ft.Colors.AMBER_400),
                ], spacing=8),
                bgcolor="#3d3a1f",
                border_radius=6,
                padding=10,
            ))

        content_controls.append(ft.Text("Member Ledger", weight="bold", color=ft.Colors.BLUE_400, size=18))

        # Scrollable table wrapper
        content_controls.append(ft.Container(
            content=ft.Column([ledger_table], scroll=ft.ScrollMode.AUTO),
            height=280 if len(ledger_table_rows) > 6 else None,
        ))

        # Refund section
        if refund_controls:
            content_controls.append(ft.Divider(height=10))
            content_controls.extend(refund_controls)

        details_dialog.content.controls = content_controls
        details_dialog.content.width = 800
        details_dialog.content.height = 580

        details_dialog.open = True
        page.update()
    
    def edit_member(member):
        """Open dialog to edit member details"""
        edit_member_mode["enabled"] = True
        edit_member_mode["member_id"] = member.id
        
        member_dialog.title.value = f"Edit Member - {member.name}"
        member_dialog.content.controls[0].value = member.name
        member_dialog.content.controls[1].value = member.ippis_number or ""
        member_dialog.content.controls[2].value = member.contact or ""
        member_dialog.content.controls[3].value = member.email or ""
        member_dialog.content.controls[4].value = member.status.value
        
        member_dialog.open = True
        page.update()
    
    def delete_member_record(member):
        """Delete member with modal confirmation and undo capability"""
        
        # Store member data for undo
        member_backup = {
            'id': member.id,
            'name': member.name,
            'contact': member.contact,
            'email': member.email,
            'ippis_number': member.ippis_number,
            'status': member.status
        }
        
        def perform_delete():
            try:
                error_logger.info(f"Deleting member {member.id}: {member.name}")
                delete_member(member.id)
                
                # Record undo action
                def undo_delete():
                    from database.models import MemberStatus
                    create_member(
                        name=member_backup['name'],
                        contact=member_backup['contact'],
                        email=member_backup['email'],
                        ippis_number=member_backup['ippis_number'],
                        status=member_backup['status']
                    )
                
                def redo_delete():
                    delete_member(member.id)
                
                undo_manager.record_action(f"Delete member '{member.name}'", undo_delete, redo_delete)
                
                refresh_members()
                error_logger.info(f"Member {member.id} deleted successfully")
                ToastNotification.show(page, f"✓ Member '{member.name}' deleted!", NotificationType.SUCCESS)
            except Exception as e:
                error_logger.exception(f"Error deleting member {member.id}: {str(e)}")
                error_msg = UserFriendlyError.get_message('db_commit_error', str(e))
                ToastNotification.show(page, error_msg, NotificationType.ERROR)
        
        # Show modal confirmation dialog
        confirm_dialog = ConfirmDialog(
            title="Delete Member",
            content=f"Are you sure you want to delete member '{member.name}'?\nThis action can be undone.",
            confirm_text="Delete",
            cancel_text="Cancel",
            danger=True
        )
        confirm_dialog.on_confirm(perform_delete)
        confirm_dialog.show(page)
    
    # --- Summary Stat Cards ---
    def _make_stat_card(title, value_ref, color):
        """Create a summary stat card."""
        return ft.Container(
            content=ft.Column(
                controls=[
                    ft.Text(title, size=13, color=ft.Colors.GREY_400),
                    value_ref,
                ],
                spacing=4,
                horizontal_alignment=ft.CrossAxisAlignment.CENTER,
            ),
            bgcolor="#2a2a2a",
            border_radius=10,
            padding=ft.padding.symmetric(horizontal=20, vertical=12),
            border=ft.border.all(1, color),
            expand=True,
        )
    
    stat_total_members = ft.Text("0", size=24, weight="bold", color=ft.Colors.BLUE_200)
    stat_active_members = ft.Text("0", size=24, weight="bold", color=ft.Colors.GREEN_400)
    stat_total_contributions = ft.Text("₦0.00", size=24, weight="bold", color=ft.Colors.AMBER_400)
    stat_active_loans = ft.Text("0", size=24, weight="bold", color=ft.Colors.ORANGE_400)
    
    stats_row = ft.Row(
        controls=[
            _make_stat_card("Total Members", stat_total_members, ft.Colors.BLUE_900),
            _make_stat_card("Active Members", stat_active_members, ft.Colors.GREEN_900),
            _make_stat_card("Total Contributions", stat_total_contributions, ft.Colors.AMBER_900),
            _make_stat_card("Active Loans", stat_active_loans, ft.Colors.ORANGE_900),
        ],
        spacing=10,
    )
    
    # --- Search Field ---
    search_field = ft.TextField(
        hint_text="Search by name, IPPIS, contact, email...",
        prefix_icon=ft.Icons.SEARCH,
        suffix=ft.IconButton(ft.Icons.CLOSE, on_click=on_search_clear, icon_size=18),
        on_change=on_search_change,
        expand=True,
        height=42,
        text_size=16,
        border_radius=8,
    )
    
    # --- Filter Segments ---
    seg_all = ft.Segment(value="All", label=ft.Text("All (0)"))
    seg_active = ft.Segment(value="Active", label=ft.Text("Active (0)"))
    seg_inactive = ft.Segment(value="Inactive", label=ft.Text("Inactive (0)"))
    seg_suspended = ft.Segment(value="Suspended", label=ft.Text("Suspended (0)"))
    
    filter_segmented_btn = ft.SegmentedButton(
        segments=[seg_all, seg_active, seg_inactive, seg_suspended],
        selected=["All"],
        on_change=on_filter_tab_change,
        allow_multiple_selection=False,
        show_selected_icon=False,
    )
    
    # --- Member Count Indicator ---
    member_count_text = ft.Text("Showing 0 of 0 members", size=14, color=ft.Colors.GREY_400, italic=True)
    
    # --- Sortable Column Definitions ---
    column_labels = {
        "name": "Name",
        "ippis": "IPPIS",
        "contact": "Contact",
        "email": "Email",
        "status": "Status",
        "join_date": "Join Date",
        "contributions": "Total Contributions",
        "active_loans": "Active Loans",
    }
    
    col_name = ft.DataColumn(
        ft.Text("Name ▲", color=ft.Colors.BLUE_200, weight="bold"),
        on_sort=lambda e: on_column_sort("name"),
    )
    col_ippis = ft.DataColumn(
        ft.Text("IPPIS", color=ft.Colors.WHITE, weight="bold"),
        on_sort=lambda e: on_column_sort("ippis"),
    )
    col_contact = ft.DataColumn(
        ft.Text("Contact", color=ft.Colors.WHITE, weight="bold"),
        on_sort=lambda e: on_column_sort("contact"),
    )
    col_email = ft.DataColumn(
        ft.Text("Email", color=ft.Colors.WHITE, weight="bold"),
        on_sort=lambda e: on_column_sort("email"),
    )
    col_status = ft.DataColumn(
        ft.Text("Status", color=ft.Colors.WHITE, weight="bold"),
        on_sort=lambda e: on_column_sort("status"),
    )
    col_join_date = ft.DataColumn(
        ft.Text("Join Date", color=ft.Colors.WHITE, weight="bold"),
        on_sort=lambda e: on_column_sort("join_date"),
    )
    col_contributions = ft.DataColumn(
        ft.Text("Total Contributions", color=ft.Colors.WHITE, weight="bold"),
        on_sort=lambda e: on_column_sort("contributions"),
        numeric=True,
    )
    col_active_loans = ft.DataColumn(
        ft.Text("Active Loans", color=ft.Colors.WHITE, weight="bold"),
        on_sort=lambda e: on_column_sort("active_loans"),
        numeric=True,
    )
    col_actions = ft.DataColumn(
        ft.Text("Actions", color=ft.Colors.WHITE, weight="bold"),
    )
    
    sortable_columns = {
        "name": col_name,
        "ippis": col_ippis,
        "contact": col_contact,
        "email": col_email,
        "status": col_status,
        "join_date": col_join_date,
        "contributions": col_contributions,
        "active_loans": col_active_loans,
    }
    
    # Members DataTable
    members_table = ft.DataTable(
        columns=[
            col_name, col_ippis, col_contact, col_email,
            col_status, col_join_date, col_contributions, col_active_loans, col_actions,
        ],
        rows=[],
        bgcolor="#1a1a1a",
        divider_thickness=1,
        sort_column_index=0,
        sort_ascending=True,
    )
    
    # --- Pagination Controls ---
    btn_prev = ft.IconButton(
        ft.Icons.CHEVRON_LEFT,
        tooltip="Previous Page",
        on_click=lambda e: on_page_change(-1),
        disabled=True,
    )
    btn_next = ft.IconButton(
        ft.Icons.CHEVRON_RIGHT,
        tooltip="Next Page",
        on_click=lambda e: on_page_change(1),
    )
    per_page_dropdown = ft.Dropdown(
        value="20",
        options=[
            ft.dropdown.Option("10", "10 per page"),
            ft.dropdown.Option("20", "20 per page"),
            ft.dropdown.Option("50", "50 per page"),
            ft.dropdown.Option("100", "100 per page"),
        ],
        on_select=on_per_page_change,
        width=130,
        height=36,
        text_size=14,
    )
    
    # --- Export Buttons ---
    export_csv_btn = ft.OutlinedButton(
        "CSV",
        icon=ft.Icons.DOWNLOAD,
        on_click=export_members_csv,
        tooltip="Export filtered members to CSV",
    )
    export_excel_btn = ft.OutlinedButton(
        "Excel",
        icon=ft.Icons.TABLE_CHART,
        on_click=export_members_excel,
        tooltip="Export filtered members to Excel",
    )
    
    # Initial render
    filter_members_table("")
    
    # Table wrapper with scrolling
    table_wrapper = ft.Container(
        content=ft.Column(
            controls=[members_table],
            scroll=ft.ScrollMode.AUTO,
        ),
        expand=True,
        bgcolor="#2a2a2a",
        border_radius=10,
        padding=10,
    )
    
    # Add button
    def open_add_member_dialog():
        edit_member_mode["enabled"] = False
        edit_member_mode["member_id"] = None
        member_dialog.title.value = "Add New Member"
        member_dialog.content.controls[0].value = ""
        member_dialog.content.controls[1].value = ""
        member_dialog.content.controls[2].value = ""
        member_dialog.content.controls[3].value = ""
        member_dialog.content.controls[4].value = "ACTIVE"
        member_dialog.open = True
        page.update()
    
    add_member_button = ft.ElevatedButton(
        "Add Member",
        icon=ft.Icons.ADD,
        on_click=lambda e: open_add_member_dialog(),
    )
    
    # Refresh button
    refresh_button = ft.IconButton(
        ft.Icons.REFRESH,
        tooltip="Refresh",
        on_click=lambda e: refresh_members(),
    )
    
    # Top controls row: search + buttons + export
    top_row = ft.Row(
        controls=[
            search_field,
            add_member_button,
            refresh_button,
            export_csv_btn,
            export_excel_btn,
        ],
        spacing=10,
    )
    
    # Bottom row: pagination + member count
    bottom_row = ft.Row(
        controls=[
            member_count_text,
            ft.Container(expand=True),  # spacer
            per_page_dropdown,
            btn_prev,
            btn_next,
        ],
        spacing=5,
        alignment=ft.MainAxisAlignment.SPACE_BETWEEN,
    )
    
    # Main content
    content = ft.Container(
        content=ft.Column(
            controls=[
                ft.Text("Member Management", size=26, weight="bold", color=ft.Colors.BLUE_200),
                stats_row,
                top_row,
                filter_segmented_btn,
                table_wrapper,
                bottom_row,
            ],
            spacing=15,
            expand=True,
        ),
        padding=20,
        bgcolor="#1a1a1a",
        expand=True,
    )
    
    # Create sidebar overlay
    sidebar_wrapper, backdrop, sidebar_visible, toggle_sidebar, close_sidebar = create_sidebar_overlay(page)
    
    # Create burger menu button
    burger_button = create_burger_menu(toggle_sidebar)
    
    # Create app bar with burger menu
    app_bar = ft.AppBar(
        title=ft.Text("Member Management", size=22, weight="bold", color=ft.Colors.WHITE),
        bgcolor=ft.Colors.BLUE_900,
        leading=burger_button,
        actions=[
            ft.IconButton(
                ft.Icons.LOGOUT,
                tooltip="Logout",
                on_click=lambda _: page.go("/login"),
            )
        ],
    )
    
    page.overlay.append(member_dialog)
    page.overlay.append(details_dialog)
    
    return ft.View(
        route="/members",
        controls=[
            app_bar,
            ft.Stack(
                controls=[
                    content,
                    backdrop,
                    sidebar_wrapper,
                ],
                expand=True,
            ),
        ],
    )

