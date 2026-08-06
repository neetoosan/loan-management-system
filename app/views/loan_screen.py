import flet as ft
from database.connection import (
    get_all_loans,
    get_all_members,
    get_member_by_id,
    create_loan,
    update_loan,
    record_repayment,
    create_refund,
    get_loans_by_member,
    calculate_interest,
    can_member_take_loan,
    calculate_loan_details,
    create_loan_with_validation,
    process_repayment_advanced,
    get_loan_summary,
    get_member_loan_status,
    get_loan_total_due_amount,
    get_loan_balance_amount,
    delete_loan as db_delete_loan,
    get_refunds_by_loan,
    process_refund,
    check_and_flag_overdue_borrowers,
)
from database.models import LoanStatus
from components.navigation import create_app_bar
from components.loan_details_dialog import create_loan_details_dialog
from components.burger_menu import create_sidebar_overlay, create_burger_menu
from components.top_up_loan import create_top_up_loan_dialog
from components.responsive import ResponsiveConfig, get_responsive_padding, get_responsive_font_size
from components.error_handler import error_logger, UserFriendlyError
from components.ui_components import ToastNotification, NotificationType
from datetime import datetime, timedelta
import csv
import os


def navigate_to(page: ft.Page, route: str):
    """Navigate to a different route (replaces deprecated page.go)"""
    page.route = route
    page.update()


def LoanScreen(page: ft.Page):
    """Loans management screen with responsive DataTable and dialogs"""
    
    # State management
    loans_list = get_all_loans()
    members_list = get_all_members()
    members_dict = {m.id: m.name for m in members_list}
    members_ippis = {m.id: m.ippis_number for m in members_list}
    
    # Preload non-members for performance (avoid database queries in loop)
    from database.connection import get_all_non_members
    non_members_list = get_all_non_members()
    non_members_dict = {nm.id: nm.name for nm in non_members_list}
    non_members_ippis = {nm.id: nm.ippis_number for nm in non_members_list}
    
    # Auto-flag borrowers with 90+ day overdue loans
    check_and_flag_overdue_borrowers()
    # Reload to get updated flags
    members_list = get_all_members()
    non_members_list = get_all_non_members()
    # Flag lookup dicts
    members_flagged = {m.id: m.is_flagged for m in members_list}
    non_members_flagged = {nm.id: nm.is_flagged for nm in non_members_list}
    
    # Pagination state
    pagination = {
        "current_page": 1,
        "items_per_page": 50,
        "total_items": len(loans_list),
        "total_pages": max(1, (len(loans_list) + 49) // 50),
        "all_filtered_loans": loans_list.copy(),  # Store filtered results
    }
    
    # Status filter state
    active_status_filter = {"value": "All"}
    
    # Sort state: default sort by updated_at descending
    sort_state = {"column": "updated_at", "ascending": False}
    
    # Loan dialog state
    loan_type = {"value": "member"}  # "member" or "non-member"
    selected_member_id = {"value": None}
    current_loan_for_repayment = {"value": None}  # Store the loan being repaid
    
    # ==================== SEARCHABLE MEMBER PICKER ====================
    # Build member data for search
    def _build_member_data():
        data = []
        for m in members_list:
            label = m.name
            if m.ippis_number:
                label += f"  (IPPIS: {m.ippis_number})"
            data.append({"id": m.id, "name": m.name, "ippis": m.ippis_number or "", "label": label, "flagged": m.is_flagged})
        return data
    member_data = _build_member_data()

    # ListView to display matching members
    member_results_list = ft.ListView(height=150, spacing=2, padding=5)

    # Chip showing selected member
    _selected_chip_text = ft.Text("", size=15, color=ft.Colors.GREEN_400, weight="bold")

    def clear_member_selection():
        selected_member_id["value"] = None
        member_name_field.value = ""
        ippis_field.value = ""
        selected_member_chip.visible = False
        member_search_field.value = ""
        member_results_container.visible = True
        _update_member_results("")
        page.update()

    selected_member_chip = ft.Container(
        content=ft.Row(
            controls=[
                ft.Icon(ft.Icons.PERSON, color=ft.Colors.GREEN_400, size=18),
                _selected_chip_text,
                ft.IconButton(
                    ft.Icons.CLOSE, icon_size=16, icon_color=ft.Colors.RED_400,
                    tooltip="Clear selection",
                    on_click=lambda e: clear_member_selection(),
                ),
            ],
            spacing=5,
            vertical_alignment=ft.CrossAxisAlignment.CENTER,
        ),
        bgcolor="#1a3a1a",
        border_radius=6,
        padding=ft.padding.symmetric(horizontal=10, vertical=4),
        visible=False,
    )

    def _on_member_search_change(e):
        term = (e.control.value or "").strip().lower()
        _update_member_results(term)
        member_results_container.visible = True
        page.update()

    def _update_member_results(term):
        member_results_list.controls.clear()
        if not term:
            matches = member_data[:20]
        else:
            matches = [
                md for md in member_data
                if term in md["name"].lower() or term in md["ippis"].lower()
            ][:20]

        if not matches:
            member_results_list.controls.append(
                ft.Container(
                    content=ft.Text("No members found", size=14, color=ft.Colors.GREY_400, italic=True),
                    padding=10,
                )
            )
        else:
            for md in matches:
                row_controls = []
                if md.get("flagged"):
                    row_controls.append(ft.Icon(ft.Icons.FLAG, color=ft.Colors.RED_400, size=16, tooltip="Flagged: Overdue 90+ days"))
                row_controls.append(ft.Text(md["label"], size=15, color=ft.Colors.WHITE))
                member_results_list.controls.append(
                    ft.Container(
                        content=ft.Row(row_controls, spacing=5, vertical_alignment=ft.CrossAxisAlignment.CENTER),
                        padding=ft.padding.symmetric(horizontal=10, vertical=8),
                        border_radius=4,
                        bgcolor="#2a2a2a",
                        on_click=lambda e, m=md: _select_member_from_list(m),
                        ink=True,
                    )
                )

    def _select_member_from_list(md):
        selected_member_id["value"] = md["id"]
        member = get_member_by_id(md["id"])
        if member:
            member_name_field.value = member.name
            ippis_field.value = member.ippis_number or "N/A"
        _selected_chip_text.value = md["label"]
        selected_member_chip.visible = True
        member_search_field.value = ""
        member_results_container.visible = False
        page.update()

    member_search_field = ft.TextField(
        hint_text="Type name or IPPIS to search...",
        prefix_icon=ft.Icons.SEARCH,
        on_change=_on_member_search_change,
        width=450,
        height=42,
        text_size=16,
        border_radius=8,
        dense=True,
    )

    member_results_container = ft.Container(
        content=member_results_list,
        bgcolor="#1e1e1e",
        border=ft.border.all(1, ft.Colors.GREY_800),
        border_radius=6,
        width=450,
        visible=True,
    )

    # Initialize with first 20 members
    _update_member_results("")

    # Full picker widget to place in dialog
    member_search_container = ft.Column(
        controls=[
            selected_member_chip,
            member_search_field,
            member_results_container,
        ],
        spacing=5,
    )

    # ==================== FORM FIELDS FOR DIALOG ====================
    
    member_name_field = ft.TextField(
        label="Member Name",
        read_only=True,
        width=450,
        dense=True,
    )
    
    ippis_field = ft.TextField(
        label="IPPIS Number",
        read_only=True,
        width=450,
        dense=True,
    )
    
    # Non-member fields (manual input)
    nonmember_name_field = ft.TextField(
        label="Full Name",
        width=450,
        visible=False,
        dense=True,
    )
    
    nonmember_ippis_field = ft.TextField(
        label="IPPIS Number (if applicable)",
        width=450,
        visible=False,
        dense=True,
    )
    
    # Common fields
    loan_amount_field = ft.TextField(
        label="Loan Amount (₦)",
        keyboard_type="number",
        width=450,
        dense=True,
        on_change=lambda e: on_amount_changed(e),
    )
    
    batch_number_field = ft.TextField(
        label="Batch Number",
        width=450,
        dense=True,
    )
    
    cheque_number_field = ft.TextField(
        label="Cheque Number",
        width=450,
        dense=True,
    )
    
    # Guarantor fields (for both members and non-members)
    guarantor_name_field = ft.TextField(
        label="Guarantor's Name",
        width=450,
        dense=True,
    )
    
    guarantor_phone_field = ft.TextField(
        label="Guarantor's Phone Number",
        keyboard_type="phone",
        width=450,
        dense=True,
    )
    
    loan_duration_field = ft.TextField(
        label="Loan Duration (months)",
        keyboard_type="number",
        value="12",
        width=450,
        dense=True,
        on_change=lambda e: on_amount_changed(e),
    )
    
    # Loan issuance date
    loan_date_field = ft.TextField(
        label="Loan Issue Date (YYYY-MM-DD)",
        width=450,
        dense=True,
        value=datetime.now().strftime("%Y-%m-%d"),
        hint_text="Date the loan was issued",
    )
    
    # Interest rate input (manual entry with 50% cap)
    interest_rate_field = ft.TextField(
        label="Interest Rate (%)",
        keyboard_type="number",
        value="10",
        width=450,
        dense=True,
        on_change=lambda e: on_interest_changed(e),
        hint_text="Enter rate (max 50%)",
    )
    
    # Interest amount display (read-only)
    interest_amount_field = ft.TextField(
        label="Total Interest Amount (₦)",
        read_only=True,
        width=450,
        dense=True,
    )
    
    # ==================== REPAYMENT FORM FIELDS ====================
    repayment_loan_info = ft.Text("", size=14)
    repayment_amount_field = ft.TextField(
        label="Amount to Pay (₦)",
        keyboard_type="number",
        width=450,
        dense=True,
    )
    
    repayment_date_field = ft.TextField(
        label="Payment Date (YYYY-MM-DD)",
        width=450,
        dense=True,
        value=datetime.now().strftime("%Y-%m-%d"),
    )
    
    repayment_notes_field = ft.TextField(
        label="Notes (optional)",
        multiline=True,
        width=450,
        dense=True,
        min_lines=3,
    )
    
    # Dialog for creating/editing loans (responsive)
    loan_dialog = ft.AlertDialog(
        title=ft.Text("Add New Loan"),
        content=ft.Container(
            content=ft.Column(
                controls=[
                    # Loan Type Selection
                    ft.Dropdown(
                        label="Loan Type",
                        value="member",
                        options=[
                            ft.dropdown.Option("member", "Member"),
                            ft.dropdown.Option("non-member", "Non-Member"),
                        ],
                        width=450,
                        dense=True,
                        on_select=lambda e: on_loan_type_change(e),
                    ),
                    ft.Divider(height=10),
                    
                    # Member section
                    ft.Text("Type member name or IPPIS to search:", size=14, color=ft.Colors.GREY_400),
                    member_search_container,
                    member_name_field,
                    ippis_field,
                    
                    # Non-member section
                    nonmember_name_field,
                    nonmember_ippis_field,
                    
                    ft.Divider(height=10),
                    
                    # Loan amount
                    loan_amount_field,
                    
                    # Batch & Cheque numbers
                    batch_number_field,
                    cheque_number_field,
                    
                    # Guarantor details
                    ft.Text("Guarantor Information:", weight="bold", size=14),
                    guarantor_name_field,
                    guarantor_phone_field,
                    
                    # Duration
                    loan_duration_field,
                    
                    # Loan issuance date
                    loan_date_field,
                    
                    # Interest rate
                    ft.Text("Interest Rate:", weight="bold", size=14),
                    interest_rate_field,
                    
                    # Interest amount display
                    interest_amount_field,
                ],
                spacing=8,
                scroll=ft.ScrollMode.AUTO,
                tight=True,
            ),
            padding=15,
        ),
        actions=[
            ft.TextButton("Cancel", on_click=lambda e: close_loan_dialog()),
            ft.TextButton("Create Loan", on_click=lambda e: create_new_loan()),
        ],
    )
    
    # Dialog for recording repayment with date input
    repayment_dialog = ft.AlertDialog(
        title=ft.Text("Record Repayment"),
        content=ft.Column(
            controls=[
                repayment_loan_info,
                ft.Divider(height=10),
                repayment_amount_field,
                repayment_date_field,
                repayment_notes_field,
            ],
            width=450,
            spacing=8,
            scroll=ft.ScrollMode.AUTO,
            tight=True,
        ),
        actions=[
            ft.TextButton("Cancel", on_click=lambda e: close_repayment_dialog()),
            ft.TextButton("Confirm", on_click=lambda e: confirm_repayment()),
        ],
    )
    
    # ==================== LOAN DIALOG HANDLERS ====================
    
    def on_loan_type_change(e):
        """Handle loan type change - show/hide fields based on member type"""
        loan_type["value"] = e.control.value
        is_member = loan_type["value"] == "member"
        
        # Show/hide member fields
        member_search_container.visible = is_member
        member_name_field.visible = is_member
        ippis_field.visible = is_member
        
        # Show/hide non-member fields
        nonmember_name_field.visible = not is_member
        nonmember_ippis_field.visible = not is_member
        
        # Clear fields
        member_search_field.value = ""
        member_name_field.value = ""
        ippis_field.value = ""
        nonmember_name_field.value = ""
        nonmember_ippis_field.value = ""
        selected_member_chip.visible = False
        member_results_container.visible = True
        _update_member_results("")
        
        selected_member_id["value"] = None
        page.update()
    
    def on_amount_changed(e):
        """Handle loan amount or duration change and recalculate interest"""
        calculate_total_interest()
    
    def on_interest_changed(e):
        """Handle interest rate input change - validate cap and recalculate"""
        try:
            rate_str = interest_rate_field.value or "0"
            rate = float(rate_str)
            
            # Enforce 50% cap for both members and non-members
            if rate > 50:
                interest_rate_field.value = "50"
                ToastNotification.show(page, "✗ Interest rate capped at 50%", NotificationType.ERROR)
                page.update()
                return
            
            calculate_total_interest()
        except ValueError:
            # Invalid input, just recalculate with what we have
            calculate_total_interest()
    
    def calculate_total_interest():
        """Calculate and update total interest based on amount, duration, and interest rate
        
        For MEMBERS: Flat-rate interest = (amount × rate) / 100 (NOT multiplied by duration)
        For NON-MEMBERS: Monthly interest = (amount × rate × months) / 100
        """
        try:
            amount_str = loan_amount_field.value or "0"
            duration_str = loan_duration_field.value or "1"
            rate_str = interest_rate_field.value or "0"
            
            amount = float(amount_str)
            duration = int(duration_str) if duration_str else 1
            interest_rate = float(rate_str)
            
            # Validate interest rate cap (50% max)
            if interest_rate > 50:
                interest_rate = 50
                interest_rate_field.value = "50"
            
            is_member = loan_type["value"] == "member"
            
            if is_member:
                # Members: Flat-rate simple interest (NOT multiplied by duration)
                # Example: ₦50,000 at 5% = ₦2,500 total interest (fixed)
                total_interest = (amount * interest_rate) / 100
            else:
                # Non-members: Monthly compound interest (applied per month)
                # Example: ₦50,000 at 2% per month for 12 months = ₦12,000 interest
                total_interest = (amount * interest_rate * duration) / 100
            
            # Display interest amount
            interest_amount_field.value = f"₦{total_interest:.2f}"
            page.update()
        except Exception as ex:
            print(f"Error calculating interest: {ex}")
            interest_amount_field.value = "₦0.00"
            page.update()
    
    def close_loan_dialog():
        """Close loan dialog and reset all fields"""
        loan_dialog.open = False
        
        # Reset to member type
        loan_type["value"] = "member"
        member_search_container.visible = True
        nonmember_name_field.visible = False
        nonmember_ippis_field.visible = False
        
        # Clear all fields
        member_search_field.value = ""
        selected_member_chip.visible = False
        member_results_container.visible = True
        _update_member_results("")
        member_name_field.value = ""
        ippis_field.value = ""
        nonmember_name_field.value = ""
        nonmember_ippis_field.value = ""
        loan_amount_field.value = ""
        batch_number_field.value = ""
        cheque_number_field.value = ""
        loan_duration_field.value = "12"
        loan_date_field.value = datetime.now().strftime("%Y-%m-%d")
        interest_rate_field.value = "10"
        interest_amount_field.value = ""
        selected_member_id["value"] = None
        page.update()
    
    def close_repayment_dialog():
        """Close repayment dialog"""
        repayment_dialog.open = False
        page.update()
    
    def create_new_loan():
        """Create a new loan with member or non-member details"""
        try:
            from database.connection import create_non_member
            
            amount_str = loan_amount_field.value or "0"
            duration_str = loan_duration_field.value or "1"
            interest_rate_str = interest_rate_field.value or "0"
            batch_number = batch_number_field.value
            cheque_number = cheque_number_field.value
            guarantor_name = guarantor_name_field.value
            guarantor_phone = guarantor_phone_field.value
            loan_date_str = loan_date_field.value or datetime.now().strftime("%Y-%m-%d")
            
            # INPUT VALIDATION
            # Validate basic inputs
            try:
                amount = float(amount_str)
                duration = int(duration_str) if duration_str else 1
                interest_rate = float(interest_rate_str)
            except ValueError as e:
                error_logger.warning(f"Invalid loan input: amount={amount_str}, duration={duration_str}, rate={interest_rate_str}")
                error_msg = UserFriendlyError.get_message('validation_error', "Invalid amount, duration, or interest rate")
                ToastNotification.show(page, f"✗ {error_msg}", NotificationType.ERROR)
                page.update()
                return
            
            # Validate amount and duration
            if amount <= 0:
                error_logger.warning(f"Loan amount validation failed: {amount} <= 0")
                error_msg = UserFriendlyError.get_message('validation_error', "Loan amount must be greater than 0")
                ToastNotification.show(page, f"✗ {error_msg}", NotificationType.ERROR)
                page.update()
                return
            
            if duration <= 0:
                error_logger.warning(f"Loan duration validation failed: {duration} <= 0")
                error_msg = UserFriendlyError.get_message('validation_error', "Duration must be at least 1 month")
                ToastNotification.show(page, f"✗ {error_msg}", NotificationType.ERROR)
                page.update()
                return
            
            if interest_rate < 0 or interest_rate > 50:
                error_logger.warning(f"Interest rate validation failed: {interest_rate} not in 0-50%")
                error_msg = UserFriendlyError.get_message('validation_error', "Interest rate must be between 0% and 50%")
                ToastNotification.show(page, f"✗ {error_msg}", NotificationType.ERROR)
                page.update()
                return
            
            if not guarantor_name or not guarantor_name.strip():
                error_logger.warning("Guarantor name validation failed: empty")
                error_msg = UserFriendlyError.get_message('missing_required_field', "Guarantor name is required")
                ToastNotification.show(page, f"✗ {error_msg}", NotificationType.ERROR)
                page.update()
                return
            
            if not guarantor_phone or not guarantor_phone.strip():
                error_logger.warning("Guarantor phone validation failed: empty")
                error_msg = UserFriendlyError.get_message('missing_required_field', "Guarantor phone is required")
                ToastNotification.show(page, f"✗ {error_msg}", NotificationType.ERROR)
                page.update()
                return
            
            # Parse loan issue date
            try:
                start_date = datetime.strptime(loan_date_str, "%Y-%m-%d")
            except Exception as e:
                error_logger.warning(f"Invalid loan date format: {loan_date_str}")
                error_logger.info("Using current date for loan")
                start_date = datetime.now()
            
            is_member = loan_type["value"] == "member"
            member_id = None
            non_member_id = None
            
            if is_member:
                # Member loan - must have member selected
                if not selected_member_id["value"]:
                    error_logger.warning("Member loan creation: no member selected")
                    error_msg = UserFriendlyError.get_message('missing_required_field', "Please select a member")
                    ToastNotification.show(page, f"✗ {error_msg}", NotificationType.ERROR)
                    page.update()
                    return
                
                member_id = selected_member_id["value"]
                error_logger.info(f"Creating member loan: member_id={member_id}, amount={amount}, duration={duration}m")
                
                # CHECK BUSINESS RULES: Can member take loan?
                can_take, reason = can_member_take_loan(member_id)
                if not can_take:
                    error_logger.warning(f"Member {member_id} cannot take loan: {reason}")
                    ToastNotification.show(page, f"✗ {reason}", NotificationType.ERROR)
                    page.update()
                    return
                
                # Use advanced loan creation with validation
                success, result, info = create_loan_with_validation(
                    member_id=member_id,
                    amount=amount,
                    interest_rate=interest_rate,
                    duration_months=duration,
                    start_date=start_date,
                    batch_number=batch_number or None,
                    cheque_number=cheque_number or None,
                    guarantor_name=guarantor_name,
                    guarantor_phone=guarantor_phone,
                    is_member=True  # Use flat-rate interest for members
                )
                
                if not success:
                    error_logger.error(f"Member loan creation failed: {result}")
                    ToastNotification.show(page, f"✗ {result}", NotificationType.ERROR)
                    page.update()
                    return
                
                error_logger.info(f"Member loan created successfully: loan_id={info['loan_id']}")
                # Show success with loan details
                close_loan_dialog()
                refresh_loans()
                msg = f"✓ Loan created!\nLoan ID: {info['loan_id']}\nTotal Due: ₦{info['total_due']:.2f}\nMonthly Payment: ₦{info['monthly_payment']:.2f}"
                ToastNotification.show(page, msg, NotificationType.INFO)
                page.update()
            
            else:
                # Non-member loan - create non-member record
                nonmember_name = nonmember_name_field.value
                nonmember_ippis = nonmember_ippis_field.value
                
                if not nonmember_name or not nonmember_name.strip():
                    error_logger.warning("Non-member loan creation: no name provided")
                    error_msg = UserFriendlyError.get_message('missing_required_field', "Please enter non-member full name")
                    ToastNotification.show(page, f"✗ {error_msg}", NotificationType.ERROR)
                    page.update()
                    return
                
                error_logger.info(f"Creating non-member loan: name={nonmember_name}, amount={amount}")
                
                # Create non-member record
                from database.connection import create_non_member
                try:
                    non_member = create_non_member(
                        name=nonmember_name.strip(),
                        contact=guarantor_phone.strip(),
                        email=None,
                        ippis_number=nonmember_ippis.strip() if nonmember_ippis else None
                    )
                except Exception as e:
                    error_logger.exception(f"Failed to create non-member record: {str(e)}")
                    error_msg = UserFriendlyError.get_message('db_commit_error', "Failed to create non-member record")
                    ToastNotification.show(page, f"✗ {error_msg}", NotificationType.ERROR)
                    page.update()
                    return
                
                if not non_member:
                    error_logger.error("Non-member creation returned None")
                    error_msg = UserFriendlyError.get_message('db_commit_error', "Error creating non-member record")
                    ToastNotification.show(page, f"✗ {error_msg}", NotificationType.ERROR)
                    page.update()
                    return
                
                # Calculate loan details for non-member (monthly compound interest)
                try:
                    loan_details = calculate_loan_details(amount, interest_rate, duration, is_member=False)
                except ValueError as e:
                    error_logger.warning(f"Loan details calculation failed: {str(e)}")
                    ToastNotification.show(page, f"✗ {str(e)}", NotificationType.ERROR)
                    return
                
                end_date = start_date + timedelta(days=30 * duration)
                
                # Create loan for non-member
                try:
                    new_loan = create_loan(
                        member_id=None,
                        amount=amount,
                        interest_rate=interest_rate,
                        start_date=start_date,
                        end_date=end_date,
                        is_member=False,  # Use monthly compound interest for non-members
                        batch_number=batch_number or None,
                        cheque_number=cheque_number or None,
                        non_member_id=non_member.id,
                        guarantor_name=guarantor_name,
                        guarantor_phone=guarantor_phone,
                        duration_months=duration
                    )
                except Exception as e:
                    error_logger.exception(f"Failed to create non-member loan: {str(e)}")
                    error_msg = UserFriendlyError.get_message('db_commit_error', str(e))
                    ToastNotification.show(page, f"✗ {error_msg}", NotificationType.ERROR)
                    page.update()
                    return
                
                if new_loan:
                    error_logger.info(f"Non-member loan created successfully: loan_id={new_loan.id}")
                    close_loan_dialog()
                    refresh_loans()
                    msg = f"✓ Non-member loan created!\nLoan ID: {new_loan.id}\nTotal Due: ₦{loan_details['total_due']:.2f}"
                    ToastNotification.show(page, msg, NotificationType.INFO)
                    page.update()
                else:
                    error_logger.error("Non-member loan creation returned None")
                    error_msg = UserFriendlyError.get_message('db_commit_error', "Error creating loan")
                    ToastNotification.show(page, f"✗ {error_msg}", NotificationType.ERROR)
                    page.update()
        
        except Exception as ex:
            error_logger.exception(f"Unexpected error in create_new_loan: {str(ex)}")
            error_msg = UserFriendlyError.get_message('operation_failed', str(ex))
            ToastNotification.show(page, f"✗ {error_msg}", NotificationType.ERROR)
            page.update()
    
    def confirm_repayment():
        """Record a loan repayment with date and handle refunds"""
        try:
            loan = current_loan_for_repayment["value"]
            if not loan:
                error_logger.warning("Repayment processing: no loan selected")
                error_msg = UserFriendlyError.get_message('missing_required_field', "Please select a loan")
                ToastNotification.show(page, f"✗ {error_msg}", NotificationType.ERROR)
                page.update()
                return
            
            amount_str = repayment_amount_field.value or "0"
            payment_date_str = repayment_date_field.value
            notes = repayment_notes_field.value
            
            # Validate amount
            try:
                amount = float(amount_str)
            except ValueError as e:
                error_logger.warning(f"Invalid repayment amount: {amount_str}")
                error_msg = UserFriendlyError.get_message('validation_error', "Please enter a valid repayment amount")
                ToastNotification.show(page, f"✗ {error_msg}", NotificationType.ERROR)
                page.update()
                return
            
            if amount <= 0:
                error_logger.warning(f"Repayment amount validation failed: {amount} <= 0")
                error_msg = UserFriendlyError.get_message('validation_error', "Repayment amount must be greater than 0")
                ToastNotification.show(page, f"✗ {error_msg}", NotificationType.ERROR)
                page.update()
                return
            
            # Parse date
            try:
                payment_date = datetime.strptime(payment_date_str, "%Y-%m-%d")
            except Exception as e:
                error_logger.warning(f"Invalid payment date format: {payment_date_str}, using current date")
                payment_date = datetime.now()
            
            error_logger.info(f"Processing repayment: loan_id={loan.id}, amount={amount}, date={payment_date.strftime('%Y-%m-%d')}")
            
            # Use advanced repayment processing
            result = process_repayment_advanced(
                loan_id=loan.id,
                amount_paid=amount,
                payment_date=payment_date,
                notes=notes or None
            )
            
            if not result['success']:
                error_logger.error(f"Repayment processing failed for loan {loan.id}: {result['message']}")
                ToastNotification.show(page, f"✗ {result['message']}", NotificationType.ERROR)
                page.update()
                return
            
            error_logger.info(f"Repayment recorded successfully for loan {loan.id}: ₦{amount}")
            
            close_repayment_dialog()
            refresh_loans()
            
            # Build success message
            msg = f"✓ Payment Recorded: ₦{amount:.2f}\n"
            msg += f"Status: {result['loan_status']}\n"
            msg += f"Balance Remaining: ₦{result['balance_remaining']:.2f}"
            
            if result['loan_fully_paid']:
                msg += "\n✓ LOAN FULLY PAID!"
                error_logger.info(f"Loan {loan.id} fully paid!")
            
            if result['refund_created']:
                msg += f"\n💰 Refund Created: ₦{result['refund_amount']:.2f} (PENDING)"
                error_logger.info(f"Refund created for loan {loan.id}: ₦{result['refund_amount']:.2f}")
            
            ToastNotification.show(page, msg, NotificationType.INFO)
            page.update()
                
        except ValueError as e:
            error_logger.warning(f"Value error in repayment: {str(e)}")
            error_msg = UserFriendlyError.get_message('validation_error', str(e))
            ToastNotification.show(page, f"✗ {error_msg}", NotificationType.ERROR)
            page.update()
        except Exception as ex:
            error_logger.exception(f"Unexpected error recording repayment: {str(ex)}")
            error_msg = UserFriendlyError.get_message('operation_failed', str(ex))
            ToastNotification.show(page, f"✗ {error_msg}", NotificationType.ERROR)
            page.update()
    
    
    def refresh_loans():
        """Refresh loans list and reinitialize pagination"""
        nonlocal loans_list
        loans_list = get_all_loans()
        # Also refresh non-member cache
        try:
            from database.connection import get_all_non_members as reload_non_members
            refreshed_nm = reload_non_members()
            non_members_dict.clear()
            non_members_ippis.clear()
            for nm in refreshed_nm:
                non_members_dict[nm.id] = nm.name
                non_members_ippis[nm.id] = nm.ippis_number
        except Exception:
            pass
        # Reset pagination to the beginning
        pagination["current_page"] = 1
        pagination["total_items"] = len(loans_list)
        pagination["total_pages"] = max(1, (len(loans_list) + 49) // 50)
        # Clear search field
        search_field.value = ""
        update_summary_stats()
        update_filter_tab_counts()
        update_loans_table()
    
    def on_search_changed(e):
        """Search loans — runs directly on UI thread to avoid threading errors"""
        search_value = e.control.value
        try:
            filter_loans_table(search_value)
        except Exception as err:
            error_logger.error(f"Error in search: {err}")
    
    # Search field
    search_field = ft.TextField(
        label="Search by name or IPPIS",
        prefix_icon=ft.Icons.SEARCH,
        width=300,
        dense=True,
        on_change=on_search_changed,
        color=ft.Colors.WHITE,
        label_style=ft.TextStyle(color=ft.Colors.BLUE_200),
        border_color=ft.Colors.BLUE_400,
        focused_border_color=ft.Colors.CYAN_400,
        cursor_color=ft.Colors.WHITE,
        bgcolor="#2a2a2a",
        border_radius=8,
    )
    
    def _is_loan_overdue(loan):
        """Check if a loan is overdue (active and past end date)"""
        if loan.status == LoanStatus.ACTIVE and loan.end_date:
            return loan.end_date < datetime.now()
        return False
    
    def _get_sort_key(item, col, asc):
        """Get sort key for a filtered loan tuple (loan, name, ippis)"""
        loan, member_name, ippis_number = item
        total_amount = get_loan_total_due_amount(loan)
        balance = get_loan_balance_amount(loan)
        
        if col == "name":
            return member_name.lower()
        elif col == "amount":
            return loan.amount
        elif col == "total":
            return total_amount
        elif col == "balance":
            return balance
        elif col == "status":
            return loan.status.value
        elif col == "amount_paid":
            return loan.amount_repaid
        elif col == "interest":
            return loan.total_interest
        else:  # updated_at (default)
            return loan.updated_at or loan.created_at or datetime.min
    
    # ==================== EXPORT FUNCTIONS ====================
    def _compute_loan_status_text(loan):
        """Return display status for a loan (matches table logic)."""
        if _is_loan_overdue(loan):
            return "OVERDUE"
        # Only show Refund Due if there are pending refunds
        loan_refunds = get_refunds_by_loan(loan.id)
        if any(r.status == "PENDING" for r in loan_refunds):
            return "Refund Due"
        return loan.status.value

    def export_loans_csv(e):
        """Export the currently filtered loan table to CSV."""
        try:
            filtered = pagination.get("all_filtered_loans", [])
            if not filtered:
                ToastNotification.show(page, "No loans to export", NotificationType.WARNING)
                page.update()
                return

            filter_name = active_status_filter["value"]
            downloads_dir = os.path.join(os.path.expanduser("~"), "Downloads")
            os.makedirs(downloads_dir, exist_ok=True)
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"loans_{filter_name}_{timestamp}.csv"
            filepath = os.path.join(downloads_dir, filename)

            headers = ["S/N", "IPPIS No", "Name", "Amount", "Interest",
                       "Total", "Monthly Repayment", "Amount Paid", "Balance", "Status"]

            with open(filepath, "w", newline="", encoding="utf-8") as f:
                writer = csv.writer(f)
                writer.writerow(headers)
                for idx, (loan, member_name, ippis_number) in enumerate(filtered, 1):
                    total = get_loan_total_due_amount(loan)
                    balance = get_loan_balance_amount(loan)
                    duration = 12
                    if loan.end_date and loan.start_date:
                        duration = max(1, (loan.end_date - loan.start_date).days // 30)
                    monthly = total / duration if duration > 0 else 0
                    status_text = _compute_loan_status_text(loan)
                    writer.writerow([
                        idx, ippis_number or "N/A", member_name,
                        f"{loan.amount:.2f}", f"{loan.total_interest:.2f}",
                        f"{total:.2f}", f"{monthly:.2f}",
                        f"{loan.amount_repaid:.2f}", f"{balance:.2f}", status_text,
                    ])

            ToastNotification.show(page, f"✓ Exported {len(filtered)} {filter_name} loans to Downloads/{filename}", NotificationType.SUCCESS)
            error_logger.info(f"Exported {len(filtered)} loans to {filepath}")
        except Exception as ex:
            error_logger.exception(f"CSV export failed: {ex}")
            ToastNotification.show(page, f"Export failed: {ex}", NotificationType.ERROR)
            page.update()

    def export_loans_excel(e):
        """Export the currently filtered loan table to Excel."""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment

            filtered = pagination.get("all_filtered_loans", [])
            if not filtered:
                ToastNotification.show(page, "No loans to export", NotificationType.WARNING)
                page.update()
                return

            filter_name = active_status_filter["value"]
            downloads_dir = os.path.join(os.path.expanduser("~"), "Downloads")
            os.makedirs(downloads_dir, exist_ok=True)
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"loans_{filter_name}_{timestamp}.xlsx"
            filepath = os.path.join(downloads_dir, filename)

            wb = Workbook()
            ws = wb.active
            ws.title = f"Loans - {filter_name}"

            headers = ["S/N", "IPPIS No", "Name", "Amount", "Interest",
                       "Total", "Monthly Repayment", "Amount Paid", "Balance", "Status"]
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="1a237e", end_color="1a237e", fill_type="solid")

            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")

            for row_idx, (loan, member_name, ippis_number) in enumerate(filtered, 2):
                total = get_loan_total_due_amount(loan)
                balance = get_loan_balance_amount(loan)
                duration = 12
                if loan.end_date and loan.start_date:
                    duration = max(1, (loan.end_date - loan.start_date).days // 30)
                monthly = total / duration if duration > 0 else 0
                status_text = _compute_loan_status_text(loan)

                ws.cell(row=row_idx, column=1, value=row_idx - 1)
                ws.cell(row=row_idx, column=2, value=ippis_number or "N/A")
                ws.cell(row=row_idx, column=3, value=member_name)
                ws.cell(row=row_idx, column=4, value=round(loan.amount, 2))
                ws.cell(row=row_idx, column=5, value=round(loan.total_interest, 2))
                ws.cell(row=row_idx, column=6, value=round(total, 2))
                ws.cell(row=row_idx, column=7, value=round(monthly, 2))
                ws.cell(row=row_idx, column=8, value=round(loan.amount_repaid, 2))
                ws.cell(row=row_idx, column=9, value=round(balance, 2))
                ws.cell(row=row_idx, column=10, value=status_text)

            for col in ws.columns:
                max_len = 0
                col_letter = col[0].column_letter
                for cell in col:
                    try:
                        if cell.value:
                            max_len = max(max_len, len(str(cell.value)))
                    except:
                        pass
                ws.column_dimensions[col_letter].width = min(max_len + 3, 40)

            wb.save(filepath)
            ToastNotification.show(page, f"✓ Exported {len(filtered)} {filter_name} loans to Downloads/{filename}", NotificationType.SUCCESS)
            error_logger.info(f"Exported {len(filtered)} loans to {filepath}")
        except Exception as ex:
            error_logger.exception(f"Excel export failed: {ex}")
            ToastNotification.show(page, f"Export failed: {ex}", NotificationType.ERROR)
            page.update()

    def update_loans_table():
        """Update loans table with paginated loans"""
        filter_loans_table("")
    
    def filter_loans_table(search_term):
        """Filter loans table based on status filter, search term, smart sort, with pagination"""
        rows = []
        search_lower = search_term.lower() if search_term else ""
        filtered_loans = []
        now = datetime.now()
        
        # Reset to page 1 only if there's a new search
        if search_term:
            pagination["current_page"] = 1
        
        # Step 1: Apply status filter
        status_filter = active_status_filter["value"]
        
        for loan in loans_list:
            # Status filter
            if status_filter != "All":
                if status_filter == "Overdue":
                    if not _is_loan_overdue(loan):
                        continue
                elif status_filter == "Active":
                    # Active excludes overdue and paid loans
                    if loan.status != LoanStatus.ACTIVE or _is_loan_overdue(loan):
                        continue
                elif status_filter == "Refund":
                    # Refund: loans with pending refund records
                    loan_refunds = get_refunds_by_loan(loan.id)
                    if not any(r.status == "PENDING" for r in loan_refunds):
                        continue
                elif status_filter == "Paid":
                    # Paid: must be PAID status AND no pending refunds
                    if loan.status != LoanStatus.PAID:
                        continue
                    loan_refunds = get_refunds_by_loan(loan.id)
                    if any(r.status == "PENDING" for r in loan_refunds):
                        continue
                else:
                    if loan.status.value != status_filter:
                        continue
            
            # Get borrower name based on member type
            if loan.is_member and loan.member_id:
                member_name = members_dict.get(loan.member_id, "Unknown")
                ippis_number = members_ippis.get(loan.member_id, "N/A") or "N/A"
            elif not loan.is_member and loan.non_member_id:
                member_name = non_members_dict.get(loan.non_member_id, "Unknown")
                ippis_number = non_members_ippis.get(loan.non_member_id, "N/A") or "N/A"
            else:
                member_name = "Unknown"
                ippis_number = "N/A"
            
            # Step 2: Apply search filter
            if search_lower == "" or search_lower in member_name.lower() or search_lower in str(ippis_number).lower():
                filtered_loans.append((loan, member_name, ippis_number))
        
        # Step 3: Smart sort — paid loans to bottom, then by selected column
        col = sort_state["column"]
        asc = sort_state["ascending"]
        
        def composite_sort_key(item):
            loan = item[0]
            # Primary: paid loans go to bottom (1), everything else stays on top (0)
            paid_rank = 1 if loan.status == LoanStatus.PAID else 0
            # Secondary: the user-selected sort column
            sort_val = _get_sort_key(item, col, asc)
            return (paid_rank, sort_val)
        
        try:
            filtered_loans.sort(key=composite_sort_key, reverse=(not asc))
        except TypeError:
            # Fallback if mixed types in sort
            pass
        
        # For descending sorts, paid loans still need to be at the bottom.
        # Since we used (paid_rank, sort_val) with reverse, we need to re-separate:
        # Split into non-paid and paid, sort each, then concatenate
        non_paid = [item for item in filtered_loans if item[0].status != LoanStatus.PAID]
        paid = [item for item in filtered_loans if item[0].status == LoanStatus.PAID]
        
        try:
            non_paid.sort(key=lambda item: _get_sort_key(item, col, asc), reverse=(not asc))
            paid.sort(key=lambda item: _get_sort_key(item, col, asc), reverse=(not asc))
        except TypeError:
            pass
        
        filtered_loans = non_paid + paid
        
        # Update pagination info
        pagination["all_filtered_loans"] = filtered_loans
        pagination["total_items"] = len(filtered_loans)
        pagination["total_pages"] = max(1, (len(filtered_loans) + pagination["items_per_page"] - 1) // pagination["items_per_page"])
        
        # Ensure current page is valid
        pagination["current_page"] = max(1, min(pagination["current_page"], pagination["total_pages"]))
        
        # Calculate pagination range
        start_idx = (pagination["current_page"] - 1) * pagination["items_per_page"]
        end_idx = start_idx + pagination["items_per_page"]
        page_loans = filtered_loans[start_idx:end_idx]
        
        # Create rows for current page only
        for idx, (loan, member_name, ippis_number) in enumerate(page_loans, 1):
            total_amount = get_loan_total_due_amount(loan)
            balance = get_loan_balance_amount(loan)
            
            duration_months = 12
            if loan.end_date and loan.start_date:
                duration_months = max(1, (loan.end_date - loan.start_date).days // 30)
            
            monthly_repayment = total_amount / duration_months if duration_months > 0 else 0
            
            # Determine overdue status
            is_overdue = _is_loan_overdue(loan)
            
            # Determine refund status — only show Refund Due when there are PENDING refunds
            loan_refunds = get_refunds_by_loan(loan.id)
            pending_refunds = [r for r in loan_refunds if r.status == "PENDING"]
            pending_refund_total = sum(r.refund_amount for r in pending_refunds)
            has_refund_due = len(pending_refunds) > 0
            
            # Row background color based on status
            if has_refund_due:
                row_color = "#3d3a1f"  # Amber tint for refund due
            elif is_overdue:
                row_color = "#3d1f1f"  # Subtle red tint for overdue
            elif loan.status == LoanStatus.PAID and not has_refund_due:
                row_color = "#1f2d1f"  # Subtle green tint for paid
            else:
                row_color = None  # Default
            
            # Status display text and color — pending refund takes priority over paid
            if has_refund_due:
                status_text = f"Refund Due"
                status_color = ft.Colors.AMBER_400
            elif is_overdue:
                status_text = "OVERDUE"
                status_color = ft.Colors.RED_400
            elif loan.status == LoanStatus.ACTIVE:
                status_text = "Active"
                status_color = ft.Colors.GREEN_400
            elif loan.status == LoanStatus.PENDING:
                status_text = "Pending"
                status_color = ft.Colors.AMBER_400
            elif loan.status == LoanStatus.PAID:
                status_text = "Paid"
                status_color = ft.Colors.GREY_600
            elif loan.status == LoanStatus.DEFAULTED:
                status_text = "Defaulted"
                status_color = ft.Colors.RED_300
            else:
                status_text = loan.status.value
                status_color = ft.Colors.GREY
            
            # Balance color — red if overdue, orange otherwise
            balance_color = ft.Colors.RED_400 if is_overdue else ft.Colors.ORANGE_400
            
            # Check if borrower is flagged
            borrower_flagged = False
            if loan.is_member and loan.member_id:
                borrower_flagged = members_flagged.get(loan.member_id, False)
            elif not loan.is_member and loan.non_member_id:
                borrower_flagged = non_members_flagged.get(loan.non_member_id, False)
            
            # Name cell with optional flag icon
            if borrower_flagged:
                name_cell_content = ft.Row(
                    [ft.Icon(ft.Icons.FLAG, color=ft.Colors.RED_400, size=16, tooltip="Flagged: Overdue 90+ days"),
                     ft.Text(member_name, color=ft.Colors.WHITE)],
                    spacing=4, vertical_alignment=ft.CrossAxisAlignment.CENTER,
                )
            else:
                name_cell_content = ft.Text(member_name, color=ft.Colors.WHITE)
            
            rows.append(
                ft.DataRow(
                    color=row_color,
                    cells=[
                        ft.DataCell(ft.Text(str(start_idx + idx), color=ft.Colors.WHITE)),
                        ft.DataCell(ft.Text(str(ippis_number), color=ft.Colors.WHITE)),
                        ft.DataCell(name_cell_content),
                        ft.DataCell(ft.Text(f"₦{loan.amount:,.2f}", color=ft.Colors.GREEN_400, weight="bold")),
                        ft.DataCell(ft.Text(f"₦{loan.total_interest:,.2f}", color=ft.Colors.ORANGE_400)),
                        ft.DataCell(ft.Text(f"₦{total_amount:,.2f}", color=ft.Colors.BLUE_200, weight="bold")),
                        ft.DataCell(ft.Text(f"₦{monthly_repayment:,.2f}", color=ft.Colors.GREY)),
                        ft.DataCell(ft.Text(f"₦{loan.amount_repaid:,.2f}", color=ft.Colors.BLUE_200)),
                        ft.DataCell(ft.Text(f"₦{balance:,.2f}", color=balance_color, weight="bold")),
                        ft.DataCell(
                            ft.Container(
                                content=ft.Text(status_text, size=13, weight="bold", color=status_color),
                                bgcolor="#1a1a1a" if not is_overdue else "#4d1a1a",
                                border_radius=4,
                                padding=ft.padding.symmetric(horizontal=8, vertical=2),
                            )
                        ),
                        ft.DataCell(
                            ft.Row(
                                controls=[
                                    ft.IconButton(
                                        ft.Icons.VISIBILITY,
                                        tooltip="View Details",
                                        on_click=lambda e, l=loan: open_loan_details(l),
                                        icon_size=20,
                                        style=ft.ButtonStyle(padding=ft.padding.all(4)),
                                    ),
                                    ft.IconButton(
                                        ft.Icons.ATTACH_MONEY,
                                        tooltip="Record Payment",
                                        on_click=lambda e, l=loan: open_repayment_dialog(l),
                                        icon_size=20,
                                        style=ft.ButtonStyle(padding=ft.padding.all(4)),
                                    ),
                                    ft.IconButton(
                                        ft.Icons.DELETE,
                                        tooltip="Delete",
                                        on_click=lambda e, l=loan: delete_loan(l),
                                        icon_size=20,
                                        icon_color=ft.Colors.RED_400,
                                        style=ft.ButtonStyle(padding=ft.padding.all(4)),
                                    ),
                                ],
                                spacing=0,
                                tight=True,
                            )
                        ),
                    ]
                )
            )
        
        loans_table.rows = rows
        update_pagination_info()
        safe_update()
    
    def update_pagination_info():
        """Update pagination info text"""
        total = pagination["total_items"]
        current = pagination["current_page"]
        total_pages = pagination["total_pages"]
        items_per_page = pagination["items_per_page"]
        
        start_item = (current - 1) * items_per_page + 1 if total > 0 else 0
        end_item = min(current * items_per_page, total)
        
        pagination_info_text.value = f"Page {current}/{total_pages} | Showing {start_item}-{end_item} of {total} loans"
        
        # Enable/disable pagination buttons
        prev_button.disabled = current <= 1
        next_button.disabled = current >= total_pages
    
    def safe_update():
        """Wrapper for page.update()"""
        try:
            page.update()
        except Exception:
            pass
    
    def show_loading_overlay(message="Loading..."):
        """Show loading overlay centered on screen"""
        overlay = ft.Container(
            content=ft.Column(
                controls=[
                    ft.ProgressRing(
                        width=60,
                        height=60,
                        stroke_width=4,
                        color=ft.Colors.BLUE_400,
                    ),
                    ft.Text(message, size=16, color=ft.Colors.BLUE_200),
                ],
                horizontal_alignment=ft.CrossAxisAlignment.CENTER,
                alignment=ft.MainAxisAlignment.CENTER,
                spacing=10,
            ),
            expand=True,
            alignment=ft.Alignment(0, 0),
            bgcolor="rgba(0, 0, 0, 0.7)",
        )
        page.overlay.append(overlay)
        try:
            page.update()
        except Exception:
            pass
        return overlay
    
    def hide_loading_overlay(overlay):
        """Hide loading overlay"""
        try:
            if overlay in page.overlay:
                page.overlay.remove(overlay)
                page.update()
        except Exception:
            pass
    
    def go_to_previous_page(e):
        """Go to previous page"""
        if pagination["current_page"] > 1:
            pagination["current_page"] -= 1
            filter_loans_table(search_field.value or "")
    
    def go_to_next_page(e):
        """Go to next page"""
        if pagination["total_pages"] == 0:
            pagination["total_pages"] = 1
        
        if pagination["current_page"] < pagination["total_pages"]:
            pagination["current_page"] += 1
            filter_loans_table(search_field.value or "")
    
    def open_repayment_dialog(loan):
        """Open repayment dialog and populate loan info"""
        current_loan_for_repayment["value"] = loan
        member_name = members_dict.get(loan.member_id, "Unknown")
        repayment_loan_info.value = f"Loan #{loan.id} - {member_name}: ₦{loan.amount:.2f}"
        repayment_amount_field.value = ""
        repayment_date_field.value = datetime.now().strftime("%Y-%m-%d")
        repayment_notes_field.value = ""
        repayment_dialog.open = True
        page.update()
    
    def delete_loan(loan):
        """Show confirmation dialog before deleting a loan"""
        # Determine borrower name for confirmation message
        if loan.is_member and loan.member_id:
            borrower_name = members_dict.get(loan.member_id, "Unknown")
        elif not loan.is_member and loan.non_member_id:
            from database.connection import get_non_member_by_id
            non_member = get_non_member_by_id(loan.non_member_id)
            borrower_name = non_member.name if non_member else "Unknown"
        else:
            borrower_name = "Unknown"
        
        def confirm_delete(e):
            """Actually delete the loan after confirmation"""
            delete_confirm_dialog.open = False
            page.update()
            
            try:
                success = db_delete_loan(loan.id)
                if success:
                    error_logger.info(f"Loan {loan.id} for {borrower_name} deleted successfully")
                    refresh_loans()
                    ToastNotification.show(page, f"✓ Loan #{loan.id} for {borrower_name} has been deleted.", NotificationType.SUCCESS)
                else:
                    error_logger.error(f"Failed to delete loan {loan.id}: loan not found")
                    ToastNotification.show(page, "✗ Failed to delete loan. It may have already been removed.", NotificationType.ERROR)
            except Exception as ex:
                error_logger.exception(f"Error deleting loan {loan.id}: {str(ex)}")
                ToastNotification.show(page, f"✗ Error deleting loan: {str(ex)}", NotificationType.ERROR)
            page.update()
        
        def cancel_delete(e):
            delete_confirm_dialog.open = False
            page.update()
        
        delete_confirm_dialog = ft.AlertDialog(
            modal=True,
            title=ft.Text("⚠️ Confirm Delete"),
            content=ft.Text(
                f"Are you sure you want to permanently delete this loan?\n\n"
                f"Borrower: {borrower_name}\n"
                f"Amount: ₦{loan.amount:,.2f}\n"
                f"Status: {loan.status.value}\n\n"
                f"This will also delete all repayment and refund records for this loan. "
                f"This action cannot be undone."
            ),
            actions=[
                ft.TextButton("Cancel", on_click=cancel_delete),
                ft.TextButton(
                    "Delete",
                    on_click=confirm_delete,
                    style=ft.ButtonStyle(color=ft.Colors.RED_400),
                ),
            ],
            actions_alignment=ft.MainAxisAlignment.END,
        )
        
        page.overlay.append(delete_confirm_dialog)
        delete_confirm_dialog.open = True
        page.update()
    
    def open_loan_details(loan):
        """Open loan details dialog with payment history"""
        try:
            details_dialog = create_loan_details_dialog(loan, page)
            page.overlay.append(details_dialog)
            details_dialog.open = True
            page.update()
        except Exception as err:
            error_logger.error(f"Error loading loan details: {err}")
            ToastNotification.show(page, f"✗ Error loading loan details: {str(err)}", NotificationType.ERROR)
    
    # Pagination UI Controls
    pagination_info_text = ft.Text(
        "Page 1/1 | Showing 0-0 of 0 loans",
        size=14,
        color=ft.Colors.GREY,
    )
    
    prev_button = ft.IconButton(
        ft.Icons.ARROW_BACK,
        tooltip="Previous Page",
        on_click=go_to_previous_page,
        disabled=True,
    )
    
    next_button = ft.IconButton(
        ft.Icons.ARROW_FORWARD,
        tooltip="Next Page",
        on_click=go_to_next_page,
        disabled=True,
    )
    
    pagination_row = ft.Row(
        controls=[
            prev_button,
            pagination_info_text,
            next_button,
        ],
        spacing=20,
        alignment=ft.MainAxisAlignment.CENTER,
    )
    
    # ==================== COLUMN SORTING ====================
    def on_column_sort(col_key):
        """Handle column sort — toggle direction if same column, else sort ascending"""
        def handler(e):
            if sort_state["column"] == col_key:
                sort_state["ascending"] = not sort_state["ascending"]
            else:
                sort_state["column"] = col_key
                sort_state["ascending"] = True
            filter_loans_table(search_field.value or "")
        return handler
    
    # Loans DataTable with sortable columns
    loans_table = ft.DataTable(
        columns=[
            ft.DataColumn(ft.Text("S/N", color=ft.Colors.WHITE, weight="bold")),
            ft.DataColumn(ft.Text("IPPIS NO", color=ft.Colors.WHITE, weight="bold")),
            ft.DataColumn(
                ft.Text("NAME", color=ft.Colors.WHITE, weight="bold"),
                on_sort=on_column_sort("name"),
            ),
            ft.DataColumn(
                ft.Text("AMOUNT", color=ft.Colors.WHITE, weight="bold"),
                on_sort=on_column_sort("amount"),
                numeric=True,
            ),
            ft.DataColumn(
                ft.Text("INTEREST", color=ft.Colors.WHITE, weight="bold"),
                on_sort=on_column_sort("interest"),
                numeric=True,
            ),
            ft.DataColumn(
                ft.Text("TOTAL", color=ft.Colors.WHITE, weight="bold"),
                on_sort=on_column_sort("total"),
                numeric=True,
            ),
            ft.DataColumn(ft.Text("MONTHLY REPAYMENT", color=ft.Colors.WHITE, weight="bold")),
            ft.DataColumn(
                ft.Text("AMOUNT PAID", color=ft.Colors.WHITE, weight="bold"),
                on_sort=on_column_sort("amount_paid"),
                numeric=True,
            ),
            ft.DataColumn(
                ft.Text("BALANCE", color=ft.Colors.WHITE, weight="bold"),
                on_sort=on_column_sort("balance"),
                numeric=True,
            ),
            ft.DataColumn(
                ft.Text("STATUS", color=ft.Colors.WHITE, weight="bold"),
                on_sort=on_column_sort("status"),
            ),
            ft.DataColumn(ft.Text("ACTIONS", color=ft.Colors.WHITE, weight="bold")),
        ],
        rows=[],
        bgcolor="#1a1a1a",
        divider_thickness=1,
        horizontal_lines=ft.border.BorderSide(1, "#333333"),
        sort_column_index=None,
        sort_ascending=False,
    )
    
    # ==================== SUMMARY STATS BAR ====================
    def _create_stat_card(icon, value_text, label_text, color):
        """Create a compact stat card"""
        return ft.Container(
            content=ft.Row(
                controls=[
                    ft.Icon(icon, size=22, color=color),
                    ft.Column(
                        controls=[
                            ft.Text(value_text, size=18, weight="bold", color=ft.Colors.WHITE),
                            ft.Text(label_text, size=12, color=ft.Colors.GREY_400),
                        ],
                        spacing=0,
                        tight=True,
                    ),
                ],
                spacing=8,
                vertical_alignment=ft.CrossAxisAlignment.CENTER,
            ),
            bgcolor="#252525",
            border_radius=8,
            padding=ft.padding.symmetric(horizontal=14, vertical=10),
            shadow=ft.BoxShadow(blur_radius=4, color=ft.Colors.BLACK, spread_radius=0),
        )
    
    stat_total_outstanding = ft.Text("₦0.00", size=18, weight="bold", color=ft.Colors.WHITE)
    stat_active_count = ft.Text("0", size=18, weight="bold", color=ft.Colors.WHITE)
    stat_overdue_count = ft.Text("0", size=18, weight="bold", color=ft.Colors.WHITE)
    stat_total_repaid = ft.Text("₦0.00", size=18, weight="bold", color=ft.Colors.WHITE)
    
    stats_row = ft.Row(
        controls=[
            ft.Container(
                content=ft.Row(
                    controls=[
                        ft.Icon(ft.Icons.ACCOUNT_BALANCE_WALLET, size=22, color=ft.Colors.ORANGE_400),
                        ft.Column(
                            controls=[stat_total_outstanding, ft.Text("Outstanding", size=12, color=ft.Colors.GREY_400)],
                            spacing=0, tight=True,
                        ),
                    ],
                    spacing=8, vertical_alignment=ft.CrossAxisAlignment.CENTER,
                ),
                bgcolor="#252525", border_radius=8,
                padding=ft.padding.symmetric(horizontal=14, vertical=10),
                expand=True,
            ),
            ft.Container(
                content=ft.Row(
                    controls=[
                        ft.Icon(ft.Icons.TRENDING_UP, size=22, color=ft.Colors.GREEN_400),
                        ft.Column(
                            controls=[stat_active_count, ft.Text("Active Loans", size=12, color=ft.Colors.GREY_400)],
                            spacing=0, tight=True,
                        ),
                    ],
                    spacing=8, vertical_alignment=ft.CrossAxisAlignment.CENTER,
                ),
                bgcolor="#252525", border_radius=8,
                padding=ft.padding.symmetric(horizontal=14, vertical=10),
                expand=True,
            ),
            ft.Container(
                content=ft.Row(
                    controls=[
                        ft.Icon(ft.Icons.WARNING_AMBER, size=22, color=ft.Colors.RED_400),
                        ft.Column(
                            controls=[stat_overdue_count, ft.Text("Overdue", size=12, color=ft.Colors.GREY_400)],
                            spacing=0, tight=True,
                        ),
                    ],
                    spacing=8, vertical_alignment=ft.CrossAxisAlignment.CENTER,
                ),
                bgcolor="#252525", border_radius=8,
                padding=ft.padding.symmetric(horizontal=14, vertical=10),
                expand=True,
            ),
            ft.Container(
                content=ft.Row(
                    controls=[
                        ft.Icon(ft.Icons.PAYMENTS, size=22, color=ft.Colors.BLUE_200),
                        ft.Column(
                            controls=[stat_total_repaid, ft.Text("Total Repaid", size=12, color=ft.Colors.GREY_400)],
                            spacing=0, tight=True,
                        ),
                    ],
                    spacing=8, vertical_alignment=ft.CrossAxisAlignment.CENTER,
                ),
                bgcolor="#252525", border_radius=8,
                padding=ft.padding.symmetric(horizontal=14, vertical=10),
                expand=True,
            ),
        ],
        spacing=12,
    )
    
    def update_summary_stats():
        """Recalculate and update summary stat cards"""
        now = datetime.now()
        total_outstanding = 0.0
        active_count = 0
        overdue_count = 0
        total_repaid = 0.0
        
        for loan in loans_list:
            total_repaid += loan.amount_repaid
            if loan.status != LoanStatus.PAID:
                balance = get_loan_balance_amount(loan)
                total_outstanding += max(0, balance)
            if loan.status == LoanStatus.ACTIVE:
                active_count += 1
                if loan.end_date and loan.end_date < now:
                    overdue_count += 1
        
        stat_total_outstanding.value = f"₦{total_outstanding:,.2f}"
        stat_active_count.value = str(active_count)
        stat_overdue_count.value = str(overdue_count)
        stat_total_repaid.value = f"₦{total_repaid:,.2f}"
    
    # ==================== STATUS FILTER TABS ====================
    def on_filter_tab_change(e):
        """Handle filter tab selection via SegmentedButton"""
        selected = e.control.selected
        if selected:
            active_status_filter["value"] = selected[0]
        else:
            active_status_filter["value"] = "All"
            e.control.selected = ["All"]
        pagination["current_page"] = 1
        filter_loans_table(search_field.value or "")
    
    # Segment references for dynamic label updates
    seg_all = ft.Segment("All", label=ft.Text("All"))
    seg_active = ft.Segment("Active", label=ft.Text("Active"))
    seg_refund = ft.Segment("Refund", label=ft.Text("Refund"))
    seg_paid = ft.Segment("Paid", label=ft.Text("Paid"))
    seg_overdue = ft.Segment("Overdue", label=ft.Text("Overdue"))
    
    filter_tabs = ft.SegmentedButton(
        selected=["All"],
        allow_empty_selection=False,
        allow_multiple_selection=False,
        on_change=on_filter_tab_change,
        segments=[seg_all, seg_active, seg_refund, seg_paid, seg_overdue],
        show_selected_icon=False,
    )
    
    def update_filter_tab_counts():
        """Update segment labels with counts. Overdue is separate from Active."""
        now = datetime.now()
        counts = {"All": 0, "Active": 0, "Refund": 0, "Paid": 0, "Overdue": 0}
        for loan in loans_list:
            counts["All"] += 1
            # Check for pending refunds — takes priority
            loan_refunds = get_refunds_by_loan(loan.id)
            has_refund = any(r.status == "PENDING" for r in loan_refunds)
            
            if has_refund:
                counts["Refund"] += 1
            elif loan.status == LoanStatus.ACTIVE and loan.end_date and loan.end_date < now:
                counts["Overdue"] += 1
            elif loan.status == LoanStatus.ACTIVE:
                counts["Active"] += 1
            elif loan.status == LoanStatus.PAID:
                counts["Paid"] += 1
        
        seg_all.label = ft.Text(f"All ({counts['All']})", size=14)
        seg_active.label = ft.Text(f"Active ({counts['Active']})", size=14)
        seg_refund.label = ft.Text(f"Refund ({counts['Refund']})", size=14)
        seg_paid.label = ft.Text(f"Paid ({counts['Paid']})", size=14)
        seg_overdue.label = ft.Text(f"Overdue ({counts['Overdue']})", size=14)
    
    # Initialize stats and tab counts
    update_summary_stats()
    update_filter_tab_counts()
    update_loans_table()
    
    # Add button
    def open_loan_dialog(e):
        loan_dialog.open = True
        page.update()
    
    add_loan_button = ft.ElevatedButton(
        "Add New Loan",
        icon=ft.Icons.ADD,
        on_click=open_loan_dialog,
    )
    
    # Top-up loan button
    def open_topup_dialog(e):
        topup_dialog_obj, close_topup, refresh_list = create_top_up_loan_dialog(page)
        page.overlay.append(topup_dialog_obj)
        topup_dialog_obj.open = True
        topup_dialog_obj.on_dismiss = lambda e: refresh_loans()
        page.update()
    
    topup_loan_button = ft.ElevatedButton(
        "Top-up Loan",
        icon=ft.Icons.ADD_CIRCLE,
        on_click=open_topup_dialog,
    )
    
    # Refresh button
    refresh_button = ft.IconButton(
        ft.Icons.REFRESH,
        tooltip="Refresh",
        on_click=lambda e: refresh_loans(),
    )
    
    # Export button
    export_button = ft.PopupMenuButton(
        icon=ft.Icons.DOWNLOAD,
        tooltip="Export current view",
        items=[
            ft.PopupMenuItem(content=ft.Text("Export as CSV"), icon=ft.Icons.DESCRIPTION, on_click=export_loans_csv),
            ft.PopupMenuItem(content=ft.Text("Export as Excel"), icon=ft.Icons.TABLE_CHART, on_click=export_loans_excel),
        ],
    )

    # Top controls
    top_row = ft.Row(
        controls=[
            search_field,
            add_loan_button,
            topup_loan_button,
            refresh_button,
            export_button,
        ],
        spacing=10,
    )
    
    # Table wrapper with horizontal + vertical scrolling
    table_wrapper = ft.Container(
        content=ft.Column(
            controls=[
                ft.Row(
                    controls=[loans_table],
                    scroll=ft.ScrollMode.AUTO,
                ),
                ft.Divider(height=10),
                pagination_row,
            ],
            scroll=ft.ScrollMode.AUTO,
        ),
        expand=True,
        bgcolor="#2a2a2a",
        border_radius=10,
        padding=10,
    )
    
    # Main content
    padding = get_responsive_padding(page)
    
    content = ft.Container(
        content=ft.Column(
            controls=[
                ft.Text("Loan Management", size=get_responsive_font_size(page, 20), weight="bold", color=ft.Colors.BLUE_200),
                stats_row,
                top_row,
                filter_tabs,
                table_wrapper,
            ],
            spacing=12,
            expand=True,
        ),
        padding=padding,
        bgcolor="#1a1a1a",
        expand=True,
    )
    
    # Create sidebar overlay
    sidebar_wrapper, backdrop, sidebar_visible, toggle_sidebar, close_sidebar = create_sidebar_overlay(page)
    
    # Create burger menu button
    burger_button = create_burger_menu(toggle_sidebar)
    
    # Create app bar with burger menu
    app_bar = ft.AppBar(
        title=ft.Text("Loans Management", size=22, weight="bold", color=ft.Colors.WHITE),
        bgcolor=ft.Colors.BLUE_900,
        leading=burger_button,
        actions=[
            ft.IconButton(
                ft.Icons.LOGOUT,
                tooltip="Logout",
                on_click=lambda _: page.go("/login"),
            )
        ],
    )
    
    page.overlay.append(loan_dialog)
    page.overlay.append(repayment_dialog)
    
    return ft.View(
        route="/loans",
        controls=[
            app_bar,
            ft.Stack(
                controls=[
                    content,
                    backdrop,
                    sidebar_wrapper,
                ],
                expand=True,
            ),
        ],
    )

